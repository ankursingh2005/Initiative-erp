from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form, Query, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, Response, RedirectResponse, StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text, func, or_
from sqlalchemy.exc import IntegrityError
from typing import List, Optional
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, timedelta, timezone
from io import BytesIO, StringIO
from collections import defaultdict
import csv
import json
import os
import re
import difflib
import importlib
from urllib.error import HTTPError, URLError
from urllib.request import Request as URLRequest, urlopen
from uuid import uuid4
import base64
import math
import calendar
import secrets
import hashlib

INDIA_TZ = timezone(timedelta(hours=5, minutes=30))


def india_datetime(value: datetime) -> datetime:
    """Normalize an incoming timestamp to naive India Standard Time."""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(INDIA_TZ).replace(tzinfo=None)


def india_today() -> date:
    return datetime.now(INDIA_TZ).date()


def verified_attendance_time(_device_time: datetime) -> datetime:
    """Return authoritative server IST without trusting the device clock.

    Comparing the two clocks caused valid attendance to be rejected when a
    browser or hosting instance had clock drift. Ignoring the submitted value
    is both more reliable and more secure: changing a phone's date or time can
    no longer backdate or forward-date the saved attendance record.
    """
    return datetime.now(INDIA_TZ).replace(tzinfo=None)
import smtplib
from email.mime.text import MIMEText
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook

load_dotenv()  # reads a local .env file (if present) into os.environ before
                # anything below calls os.getenv() - e.g. SMTP_*, SECRET_KEY.

import models
import schemas
import scheme_engine
import auth
from database import engine, get_db, Base, SessionLocal

# Creates all tables in the database if they don't already exist
Base.metadata.create_all(bind=engine)


def ensure_column(table_name: str, column_name: str, column_def: str):
    inspector = inspect(engine)
    if table_name not in inspector.get_table_names():
        return
    existing_columns = {col["name"] for col in inspector.get_columns(table_name)}
    if column_name in existing_columns:
        return
    with engine.begin() as conn:
        conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}"))


def ensure_username_not_unique():
    """One-time repair for databases created before username was made
    non-unique (see models.py User.username). ALTER TABLE ADD COLUMN
    can't drop a constraint, so on SQLite we rebuild the table without
    it; on Postgres the constraint just needs dropping by name. Safe to
    run on every startup - it's a no-op once the fix has been applied."""
    inspector = inspect(engine)
    if "users" not in inspector.get_table_names():
        return

    if engine.dialect.name == "sqlite":
        with engine.begin() as conn:
            index_list = conn.execute(text("PRAGMA index_list(users)")).fetchall()
            has_unique_username = False
            for idx in index_list:
                idx_name, is_unique = idx[1], idx[2]
                if not is_unique:
                    continue
                cols = {row[2] for row in conn.execute(text(f"PRAGMA index_info({idx_name})")).fetchall()}
                if cols == {"username"}:
                    has_unique_username = True
                    break
            if not has_unique_username:
                return

            conn.execute(text("ALTER TABLE users RENAME TO users_pre_username_fix"))
            models.User.__table__.create(bind=conn)
            columns = ", ".join(c.name for c in models.User.__table__.columns)
            conn.execute(text(
                f"INSERT INTO users ({columns}) SELECT {columns} FROM users_pre_username_fix"
            ))
            conn.execute(text("DROP TABLE users_pre_username_fix"))
    else:
        # Postgres: SQLAlchemy's default constraint name for a
        # single-column UNIQUE is "<table>_<column>_key".
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE users DROP CONSTRAINT IF EXISTS users_username_key"))


def ensure_database_schema():
    ensure_username_not_unique()
    ensure_column("stores", "code", "VARCHAR(20)")
    ensure_column("stores", "city", "VARCHAR(100)")
    ensure_column("stores", "status", "VARCHAR(20)")
    ensure_column("stores", "latitude", "FLOAT")
    ensure_column("stores", "longitude", "FLOAT")
    ensure_column("stores", "geofence_radius_m", "FLOAT")
    ensure_column("attendance_records", "second_punch_at", "DATETIME")
    ensure_column("attendance_records", "second_punch_selfie", "TEXT")

    ensure_column("categories", "code", "VARCHAR(10)")
    ensure_column("sub_categories", "category_id", "INTEGER")
    ensure_column("sub_categories", "name", "VARCHAR(100)")
    ensure_column("brands", "subcategory_id", "INTEGER")
    ensure_column("brands", "is_seeded_default", "BOOLEAN DEFAULT FALSE")
    ensure_column("products", "brand_id", "INTEGER")
    ensure_column("products", "name", "VARCHAR(150)")
    ensure_column("variants", "product_id", "INTEGER")
    ensure_column("variants", "name", "VARCHAR(150)")

    ensure_column("sales", "invoice_date", "DATE")
    ensure_column("sales", "subcategory_id", "INTEGER")
    ensure_column("sales", "product_id", "INTEGER")
    ensure_column("sales", "variant_id", "INTEGER")
    ensure_column("sales", "imei", "VARCHAR(100)")
    ensure_column("sales", "serial_no", "VARCHAR(100)")
    ensure_column("sales", "model_no", "VARCHAR(100)")
    ensure_column("sales", "customer_name", "VARCHAR(150)")
    ensure_column("sales", "gst", "FLOAT")
    ensure_column("sales", "sale_value_exact", "VARCHAR(40)")
    ensure_column("sales", "schemes", "VARCHAR(50)")
    ensure_column("sales", "schemes_other", "VARCHAR(255)")
    ensure_column("sales", "scheme_match", "VARCHAR(20)")
    ensure_column("sales", "scheme_match_other", "VARCHAR(255)")
    ensure_column("sales", "scheme_amount", "FLOAT")
    ensure_column("sales", "scheme_amount_exact", "VARCHAR(40)")
    ensure_column("sales", "claim_status", "VARCHAR(50)")
    ensure_column("sales", "claim_status_other", "VARCHAR(255)")
    ensure_column("sales", "claim_overall_status", "VARCHAR(50)")
    ensure_column("sales", "settled_date", "DATE")
    ensure_column("sales", "sales_executive", "VARCHAR(150)")
    ensure_column("sales", "upi_scheme_amount", "FLOAT")
    ensure_column("sales", "upi_scheme_amount_exact", "VARCHAR(40)")
    ensure_column("sales", "upi_claim_status", "VARCHAR(20)")
    ensure_column("sales", "backend_scheme_amount", "FLOAT")
    ensure_column("sales", "backend_scheme_amount_exact", "VARCHAR(40)")
    ensure_column("sales", "backend_claim_type", "VARCHAR(30)")
    ensure_column("sales", "backend_claim_status", "VARCHAR(20)")

    ensure_column("schemes", "brand_id", "INTEGER")
    ensure_column("schemes", "category_id", "INTEGER")
    ensure_column("schemes", "subcategory_id", "INTEGER")
    ensure_column("schemes", "product_id", "INTEGER")
    ensure_column("schemes", "variant_id", "INTEGER")
    ensure_column("schemes", "offer_type", "VARCHAR(50)")

    ensure_column("purchase_orders", "supplier_address", "VARCHAR(500)")
    ensure_column("purchase_orders", "supplier_gstin", "VARCHAR(30)")
    ensure_column("purchase_orders", "exported_to_busy", "BOOLEAN DEFAULT FALSE")
    ensure_column("purchase_orders", "exported_to_busy_at", "TIMESTAMP")
    ensure_column("purchase_orders", "approved_by_user_id", "INTEGER")
    ensure_column("purchase_orders", "approved_date", "TIMESTAMP")
    ensure_column("price_list_items", "model_no", "VARCHAR(100)")
    ensure_column("price_list_items", "serial_no", "VARCHAR(100)")
    ensure_column("price_list_items", "imei", "VARCHAR(100)")
    ensure_column("analytics_sales_rows", "brand", "VARCHAR(150)")
    ensure_column("schemes", "offer_value", "FLOAT")
    ensure_column("schemes", "calculation_method", "VARCHAR(50)")
    ensure_column("schemes", "min_qty", "INTEGER")
    ensure_column("schemes", "max_qty", "INTEGER")
    ensure_column("schemes", "applicable_branch_id", "INTEGER")
    ensure_column("schemes", "applicable_customer", "VARCHAR(100)")
    ensure_column("schemes", "applicable_dealer", "VARCHAR(100)")
    ensure_column("schemes", "circular_number", "VARCHAR(50)")
    ensure_column("schemes", "remarks", "VARCHAR(255)")
    ensure_column("schemes", "reward_type_other", "VARCHAR(100)")

    ensure_column("users", "store_id", "INTEGER")
    ensure_column("users", "category_code", "VARCHAR(20)")
    ensure_column("users", "status", "VARCHAR(20)")
    ensure_column("users", "weekoff_day", "VARCHAR(10)")
    ensure_column("users", "created_date", "TIMESTAMP")
    ensure_column("users", "reset_token", "VARCHAR(100)")
    ensure_column("users", "reset_token_expires", "TIMESTAMP")
    # Correct the legacy role spelling without leaving existing accounts
    # under a role that can no longer be selected during signup.
    with engine.begin() as conn:
        conn.execute(text("UPDATE users SET role = 'Loader' WHERE role = 'Loder'"))

    ensure_column("claim_headers", "claim_no", "VARCHAR(50)")
    ensure_column("claim_headers", "brand_id", "INTEGER")
    ensure_column("claim_headers", "invoice_no", "VARCHAR(50)")
    ensure_column("claim_headers", "branch_id", "INTEGER")
    ensure_column("claim_headers", "remarks", "VARCHAR(255)")
    ensure_column("claim_headers", "payment_amount", "FLOAT")
    ensure_column("claim_headers", "balance", "FLOAT")
    ensure_column("claim_headers", "created_date", "TIMESTAMP")
    ensure_column("claim_headers", "submission_date", "DATE")
    ensure_column("claim_headers", "approval_date", "DATE")
    ensure_column("claim_headers", "received_date", "DATE")

    # Lets Ageing Stock Analysis search by Model No. as well as Item
    # Details - see AGEING_HEADER_ALIASES / parse_ageing_stock_workbook.
    ensure_column("ageing_stock_items", "model_no", "VARCHAR(150)")


ensure_database_schema()


def ensure_default_branches():
    with SessionLocal() as db:
        default_branches = [
            {"name": "Alambagh", "code": "BR001", "city": "Lucknow", "status": "Active", "latitude": 26.8023316384953, "longitude": 80.89578659627016, "geofence_radius_m": 100},
            {"name": "Gomtinagar", "code": "BR002", "city": "Lucknow", "status": "Active", "latitude": 26.850130023596396, "longitude": 81.00713531584118, "geofence_radius_m": 100},
            {"name": "Ashiyana", "code": "BR003", "city": "Lucknow", "status": "Active", "latitude": 26.795911774284612, "longitude": 80.92089772123978, "geofence_radius_m": 100},
            {"name": "Hazratganj", "code": "BR004", "city": "Lucknow", "status": "Active", "latitude": 26.84924030483742, "longitude": 80.94773860240677, "geofence_radius_m": 100},
            {"name": "Vikas Nagar", "code": "BR005", "city": "Lucknow", "status": "Active", "latitude": 26.90188397262733, "longitude": 80.95513690261241, "geofence_radius_m": 100},
            {"name": "Warehouse", "code": "MWH", "city": "Lucknow", "status": "Active", "latitude": 26.779149508725354, "longitude": 80.88470873166187, "geofence_radius_m": 100},
            {"name": "Head Office", "code": "HO", "city": "Lucknow", "status": "Active", "latitude": 26.84904218332385, "longitude": 80.94789353821966, "geofence_radius_m": 100},
        ]
        if db.query(models.Store).count() == 0:
            for branch in default_branches:
                db.add(models.Store(**branch))
        else:
            for branch in default_branches:
                store = db.query(models.Store).filter(models.Store.code == branch["code"]).first()
                if not store:
                    db.add(models.Store(**branch))
                elif branch["code"] == "BR003":
                    # Keep the deployed Ashiyana geofence synchronized with
                    # its verified location instead of retaining stale GPS.
                    store.latitude = branch["latitude"]
                    store.longitude = branch["longitude"]
                    store.geofence_radius_m = store.geofence_radius_m or 100
                elif branch["latitude"] is not None and (not store.latitude or not store.longitude):
                    store.latitude = branch["latitude"]
                    store.longitude = branch["longitude"]
                    store.geofence_radius_m = store.geofence_radius_m or 100

        # One-time/idempotent outlet reassignment requested for the deployed
        # Hazratganj accounts. Running this at startup ensures existing users
        # are moved after deployment as soon as Head Office has been seeded.
        head_office = db.query(models.Store).filter(models.Store.code == "HO").first()
        if head_office:
            head_office_emails = {
                "singhankur7521@gmail.com",
                "ashvriyamishra6671@gmail.com",
                "ashwanis0025@gmail.com",
                "jatinkumarsaini771@gmail.com",
                "initiative.lucknow@gmail.com",
                "raju.singh26101@gmail.com",
                "srivastava.op@gmail.com",
                "pooja.tripathiji@gmail.com",
                "shreyashah19479@gmail.com",
                "shubhampanday959@gmail.com",
                "rajputswarnima8@gmail.com",
                "tabish.syed24@gmail.com",
            }
            db.query(models.User).filter(
                func.lower(models.User.email).in_(head_office_emails)
            ).update({models.User.store_id: head_office.id}, synchronize_session=False)
        db.commit()


def ensure_default_master_data():
    with SessionLocal() as db:
        default_categories = [
            {"code": "HA", "name": "Home Appliances"},
            {"code": "HE", "name": "Home Entertainment"},
            {"code": "MH", "name": "Mobiles / Handset"},
            {"code": "IT", "name": "Information Technology"},
            {"code": "ASC", "name": "Accessories"},
            {"code": "OTH", "name": "Others"},
        ]
        allowed_codes = {item["code"] for item in default_categories}
        desired_names = {item["code"]: item["name"] for item in default_categories}

        existing_categories = db.query(models.Category).all()
        for category in existing_categories:
            category_code = (category.code or "").upper()
            if category_code in allowed_codes:
                if category.name != desired_names[category_code]:
                    category.name = desired_names[category_code]
            else:
                db.delete(category)

        for item in default_categories:
            existing = db.query(models.Category).filter(models.Category.code == item["code"]).first()
            if existing:
                existing.name = item["name"]
            else:
                db.add(models.Category(**item))

        db.commit()

        categories_by_code = {
            c.code: c
            for c in db.query(models.Category).filter(models.Category.code.in_(allowed_codes)).all()
        }

        category_subcategories = {
            "HA": [
                "Air Conditioner",
                "Air Purifier",
                "Cooler",
                "Dish Washer",
                "Geyser",
                "Fan",
                "iron",
                "Heat Convector",
                "Oil Filled Radiator (OFR)",
                "Refrigerator",
                "Vacuum Cleaner",
                "Washing Machine",
                "Water Heater",
                "Water Purifier",
                "Microwave Oven",
                "Kitchen Chimney",
                "Cooktop",
                "Induction Cooker",
                "Mixer Grinder",
                "Juicer",
                "Electric Kettle",
                "Rice Cooker",
                "Room Heater",
                "Deep Freezer",
            ],
            "HE": [
                "LED TV",
                "Projector",
                "Home Theatre",
                "Soundbar",
                "Speaker",
            ],
            "IT": [
                "Laptop",
                "Desktop",
            ],
            "ASC": [
                "Mobile Charger",
                "USB Cable",
                "Power Bank",
                "Earbuds",
                "Earphones",
                "Headphones",
                "Neckband",
                "Bluetooth Speaker",
                "Laptop Bag",
                "Laptop Adapter",
                "HDMI Cable",
                "USB Drive",
                "Memory Card",
                "Mouse",
                "Keyboard",
                "Extension Board",
                "TV Wall Mount",
                "AC Stabilizer",
                "Remote",
                "Battery",
            ],
            "OTH": [
                "Digital Camera",
                "DSLR",
                "Mirrorless Camera",
                "Camera Lens",
                "Drone",
                "Gaming Console",
                "Fire TV Stick",
                "Google Chromecast",
                "Amazon Echo",
            ],
            "MH": [
                "Apple",
                "Samsung",
                "Vivo",
                "iQOO",
                "Realme",
                "Oppo",
                "Motorola",
                "Nothing",
                "Google Pixel",
            ],
        }

        for category_code, names in category_subcategories.items():
            category = categories_by_code.get(category_code)
            if not category:
                continue
            existing_subcategories = (
                db.query(models.SubCategory)
                .filter(models.SubCategory.category_id == category.id)
                .all()
            )
            expected_names = set(names)
            existing_names = {sub.name for sub in existing_subcategories}

            for sub in existing_subcategories:
                if sub.name not in expected_names:
                    db.delete(sub)

            for name in names:
                if name not in existing_names:
                    db.add(models.SubCategory(category_id=category.id, name=name))

        db.commit()

        def get_or_create_brand(name: str, fallback_subcategory_id):
            """Brand names are unique in this database, so never insert a
            second row for a name that already exists — reuse the existing
            brand instead. Its subcategory_id (used elsewhere for schemes/
            products) is only set on first creation and never overwritten,
            so seeding one category never silently reassigns a brand that
            already belongs to a different one."""
            existing = db.query(models.Brand).filter(models.Brand.name.ilike(name)).first()
            if existing:
                return existing
            brand = models.Brand(
                name=name,
                subcategory_id=fallback_subcategory_id,
                is_seeded_default=True,
            )
            db.add(brand)
            db.flush()
            return brand

        def make_brand_visible_in_category(brand: "models.Brand", category: "models.Category"):
            exists = (
                db.query(models.BrandCategoryVisibility)
                .filter(
                    models.BrandCategoryVisibility.brand_id == brand.id,
                    models.BrandCategoryVisibility.category_id == category.id,
                )
                .first()
            )
            if not exists:
                db.add(models.BrandCategoryVisibility(brand_id=brand.id, category_id=category.id))

        # Brand list per category for IDS Price System (and the rest of the
        # ERP, since Brand is a single shared table). Order matters here:
        # HA has no legacy brands to clean up, HE is processed before MH so
        # brands shared between them (e.g. "Mi") already have HE's visibility
        # recorded by the time MH's stale entries are swept, and IT is last.
        CATEGORY_BRAND_LISTS = {
            "HA": [
                "Aisen", "Amaze", "AO Smith", "Bajaj", "Bluestar", "Bosch", "Carrier",
                "Champion", "Daikin", "Dyson", "Elice", "Eurek", "Faber", "General",
                "Goderaj", "Haier", "Havells", "Hitachi", "IFB", "Kent", "LG",
                "Liebherr", "Liugurd", "Lloyd", "Luminous", "Marc", "Microtech",
                "Mitsubishi", "Philips", "Sharp", "Sunflame", "TCL", "Vgaurd",
                "Voltas", "Xiaomi",
            ],
            "HE": [
                "Aisen", "Haier", "Hisense", "LG", "Lloyd", "Mi", "OnePlus",
                "Panasonic", "Samsung", "Sony", "TCL",
            ],
            "MH": [
                "Apple", "Google", "iQOO", "Motorola", "Nothing", "OnePlus", "Oppo",
                "Realme", "Readmi", "Samsung", "Vivo", "Philips", "Lenovo Tablet",
            ],
            "IT": [
                "Apple iMac", "Apple iPad", "Apple MacBook", "Dell", "HP", "Lenovo",
            ],
        }

        def brand_reference_count(brand_id: int) -> int:
            """Rows anywhere in the ERP still pointing at this brand - if any
            exist, the brand is real transactional/master data and is only
            unassigned from a category below, never deleted outright."""
            return (
                db.query(models.Product).filter(models.Product.brand_id == brand_id).count()
                + db.query(models.Sale).filter(models.Sale.brand_id == brand_id).count()
                + db.query(models.ClaimHeader).filter(models.ClaimHeader.brand_id == brand_id).count()
                + db.query(models.UserBrand).filter(models.UserBrand.brand_id == brand_id).count()
                + db.query(models.PriceListItem).filter(models.PriceListItem.brand_id == brand_id).count()
                + db.query(models.Scheme).filter(models.Scheme.brand_id == brand_id).count()
            )

        for category_code, brand_names in CATEGORY_BRAND_LISTS.items():
            category = categories_by_code.get(category_code)
            if not category:
                continue

            desired_names_lower = {name.strip().lower() for name in brand_names}
            category_subcategory_ids = [
                row[0] for row in db.query(models.SubCategory.id)
                .filter(models.SubCategory.category_id == category.id)
                .all()
            ]

            # Every brand currently attached to this category, either as its
            # primary subcategory home or via the visibility table.
            existing_brand_ids = set()
            if category_subcategory_ids:
                existing_brand_ids.update(
                    b.id for b in db.query(models.Brand)
                    .filter(models.Brand.subcategory_id.in_(category_subcategory_ids))
                    .all()
                )
            existing_brand_ids.update(
                bcv.brand_id for bcv in db.query(models.BrandCategoryVisibility)
                .filter(models.BrandCategoryVisibility.category_id == category.id)
                .all()
            )

            # Remove every brand no longer in this category's list - the
            # "remove previous brands" part of the reseed. A brand still used
            # elsewhere (or still visible in another category) is unassigned
            # here rather than deleted. IMPORTANT: this cleanup only ever
            # touches brands this same reseed function created
            # (is_seeded_default=True). A brand a real user added through
            # the UI/API is never swept here, even if it happens to share a
            # subcategory or category-visibility row with a seeded one -
            # otherwise every restart could silently delete brands a user
            # just added and hadn't attached any sales/price-list data to
            # yet.
            for brand_id in existing_brand_ids:
                brand = db.query(models.Brand).filter(models.Brand.id == brand_id).first()
                if not brand or not brand.is_seeded_default or brand.name.strip().lower() in desired_names_lower:
                    continue
                db.query(models.BrandCategoryVisibility).filter(
                    models.BrandCategoryVisibility.brand_id == brand.id,
                    models.BrandCategoryVisibility.category_id == category.id,
                ).delete()
                if brand.subcategory_id in category_subcategory_ids:
                    brand.subcategory_id = None
                db.flush()
                if brand_reference_count(brand.id) == 0 and brand.subcategory_id is None:
                    remaining_visibility = (
                        db.query(models.BrandCategoryVisibility)
                        .filter(models.BrandCategoryVisibility.brand_id == brand.id)
                        .count()
                    )
                    if not remaining_visibility:
                        db.delete(brand)
            db.commit()

            subcategories_in_category = (
                db.query(models.SubCategory)
                .filter(models.SubCategory.category_id == category.id)
                .all()
            )
            for index, brand_name in enumerate(brand_names):
                fallback_subcategory_id = (
                    subcategories_in_category[index % len(subcategories_in_category)].id
                    if subcategories_in_category else None
                )
                brand = get_or_create_brand(brand_name, fallback_subcategory_id)
                make_brand_visible_in_category(brand, category)
            db.commit()


ensure_default_branches()
ensure_default_master_data()

app = FastAPI(title="IDSPL Scheme Management ERP")


@app.middleware("http")
async def brand_promotor_limited_access(request: Request, call_next):
    """Brand Promotor accounts may use Home and Attendance only."""
    token = request.headers.get("Authorization", "")
    if token.startswith("Bearer "):
        try:
            payload = auth.jwt.decode(
                token[7:], auth.SECRET_KEY, algorithms=[auth.ALGORITHM]
            )
            if payload.get("role") == "BrandPartner":
                allowed = {
                    "/attendance", "/attendance.html", "/home", "/home.html",
                    "/api/me", "/stores", "/auth/login", "/openapi.json", "/docs",
                }
                static_html = request.url.path.startswith("/static/") and request.url.path.endswith(".html")
                attendance_request = request.url.path == "/api/attendance" or request.url.path.startswith("/api/attendance/")
                if request.url.path not in allowed and not attendance_request and not request.url.path.startswith("/static/"):
                    return JSONResponse(
                        status_code=403,
                        content={"detail": "Brand Promotor access is limited to Home and Attendance."},
                    )
                if static_html and not request.url.path.endswith("/attendance.html"):
                    return JSONResponse(
                        status_code=403,
                        content={"detail": "Brand Promotor access is limited to Home and Attendance."},
                    )
        except Exception:
            pass
    response = await call_next(request)
    return response


@app.exception_handler(Exception)
async def handle_unexpected_error(request: Request, exc: Exception):
    """Any error we didn't explicitly raise as an HTTPException still comes
    back as JSON (with a real message) instead of a raw text/HTML 500 page.
    A non-JSON error body is what makes the frontend show a generic
    'Request failed' / 'Failed to fetch' with no useful detail."""
    return JSONResponse(status_code=500, content={"detail": f"{type(exc).__name__}: {exc}"})


# Serves the login.html / signup.html / dashboard.html pages from the
# "static" folder sitting next to this file.
app.mount("/static", StaticFiles(directory="static"), name="static")

VALID_ROLES = ["Admin", "Owner", "HR", "CategoryManager", "BrandManager", "BrandPartner", "SupportingStaff", "ServiceManager", "ServiceHead", "SalesExecutive", "AsstSalesManager", "LogisticManager", "Supervisor", "Assistant", "Loader", "ACTechnicianA", "ACTechnicianB", "Accounts", "MISExecutive", "ITEngineer", "CustomerCare", "Employee", "Cashier", "Other"]


def normalize_category_code(raw_value: Optional[str]) -> Optional[str]:
    value = (raw_value or "").strip().upper()
    if not value:
      return None
    mapping = {
        "HA": "HA",
        "HE": "HE",
        "IT": "IT",
        "MOBILE": "MH",
        "MH": "MH",
        "OTHER": "OTH",
        "OTH": "OTH",
    }
    if value in mapping:
        return mapping[value]
    raise HTTPException(status_code=400, detail="category_code must be one of: HA, HE, IT, MOBILE, OTHER")


CATEGORY_MANAGER_CODES = {"HA", "HE", "IT", "MH"}


def normalize_category_codes(raw_values) -> list:
    values = raw_values if isinstance(raw_values, (list, tuple, set)) else str(raw_values or "").split(",")
    normalized = []
    for raw in values:
        value = str(raw or "").strip().upper()
        if not value:
            continue
        if value == "ALL":
            return sorted(CATEGORY_MANAGER_CODES)
        code = normalize_category_code(value)
        if code and code not in normalized:
            normalized.append(code)
    return normalized


def category_codes_for_user(user: models.User) -> set:
    return set(normalize_category_codes(getattr(user, "category_code", None)))


def normalize_reward_type(raw_value: str) -> str:
    value = (raw_value or "").strip().lower()
    if value in {"fixed", "fixed amount", "amount"}:
        return "Fixed"
    if value in {"%", "percentage", "percent"}:
        return "Percentage"
    if value in {"target based", "target", "slab"}:
        return "Slab"
    if value in {"other"}:
        # Custom scheme types still need a concrete calculation behind them
        # for auto claim generation; treat "Other" as a flat/fixed amount.
        # The user's own label is kept separately in reward_type_other.
        return "Fixed"
    raise HTTPException(status_code=400, detail="reward_type must be one of: Fixed Amount, Target Based, %, Slab, Other")


def normalize_offer_type(raw_value: str) -> str:
    value = (raw_value or "").strip().upper()
    mapping = {
        "BACKEND": "Backend",
        "BACKEND SUPPORT": "Backend",
        "UPI": "UPI",
        "UPI OFFER": "UPI",
        "CARD": "CARD",
        "CARD OFFER": "CARD",
        "FESTIVAL OFFER": "FESTIVAL",
        "MONTHLY OFFER": "MONTHLY",
        "OTHER": "OTH",
        "OTH": "OTH",
    }
    if value in mapping:
        return mapping[value]
    raise HTTPException(status_code=400, detail="offer_type must be one of: Backend Support, UPI Offer, Card Offer, Festival Offer, Monthly Offer, Other")


def normalize_header_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


HEADER_ALIASES = {
    "sale_date": {"date", "invoicedate", "saledate", "billdate"},
    "vch_no": {"vchno", "voucher", "voucherno", "invoiceno", "billno"},
    "account": {"account", "customer", "accountname", "party"},
    "item": {"item", "product", "itemname", "description"},
    "qty": {"qty", "quantity"},
    "unit": {"unit", "uom"},
    "sales_amt": {"salesamt", "salesamount", "salevalue", "amount", "invoicevalue"},
    "cost_amt": {"costamt", "costamount", "cost"},
    "profit_loss": {"profitloss", "grossprofit"},
    "profit_percent": {"profit", "profitpercent", "profitpercentage", "marginpercent", "gppercent"},
}


def canonical_column(normalized_header: str) -> Optional[str]:
    if normalized_header == "profit":
        return "profit_percent"
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized_header in aliases:
            return canonical
    return None


def find_header_row_index(table_rows: List[List]) -> int:
    """Find the row that most likely contains expected sales headers."""
    scan_limit = min(len(table_rows), 15)
    best_index = 0
    best_score = -1

    for row_index in range(scan_limit):
        row = table_rows[row_index] or []
        normalized = [normalize_header_name(cell) for cell in row]
        canonicals = {canonical_column(name) for name in normalized if canonical_column(name)}
        canonicals.discard(None)

        # Weighted score: prefer rows that include core columns.
        score = len(canonicals)
        if "sale_date" in canonicals:
            score += 2
        if "sales_amt" in canonicals:
            score += 2
        if "qty" in canonicals:
            score += 1

        if score > best_score:
            best_score = score
            best_index = row_index

    return best_index


def parse_date_value(raw_value) -> date:
    if isinstance(raw_value, datetime):
        return raw_value.date()

    if isinstance(raw_value, date):
        return raw_value

    # Excel serial date numbers (common in xlsx uploads)
    if isinstance(raw_value, (int, float)):
        try:
            excel_datetime_utils = importlib.import_module("openpyxl.utils.datetime")
            excel_date = excel_datetime_utils.from_excel(raw_value)
            if isinstance(excel_date, datetime):
                return excel_date.date()
            if isinstance(excel_date, date):
                return excel_date
        except Exception:
            pass

    text_value = str(raw_value or "").strip()
    if not text_value:
        raise ValueError("Date is empty")

    for date_fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text_value, date_fmt).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text_value).date()
    except ValueError as exc:
        raise ValueError(f"Unsupported date value: {text_value}") from exc


def parse_float_value(raw_value, fallback: float = 0.0) -> float:
    text_value = str(raw_value or "").strip()
    if not text_value:
        return fallback

    is_negative_bracket = text_value.startswith("(") and text_value.endswith(")")
    cleaned = text_value.strip("()")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    cleaned = cleaned.replace(",", "").replace("%", "").strip()
    if cleaned in {"-", "--"}:
        return fallback

    value = float(cleaned)
    if is_negative_bracket:
        value *= -1
    return value


def parse_tabular_rows(file_ext: str, content: bytes) -> List[dict]:
    parsed_rows: List[dict] = []

    if file_ext in {".xlsx", ".xls"}:
        try:
            openpyxl_module = importlib.import_module("openpyxl")
            load_workbook = openpyxl_module.load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="Excel upload requires openpyxl package. Install: pip install openpyxl") from exc

        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        worksheet = workbook.active
        raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not raw_rows:
            return []

        header_index = find_header_row_index(raw_rows)
        headers = [normalize_header_name(cell) for cell in raw_rows[header_index]]

        for row in raw_rows[header_index + 1:]:
            row_dict = {}
            for idx, header in enumerate(headers):
                canonical = canonical_column(header)
                if canonical:
                    row_dict[canonical] = row[idx] if idx < len(row) else None
            if any(str(value or "").strip() for value in row_dict.values()):
                parsed_rows.append(row_dict)
        return parsed_rows

    if file_ext == ".csv":
        decoded = content.decode("utf-8-sig", errors="replace")
        raw_lines = [line for line in decoded.splitlines() if line.strip()]
        if not raw_lines:
            return []

        preview_rows = [next(csv.reader([line])) for line in raw_lines[:15]]
        header_index = find_header_row_index(preview_rows)

        data_lines = raw_lines[header_index:]
        reader = csv.DictReader(data_lines)
        for input_row in reader:
            row_dict = {}
            for key, value in input_row.items():
                canonical = canonical_column(normalize_header_name(key))
                if canonical:
                    row_dict[canonical] = value
            if any(str(value or "").strip() for value in row_dict.values()):
                parsed_rows.append(row_dict)
        return parsed_rows

    if file_ext == ".pdf":
        try:
            pypdf_module = importlib.import_module("pypdf")
            PdfReader = pypdf_module.PdfReader
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="PDF upload requires pypdf package. Install: pip install pypdf") from exc

        reader = PdfReader(BytesIO(content))
        text_data = "\n".join((page.extract_text() or "") for page in reader.pages)
        lines = [line.strip() for line in text_data.splitlines() if line.strip()]
        if not lines:
            return []

        header_line = None
        for line in lines:
            normalized = normalize_header_name(line)
            if "date" in normalized and "vch" in normalized and ("salesamt" in normalized or "salesamount" in normalized):
                header_line = line
                break

        if not header_line:
            return []

        header_parts = [part.strip() for part in re.split(r"\t+|\s{2,}|,", header_line) if part.strip()]
        normalized_headers = [normalize_header_name(part) for part in header_parts]

        header_index = lines.index(header_line)
        for line in lines[header_index + 1:]:
            parts = [part.strip() for part in re.split(r"\t+|\s{2,}|,", line) if part.strip()]
            if len(parts) < 4:
                continue
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                canonical = canonical_column(header)
                if canonical and idx < len(parts):
                    row_dict[canonical] = parts[idx]
            if any(str(value or "").strip() for value in row_dict.values()):
                parsed_rows.append(row_dict)
        return parsed_rows

    if file_ext in {".jpg", ".jpeg", ".png"}:
        try:
            pil_module = importlib.import_module("PIL.Image")
            Image = pil_module
            pytesseract = importlib.import_module("pytesseract")
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="Image upload requires pillow + pytesseract. Install: pip install pillow pytesseract") from exc

        image = Image.open(BytesIO(content))
        ocr_text = pytesseract.image_to_string(image)
        lines = [line.strip() for line in ocr_text.splitlines() if line.strip()]
        if not lines:
            return []

        header_line = lines[0]
        header_parts = [part.strip() for part in re.split(r"\t+|\s{2,}|,", header_line) if part.strip()]
        normalized_headers = [normalize_header_name(part) for part in header_parts]

        for line in lines[1:]:
            parts = [part.strip() for part in re.split(r"\t+|\s{2,}|,", line) if part.strip()]
            if len(parts) < 4:
                continue
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                canonical = canonical_column(header)
                if canonical and idx < len(parts):
                    row_dict[canonical] = parts[idx]
            if any(str(value or "").strip() for value in row_dict.values()):
                parsed_rows.append(row_dict)
        return parsed_rows

    raise HTTPException(status_code=400, detail="Unsupported file format. Upload Excel, CSV, PDF, JPG, JPEG, or PNG")


def build_interval_analytics(rows: List[models.IntervalSaleUpload], interval: str) -> dict:
    grouped = defaultdict(lambda: {"qty": 0.0, "sales_amt": 0.0, "cost_amt": 0.0, "profit_loss": 0.0})

    for row in rows:
        if interval == "weekly":
            iso_year, iso_week, _ = row.sale_date.isocalendar()
            group_key = f"{iso_year}-W{iso_week:02d}"
        elif interval == "monthly":
            group_key = row.sale_date.strftime("%Y-%m")
        else:
            group_key = row.sale_date.isoformat()

        bucket = grouped[group_key]
        bucket["qty"] += float(row.qty or 0)
        bucket["sales_amt"] += float(row.sales_amt or 0)
        bucket["cost_amt"] += float(row.cost_amt or 0)
        bucket["profit_loss"] += float(row.profit_loss or 0)

    points = []
    for key in sorted(grouped.keys()):
        data = grouped[key]
        sales_amt = data["sales_amt"]
        profit_percent = (data["profit_loss"] / sales_amt * 100.0) if sales_amt else 0.0
        points.append({
            "label": key,
            "qty": round(data["qty"], 2),
            "sales_amt": round(sales_amt, 2),
            "cost_amt": round(data["cost_amt"], 2),
            "profit_loss": round(data["profit_loss"], 2),
            "profit_percent": round(profit_percent, 2),
        })

    total_qty = sum(point["qty"] for point in points)
    total_sales = sum(point["sales_amt"] for point in points)
    total_cost = sum(point["cost_amt"] for point in points)
    total_profit = sum(point["profit_loss"] for point in points)
    total_profit_percent = (total_profit / total_sales * 100.0) if total_sales else 0.0

    top_items = defaultdict(lambda: {"qty": 0.0, "sales_amt": 0.0, "profit_loss": 0.0})
    for row in rows:
        item_name = (row.item or "Unknown").strip() or "Unknown"
        top_items[item_name]["qty"] += float(row.qty or 0)
        top_items[item_name]["sales_amt"] += float(row.sales_amt or 0)
        top_items[item_name]["profit_loss"] += float(row.profit_loss or 0)

    top_item_rows = sorted(
        [
            {
                "item": item,
                "qty": round(values["qty"], 2),
                "sales_amt": round(values["sales_amt"], 2),
                "profit_loss": round(values["profit_loss"], 2),
            }
            for item, values in top_items.items()
        ],
        key=lambda entry: entry["sales_amt"],
        reverse=True,
    )[:8]

    return {
        "interval": interval,
        "totals": {
            "records": len(rows),
            "qty": round(total_qty, 2),
            "sales_amt": round(total_sales, 2),
            "cost_amt": round(total_cost, 2),
            "profit_loss": round(total_profit, 2),
            "profit_percent": round(total_profit_percent, 2),
        },
        "series": points,
        "top_items": top_item_rows,
    }


# ============================================================
# SCHEME DOCUMENT -> LLM EXTRACTION
# A promoter/brand manager attaches a scheme circular (image, PDF, or
# Excel). Claude reads it and returns structured scheme fields, which are
# used to pre-fill a Draft scheme for Admin to review. Extraction never
# raises - if it fails or ANTHROPIC_API_KEY isn't set, the scheme is just
# left as a bare Draft for Admin to fill in by hand.
# ============================================================

SCHEME_EXTRACTION_MODEL = "claude-sonnet-5"


def _document_to_llm_content_block(filename: str, content_type: str, raw_bytes: bytes) -> Optional[dict]:
    """Turn an uploaded scheme document into a Claude API content block.
    Images and PDFs are sent as-is (base64) so Claude can read tables,
    stamps, and handwriting directly. Excel/CSV files are flattened to a
    plain-text cell dump first, since the Messages API has no spreadsheet
    input type."""
    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""
    b64 = base64.b64encode(raw_bytes).decode("ascii")

    if ext in {".jpg", ".jpeg", ".png", ".webp"}:
        media_type = content_type or ("image/png" if ext == ".png" else "image/jpeg")
        return {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}}

    if ext == ".pdf":
        return {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}}

    if ext in {".xlsx", ".xls", ".csv"}:
        try:
            if ext == ".csv":
                text_dump = raw_bytes.decode("utf-8-sig", errors="replace")
            else:
                openpyxl_module = importlib.import_module("openpyxl")
                workbook = openpyxl_module.load_workbook(filename=BytesIO(raw_bytes), data_only=True, read_only=True)
                lines = []
                for sheet in workbook.worksheets:
                    lines.append(f"--- Sheet: {sheet.title} ---")
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(cell) for cell in row if cell is not None]
                        if cells:
                            lines.append(" | ".join(cells))
                text_dump = "\n".join(lines)
        except Exception:
            return None
        return {"type": "text", "text": text_dump[:20000]}

    return None


def extract_scheme_from_document(db: Session, filename: str, content_type: str, raw_bytes: bytes) -> dict:
    """Calls the Claude API to read a scheme circular and return structured
    fields. Returns {"status": "Extracted"/"Failed"/"Skipped", "data": {...}, "error": str|None}."""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return {"status": "Skipped", "data": {}, "error": "ANTHROPIC_API_KEY is not configured on the server."}

    content_block = _document_to_llm_content_block(filename, content_type, raw_bytes)
    if not content_block:
        return {"status": "Failed", "data": {}, "error": "Unsupported file type for extraction."}

    brands = [{"id": b.id, "name": b.name} for b in db.query(models.Brand).all()]
    categories = [{"id": c.id, "code": c.code, "name": c.name} for c in db.query(models.Category).all()]

    instructions = (
        "You are reading a dealer/brand scheme circular (an incentive or backend "
        "scheme notice) for an electronics retail ERP. Extract the scheme terms "
        "and reply with ONLY a JSON object - no prose, no markdown fences. Schema:\n"
        "{\n"
        '  "scheme_name": string,\n'
        '  "brand_name": string or null (the brand this scheme is for),\n'
        '  "product_name": string or null (specific product/model if the scheme is product-specific),\n'
        '  "category_hint": string or null (e.g. HA, HE, IT, Mobile - only if clearly stated),\n'
        '  "start_date": "YYYY-MM-DD" or null,\n'
        '  "end_date": "YYYY-MM-DD" or null,\n'
        '  "reward_type": one of "Fixed", "Percentage", "Slab",\n'
        '  "reward_value": number (flat amount for Fixed, percent for Percentage, 0 for Slab),\n'
        '  "slabs": [{"min_quantity": number, "reward_per_unit": number}] (only for Slab, else []),\n'
        '  "min_qty": number or 0,\n'
        '  "max_qty": number or null,\n'
        '  "offer_type": one of "Backend", "UPI", "CARD", "FESTIVAL", "MONTHLY", "OTH",\n'
        '  "circular_number": string or null,\n'
        '  "remarks": string or null (any other important terms/conditions in the document)\n'
        "}\n\n"
        "If the document mentions only a month/quarter (e.g. \"August 2026 scheme\") without "
        "exact dates, set start_date to the first day and end_date to the last day of that "
        "period. If a field truly isn't in the document, use null (or 0/[] as shown above). "
        f"Known brands in this system: {json.dumps(brands)}. Known categories: {json.dumps(categories)}."
    )

    payload = json.dumps({
        "model": SCHEME_EXTRACTION_MODEL,
        "max_tokens": 1500,
        "messages": [
            {"role": "user", "content": [content_block, {"type": "text", "text": instructions}]}
        ],
    }).encode("utf-8")

    request = URLRequest(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=45) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError) as exc:
        return {"status": "Failed", "data": {}, "error": f"Claude API request failed: {exc}"}
    except Exception as exc:
        return {"status": "Failed", "data": {}, "error": f"Unexpected error calling Claude API: {exc}"}

    try:
        text_blocks = [block["text"] for block in response_data.get("content", []) if block.get("type") == "text"]
        raw_text = "\n".join(text_blocks).strip()
        raw_text = re.sub(r"^```(json)?|```$", "", raw_text, flags=re.MULTILINE).strip()
        extracted = json.loads(raw_text)
    except Exception as exc:
        return {"status": "Failed", "data": {}, "error": f"Could not parse Claude's response as JSON: {exc}"}

    return {"status": "Extracted", "data": extracted, "error": None}


def _calculate_reward_for_interval_row(scheme: models.Scheme, row: models.IntervalSaleUpload) -> float:
    """Same reward math as scheme_engine._calculate_reward, adapted for an
    IntervalSaleUpload row (Busy profitability report import) instead of a
    Sale row entered manually."""
    if scheme.reward_type == "Fixed":
        return float(scheme.reward_value)
    if scheme.reward_type == "Percentage":
        return round((row.sales_amt or 0) * (scheme.reward_value / 100), 2)
    if scheme.reward_type == "Slab":
        applicable_slabs = [s for s in scheme.slabs if (row.qty or 0) >= s.min_quantity]
        if not applicable_slabs:
            return 0
        best_slab = max(applicable_slabs, key=lambda s: s.min_quantity)
        return round(best_slab.reward_per_unit * (row.qty or 0), 2)
    return 0


def serve_html(path: str):
    return FileResponse(path)


def _get_price_list_access_scope(current_user: models.User, db: Session):
    if current_user.role in {"Admin", "Owner", "HR", "Accounts", "MISExecutive"}:
        return None
    if current_user.role in {"BrandManager", "BrandPartner"}:
        brand_ids = [user_brand.brand_id for user_brand in (getattr(current_user, "brands", []) or [])]
        return {"brand_ids": brand_ids}
    if current_user.role == "CategoryManager":
        category_codes = category_codes_for_user(current_user)
        if not category_codes:
            return {"brand_ids": []}
        brand_rows = (
            db.query(models.Brand.id)
            .join(models.SubCategory, models.Brand.subcategory_id == models.SubCategory.id)
            .join(models.Category, models.SubCategory.category_id == models.Category.id)
            .filter(func.upper(models.Category.code).in_(category_codes))
            .all()
        )
        return {"brand_ids": [brand_id for (brand_id,) in brand_rows]}
    return {"brand_ids": []}


def _apply_price_list_visibility(query, current_user: models.User, db: Session):
    scope = _get_price_list_access_scope(current_user, db)
    if scope is None:
        return query
    brand_ids = scope.get("brand_ids") or []
    if not brand_ids:
        return query.filter(models.PriceListItem.id.is_(None))
    return query.filter(models.PriceListItem.brand_id.in_(brand_ids))


def _filter_price_list_query(query, search: Optional[str]):
    search_text = (search or "").strip()
    if not search_text:
        return query
    # Split into words and require EVERY word to match somewhere (any
    # field, and different words may match different fields) rather than
    # requiring the whole typed phrase to appear literally in one field.
    # That's what let a combined "brand + model" search like "oppo f33"
    # fail before: "Oppo" only lives in the brand name column and "F33"
    # only lives in item_details, so no single column ever contained the
    # literal phrase "oppo f33" together.
    words = search_text.lower().split()
    query = query.outerjoin(models.Brand, models.PriceListItem.brand_id == models.Brand.id)
    for word in words:
        like_term = f"%{word}%"
        query = query.filter(
            or_(
                func.lower(models.PriceListItem.item_details).like(like_term),
                func.lower(models.PriceListItem.model_no).like(like_term),
                func.lower(models.PriceListItem.serial_no).like(like_term),
                func.lower(models.PriceListItem.imei).like(like_term),
                func.lower(models.Brand.name).like(like_term),
            )
        )
    return query


def _serialize_price_list_item(item: models.PriceListItem, current_user: models.User) -> dict:
    full_access = current_user.role in {"Admin", "Owner", "HR", "Accounts", "MISExecutive"}
    brand = getattr(item, "brand", None)
    return {
        "id": item.id,
        "brand_id": item.brand_id,
        "brand_name": getattr(brand, "name", None) or "",
        "item_details": item.item_details,
        "model_no": item.model_no,
        "serial_no": item.serial_no,
        "imei": item.imei,
        "total_stock": item.total_stock if full_access else None,
        "purchase_price": item.purchase_price if full_access else None,
        "msp": item.msp,
        "isp": item.isp,
        "updated_by_username": item.updated_by_username,
        "updated_date": item.updated_date,
    }


def _parse_price_list_upload(file_bytes: bytes, filename: str):
    workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    parsed_items = []
    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_index = None
            header_keywords = {"itemdetails", "item", "model", "modelname", "itemname", "stock", "totalstock", "qty", "purchase", "purchaseprice", "cost", "msp", "ids", "idsprice", "isp", "mrp", "saleprice"}
            for idx, row in enumerate(rows):
                normalized = [_normalize_price_list_header(v) for v in row or []]
                if not any(normalized):
                    continue
                match_count = 0
                for value in normalized:
                    if value in {"itemdetails", "item", "model", "modelname", "itemname", "stock", "totalstock", "qty", "purchase", "purchaseprice", "cost", "msp", "ids", "idsprice", "isp", "mrp", "saleprice"}:
                        match_count += 1
                if match_count >= 3:
                    header_index = idx
                    break
            if header_index is None:
                continue

            header_row = rows[header_index]
            normalized_headers = [(index, _normalize_price_list_header(value)) for index, value in enumerate(header_row or [])]

            # Most price lists have only ONE selling-price column, labelled
            # either "ISP", "MRP", or "Sale Price" - those are historically
            # treated as interchangeable and mapped to the same field below.
            # Some brand sheets (e.g. iQOO) instead have BOTH a literal
            # "ISP" column and a literal "MRP" column as two distinct
            # prices. When both are present on the same header row, don't
            # collapse them into one field - map ISP to IDS Price (msp) and
            # MRP to the MRP field (isp) so neither is discarded.
            has_literal_isp = any(key == "isp" for _, key in normalized_headers)
            has_literal_mrp = any(key in {"mrp", "saleprice"} for _, key in normalized_headers)
            isp_and_mrp_both_present = has_literal_isp and has_literal_mrp

            header_map = {}
            for index, key in normalized_headers:
                if key in {"itemdetails", "item", "model", "modelname", "itemname"}:
                    header_map["item"] = index
                elif key in {"modelno", "modelnumber"}:
                    header_map["model_no"] = index
                elif key in {"serialno", "serialnumber"}:
                    header_map["serial_no"] = index
                elif key in {"imei", "imeino"}:
                    header_map["imei"] = index
                elif key in {"stock", "totalstock", "qty"}:
                    header_map["stock"] = index
                elif key in {"purchase", "purchaseprice", "cost"}:
                    header_map["purchase"] = index
                elif key in {"msp", "ids", "idsprice"}:
                    header_map["msp"] = index
                elif key == "isp":
                    header_map["msp" if isp_and_mrp_both_present else "isp"] = index
                elif key in {"mrp", "saleprice"}:
                    header_map["isp"] = index

            current_brand = None
            for row in rows[header_index + 1:]:
                if not row:
                    continue
                values = ["" if value is None else str(value).strip() for value in row]
                if not any(values):
                    continue

                item_value = values[header_map.get("item", 0)] if header_map.get("item") is not None and header_map.get("item") < len(values) else ""
                model_no_value = values[header_map.get("model_no", 0)] if header_map.get("model_no") is not None and header_map.get("model_no") < len(values) else ""
                serial_no_value = values[header_map.get("serial_no", 0)] if header_map.get("serial_no") is not None and header_map.get("serial_no") < len(values) else ""
                imei_value = values[header_map.get("imei", 0)] if header_map.get("imei") is not None and header_map.get("imei") < len(values) else ""
                stock_value = values[header_map.get("stock", 0)] if header_map.get("stock") is not None and header_map.get("stock") < len(values) else ""
                purchase_value = values[header_map.get("purchase", 0)] if header_map.get("purchase") is not None and header_map.get("purchase") < len(values) else ""
                msp_value = values[header_map.get("msp", 0)] if header_map.get("msp") is not None and header_map.get("msp") < len(values) else ""
                isp_value = values[header_map.get("isp", 0)] if header_map.get("isp") is not None and header_map.get("isp") < len(values) else ""

                if not item_value and not stock_value and not purchase_value and not msp_value and not isp_value:
                    if values[0]:
                        current_brand = values[0]
                    continue

                if not item_value:
                    continue

                parsed_items.append({
                    "brand_name": current_brand or sheet.title or "Unknown",
                    "item_details": item_value,
                    "model_no": model_no_value or None,
                    "serial_no": serial_no_value or None,
                    "imei": imei_value or None,
                    "total_stock": _parse_price_list_numeric(stock_value),
                    "purchase_price": _parse_price_list_numeric(purchase_value),
                    "msp": _parse_price_list_numeric(msp_value),
                    "isp": _parse_price_list_numeric(isp_value),
                    "source_file": filename,
                })
    finally:
        workbook.close()
    return parsed_items


def _normalize_price_list_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def _parse_price_list_numeric(value) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


# ------------------------------------------------------------------
# PRICE LIST UPLOAD FROM IMAGE / PDF (AI vision extraction)
# Same idea as the scheme-document extraction above: a photographed or
# scanned price list sheet is sent to an AI vision model, which reads the
# brand headers and item table(s) and returns rows in the same shape
# produced by _parse_price_list_upload() for Excel files, so both paths
# feed the same insert/update logic in the upload endpoint below.
#
# Provider-agnostic: whichever key is set on the server is used, checked
# in this priority order (or forced with VISION_PROVIDER=anthropic|xai|openai
# if more than one key happens to be set):
#   1. ANTHROPIC_API_KEY  - Claude. Reads images AND PDFs natively.
#   2. XAI_API_KEY        - Grok (xAI). Images only (jpg/png); PDFs are
#                            rejected with a message pointing at Excel/Claude.
#   3. OPENAI_API_KEY     - OpenAI (e.g. gpt-4o-mini). Images only, same as xAI.
# Model names can be overridden via PRICE_LIST_ANTHROPIC_MODEL /
# PRICE_LIST_XAI_MODEL / PRICE_LIST_OPENAI_MODEL if xAI/OpenAI retire or
# rename a model - defaults below are current as of this writing.
# ------------------------------------------------------------------

VISION_PROVIDER_ENV_KEYS = {
    "anthropic": "ANTHROPIC_API_KEY",
    "xai": "XAI_API_KEY",
    "openai": "OPENAI_API_KEY",
}
VISION_PROVIDER_PRIORITY = ["anthropic", "xai", "openai"]


def _get_available_vision_provider() -> Optional[str]:
    """Which AI vision provider to use for price-list image/PDF extraction,
    based on whichever API key(s) are set on the server. If more than one
    key is set, VISION_PROVIDER can force a specific one; otherwise the
    first configured key wins in VISION_PROVIDER_PRIORITY order."""
    forced = (os.getenv("VISION_PROVIDER") or "").strip().lower()
    if forced in VISION_PROVIDER_ENV_KEYS and os.getenv(VISION_PROVIDER_ENV_KEYS[forced]):
        return forced
    for provider in VISION_PROVIDER_PRIORITY:
        if os.getenv(VISION_PROVIDER_ENV_KEYS[provider]):
            return provider
    return None


def _price_list_extraction_instructions() -> str:
    return (
        "You are reading a photographed or scanned price list / rate list sheet "
        "for an electronics retail ERP. It has brand section headers (e.g. "
        "\"Bluestar\", \"Daikin\") each followed by a table of items, typically "
        "with columns similar to Item Details / Model No / Total Stock / "
        "Purchase Price / MSP / ISP - column names, order, and presence can vary "
        "(MSP may be labelled IDS Price, ISP may be labelled MRP or Sale Price, "
        "and some columns may be missing entirely). Read every row across every "
        "brand section in the document/image and reply with ONLY a JSON array - "
        "no prose, no markdown fences. Each element:\n"
        "{\n"
        '  "brand_name": string (the brand section this item is under),\n'
        '  "item_details": string (the item / model name / description),\n'
        '  "model_no": string or null,\n'
        '  "serial_no": string or null,\n'
        '  "imei": string or null,\n'
        '  "total_stock": number or null,\n'
        '  "purchase_price": number or null,\n'
        '  "msp": number or null,\n'
        '  "isp": number or null\n'
        "}\n"
        "Only include rows that clearly have an item name. Use null for anything "
        "not present or not legible - never invent or guess a value."
    )


def _call_anthropic_vision(content_block: dict, instructions: str) -> str:
    """Sends one image/PDF content block to the Claude API and returns the
    raw text of its reply (expected to be a JSON array, as instructed)."""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    model = os.getenv("PRICE_LIST_ANTHROPIC_MODEL", "claude-sonnet-5")

    payload = json.dumps({
        "model": model,
        "max_tokens": 4096,
        "messages": [
            {"role": "user", "content": [content_block, {"type": "text", "text": instructions}]}
        ],
    }).encode("utf-8")

    request = URLRequest(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=60) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError) as exc:
        raise RuntimeError(f"Claude API request failed: {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Unexpected error calling Claude API: {exc}") from exc

    try:
        text_blocks = [block["text"] for block in response_data.get("content", []) if block.get("type") == "text"]
        return "\n".join(text_blocks).strip()
    except Exception as exc:
        raise RuntimeError(f"Unexpected response format from Claude API: {exc}") from exc


SCHEME_CALCULATOR_EXTRACT_INSTRUCTIONS = {
    "sales": (
        'This image shows a phone sales report (brand, model, quantity sold, sale value). '
        'Extract every row into a JSON array, one object per model, matching exactly this shape: '
        '[{"brand":"","model":"","qty":0,"value":0}]  '
        '// qty = units sold, value = total sale value (rupees) for that row. '
        'Respond with ONLY the JSON array, no markdown fences, no commentary, no explanation.'
    ),
    "scheme": (
        'This image shows a mobile phone scheme sheet (brand, model, scheme payout % - possibly with '
        'a monthly target or slab-wise tiers). '
        'Extract every row into a JSON array, one object per model, matching exactly this shape: '
        '[{"brand":"","model":"","percent":0,"target":0,"slabs":[{"qty":0,"percent":0}]}]  '
        '// percent = flat scheme % if no target/slabs apply; target = monthly unit target if the sheet '
        'states one (else 0); slabs = tiered qty->percent breakpoints if the sheet shows a slab table '
        '(else empty array). '
        'If a row is a catch-all/default rate that applies to every model not listed elsewhere on the '
        'sheet (e.g. labeled "Remaining All Model", "All Other Models", "Others", or similar), keep '
        'that row\'s wording in "model" exactly as printed - do not shorten, rename, or drop it. '
        'Respond with ONLY the JSON array, no markdown fences, no commentary, no explanation.'
    ),
}


@app.post("/api/scheme-calculator/extract-image")
def scheme_calculator_extract_image(
    file: UploadFile = File(...),
    kind: str = Form(...),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Reads a photographed/scanned Sales Data or Scheme Sheet for the
    Scheme Calculator page and returns extracted rows as JSON. Runs
    server-side (unlike the Scheme Calculator's Excel/CSV parsing, which
    happens entirely in the browser) because it needs ANTHROPIC_API_KEY,
    which must never be exposed to client-side code."""
    if kind not in SCHEME_CALCULATOR_EXTRACT_INSTRUCTIONS:
        raise HTTPException(status_code=400, detail="kind must be 'sales' or 'scheme'")

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail="Reading a photo requires ANTHROPIC_API_KEY to be configured on the server. "
                   "Upload an Excel/CSV file instead, or ask an Admin to set that key.",
        )

    raw_bytes = file.file.read()
    try:
        image_b64, media_type = _image_bytes_to_supported_b64(
            file.filename or "upload.jpg", file.content_type or "", raw_bytes
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    content_block = {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_b64}}
    try:
        raw_text = _call_anthropic_vision(content_block, SCHEME_CALCULATOR_EXTRACT_INSTRUCTIONS[kind])
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc))

    cleaned = raw_text.replace("```json", "").replace("```", "").strip()
    try:
        rows = json.loads(cleaned)
        if not isinstance(rows, list):
            raise ValueError("Expected a JSON array")
    except (json.JSONDecodeError, ValueError) as exc:
        raise HTTPException(status_code=502, detail=f"Couldn't parse what the AI read from the image: {exc}")

    return {"rows": rows}


def _call_openai_compatible_vision(api_key: str, base_url: str, model: str, image_b64: str,
                                    media_type: str, instructions: str, provider_label: str) -> str:
    """Sends one base64 image to an OpenAI-compatible /chat/completions
    endpoint (used for both xAI/Grok and OpenAI itself, since xAI's API
    mirrors OpenAI's request/response shape) and returns the raw reply text."""
    payload = json.dumps({
        "model": model,
        "max_tokens": 4096,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": instructions},
                    {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{image_b64}"}},
                ],
            }
        ],
    }).encode("utf-8")

    request = URLRequest(
        base_url,
        data=payload,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=60) as response:
            response_data = json.loads(response.read().decode("utf-8"))
    except (HTTPError, URLError, TimeoutError) as exc:
        raise RuntimeError(f"{provider_label} API request failed: {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Unexpected error calling {provider_label} API: {exc}") from exc

    try:
        return response_data["choices"][0]["message"]["content"] or ""
    except Exception as exc:
        raise RuntimeError(f"Unexpected response format from {provider_label} API: {exc}") from exc


def _image_bytes_to_supported_b64(filename: str, content_type: str, raw_bytes: bytes) -> tuple:
    """xAI/OpenAI's vision endpoints accept jpg/png reliably; webp is
    converted to PNG with Pillow (already a project dependency) so users
    don't have to know which provider is configured to know which image
    formats are safe to upload. Returns (base64_str, media_type)."""
    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""
    if ext == ".webp":
        try:
            from PIL import Image
            image = Image.open(BytesIO(raw_bytes)).convert("RGB")
            buffer = BytesIO()
            image.save(buffer, format="PNG")
            raw_bytes = buffer.getvalue()
            media_type = "image/png"
        except Exception as exc:
            raise RuntimeError(f"Could not convert WEBP image for AI extraction: {exc}") from exc
    elif ext in {".jpg", ".jpeg"}:
        media_type = content_type or "image/jpeg"
    elif ext == ".png":
        media_type = content_type or "image/png"
    else:
        raise RuntimeError("Unsupported image type for AI price list extraction.")
    return base64.b64encode(raw_bytes).decode("ascii"), media_type


def _extract_price_list_rows_from_document(filename: str, content_type: str, raw_bytes: bytes) -> list:
    """Reads a price list image or PDF via whichever AI vision provider is
    configured on the server and returns a list of row dicts shaped like
    _parse_price_list_upload()'s output. Raises RuntimeError with a
    user-facing message on any failure (no key configured, unsupported
    file for that provider, bad/unparseable response) - the upload
    endpoint catches this and reports it back to the user."""
    provider = _get_available_vision_provider()
    if provider is None:
        raise RuntimeError(
            "Reading a price list from an image or PDF requires one AI vision API key "
            "to be configured on the server - any one of ANTHROPIC_API_KEY, XAI_API_KEY, "
            "or OPENAI_API_KEY. Upload an Excel (.xlsx/.xls) file instead, or ask an Admin "
            "to set one of those keys."
        )

    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""
    instructions = _price_list_extraction_instructions()

    if provider == "anthropic":
        content_block = _document_to_llm_content_block(filename, content_type, raw_bytes)
        if not content_block or content_block.get("type") not in ("image", "document"):
            raise RuntimeError("Unsupported file type for image/PDF price list extraction.")
        raw_text = _call_anthropic_vision(content_block, instructions)
    else:
        if ext == ".pdf":
            raise RuntimeError(
                "Reading a PDF price list currently requires ANTHROPIC_API_KEY - the "
                f"{'XAI_API_KEY' if provider == 'xai' else 'OPENAI_API_KEY'} configured on this "
                "server can only read images (.jpg/.jpeg/.png/.webp). Upload a photo/scan "
                "instead, or use an Excel file."
            )
        image_b64, media_type = _image_bytes_to_supported_b64(filename, content_type, raw_bytes)
        if provider == "xai":
            raw_text = _call_openai_compatible_vision(
                api_key=os.getenv("XAI_API_KEY"),
                base_url="https://api.x.ai/v1/chat/completions",
                model=os.getenv("PRICE_LIST_XAI_MODEL", "grok-4.5"),
                image_b64=image_b64,
                media_type=media_type,
                instructions=instructions,
                provider_label="xAI (Grok)",
            )
        else:  # openai
            raw_text = _call_openai_compatible_vision(
                api_key=os.getenv("OPENAI_API_KEY"),
                base_url="https://api.openai.com/v1/chat/completions",
                model=os.getenv("PRICE_LIST_OPENAI_MODEL", "gpt-4o-mini"),
                image_b64=image_b64,
                media_type=media_type,
                instructions=instructions,
                provider_label="OpenAI",
            )

    try:
        cleaned_text = re.sub(r"^```(json)?|```$", "", raw_text.strip(), flags=re.MULTILINE).strip()
        extracted = json.loads(cleaned_text)
    except Exception as exc:
        raise RuntimeError(f"Could not parse the AI's response as JSON: {exc}") from exc

    if not isinstance(extracted, list):
        raise RuntimeError("The AI's response was not a list of price list rows as expected.")

    parsed_items = []
    for row in extracted:
        if not isinstance(row, dict):
            continue
        item_details = (row.get("item_details") or "").strip()
        if not item_details:
            continue
        brand_name = (row.get("brand_name") or "").strip() or "Unknown"
        parsed_items.append({
            "brand_name": brand_name,
            "item_details": item_details,
            "model_no": (str(row.get("model_no")).strip() if row.get("model_no") else None),
            "serial_no": (str(row.get("serial_no")).strip() if row.get("serial_no") else None),
            "imei": (str(row.get("imei")).strip() if row.get("imei") else None),
            "total_stock": _parse_price_list_numeric(row.get("total_stock")),
            "purchase_price": _parse_price_list_numeric(row.get("purchase_price")),
            "msp": _parse_price_list_numeric(row.get("msp")),
            "isp": _parse_price_list_numeric(row.get("isp")),
            "source_file": filename,
        })
    return parsed_items


@app.get("/")
@app.get("/login")
@app.get("/login.html")
def home():
    return serve_html("static/login.html")


@app.get("/signup")
@app.get("/signup.html")
def signup_page():
    return serve_html("static/signup.html")


@app.get("/dashboard")
@app.get("/dashboard.html")
def dashboard_page():
    return serve_html("static/dashboard.html")


@app.get("/home")
@app.get("/home.html")
def app_home_page():
    return serve_html("static/home.html")


@app.get("/attendance")
@app.get("/attendance.html")
def attendance_page():
    return serve_html("static/attendance.html")


@app.get("/price-list")
@app.get("/price-list.html")
def price_list_page():
    return serve_html("static/price_list.html")


@app.get("/purchase-orders")
@app.get("/purchase-orders.html")
def purchase_orders_page():
    return serve_html("static/purchase_orders.html")


@app.get("/analytics")
@app.get("/analytics.html")
def analytics_page():
    return serve_html("static/analytics.html")


@app.get("/privacy")
@app.get("/privacy.html")
def privacy_page():
    return serve_html("static/privacy.html")


@app.get("/ageing-stock")
@app.get("/ageing-stock.html")
def ageing_stock_page():
    return serve_html("static/ageing_stock.html")


@app.get("/scheme-calculator")
@app.get("/scheme-calculator.html")
def scheme_calculator_page():
    return serve_html("static/scheme_calculator.html")


@app.get("/incentive.html")
def incentive_page():
    return serve_html("static/incentive.html")


@app.get("/api/price-list/brands")
def list_price_list_brands(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    scope = _get_price_list_access_scope(current_user, db)
    if scope is None:
        brands = db.query(models.Brand).order_by(models.Brand.name).all()
    else:
        brand_ids = scope.get("brand_ids") or []
        if not brand_ids:
            brands = []
        else:
            brands = db.query(models.Brand).filter(models.Brand.id.in_(brand_ids)).order_by(models.Brand.name).all()
    return [{"id": brand.id, "name": brand.name} for brand in brands]


@app.get("/api/price-list", response_model=List[schemas.PriceListItemOut])
def list_price_list_items(
    brand_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(models.PriceListItem).options()
    query = _apply_price_list_visibility(query, current_user, db)
    if brand_id is not None:
        query = query.filter(models.PriceListItem.brand_id == brand_id)
    query = _filter_price_list_query(query, search)
    items = query.order_by(models.PriceListItem.updated_date.desc(), models.PriceListItem.item_details.asc()).all()
    return [_serialize_price_list_item(item, current_user) for item in items]


@app.post("/api/price-list", response_model=schemas.PriceListItemOut)
def create_price_list_item(
    payload: schemas.PriceListItemCreate,
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    brand = db.query(models.Brand).filter(models.Brand.id == payload.brand_id).first()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    item = models.PriceListItem(
        brand_id=payload.brand_id,
        item_details=(payload.item_details or "").strip(),
        model_no=(payload.model_no or "").strip() or None,
        serial_no=(payload.serial_no or "").strip() or None,
        imei=(payload.imei or "").strip() or None,
        total_stock=payload.total_stock,
        purchase_price=payload.purchase_price,
        msp=payload.msp,
        isp=payload.isp,
        updated_by_user_id=current_user.id,
        updated_by_username=current_user.username,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _serialize_price_list_item(item, current_user)


@app.put("/api/price-list/{item_id}", response_model=schemas.PriceListItemOut)
def update_price_list_item(
    item_id: int,
    payload: schemas.PriceListItemUpdate,
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    item = db.query(models.PriceListItem).filter(models.PriceListItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    if payload.item_details is not None:
        item.item_details = payload.item_details.strip()
    if payload.model_no is not None:
        item.model_no = (payload.model_no or "").strip() or None
    if payload.serial_no is not None:
        item.serial_no = (payload.serial_no or "").strip() or None
    if payload.imei is not None:
        item.imei = (payload.imei or "").strip() or None
    if payload.total_stock is not None:
        item.total_stock = payload.total_stock
    if payload.purchase_price is not None:
        item.purchase_price = payload.purchase_price
    if payload.msp is not None:
        item.msp = payload.msp
    if payload.isp is not None:
        item.isp = payload.isp

    item.updated_by_user_id = current_user.id
    item.updated_by_username = current_user.username
    db.commit()
    db.refresh(item)
    return _serialize_price_list_item(item, current_user)


@app.delete("/api/price-list/{item_id}")
def delete_price_list_item(
    item_id: int,
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    item = db.query(models.PriceListItem).filter(models.PriceListItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"message": "Item deleted"}


PRICE_LIST_EXCEL_EXTENSIONS = {".xlsx", ".xls"}
PRICE_LIST_IMAGE_PDF_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}


@app.post("/api/price-list/upload")
def upload_price_list_items(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    raw_bytes = file.file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    filename = file.filename or "price-list"
    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""

    extraction_method = None
    vision_provider = None
    try:
        if ext in PRICE_LIST_EXCEL_EXTENSIONS:
            parsed_rows = _parse_price_list_upload(raw_bytes, filename)
            extraction_method = "excel"
        elif ext in PRICE_LIST_IMAGE_PDF_EXTENSIONS:
            vision_provider = _get_available_vision_provider()
            parsed_rows = _extract_price_list_rows_from_document(filename, file.content_type, raw_bytes)
            extraction_method = "ai_vision"
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Upload an Excel file (.xlsx/.xls), an image "
                       "(.jpg/.jpeg/.png/.webp), or a PDF.",
            )
    except HTTPException:
        raise
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        label = "Excel file" if ext in PRICE_LIST_EXCEL_EXTENSIONS else "file"
        raise HTTPException(status_code=400, detail=f"Unable to read {label}: {exc}") from exc

    if not parsed_rows:
        no_rows_detail = (
            "No price list items could be read from this image/PDF. Try a clearer photo "
            "or scan, or use an Excel file instead."
            if extraction_method == "ai_vision"
            else "No price list items could be found in this file."
        )
        raise HTTPException(status_code=400, detail=no_rows_detail)

    inserted = 0
    updated = 0
    skipped = 0
    created_brands = []

    for row in parsed_rows:
        item_details = (row.get("item_details") or "").strip()
        if not item_details:
            skipped += 1
            continue

        brand = db.query(models.Brand).filter(func.lower(models.Brand.name) == row["brand_name"].strip().lower()).first()
        if not brand:
            brand = models.Brand(name=row["brand_name"].strip(), subcategory_id=None)
            db.add(brand)
            db.commit()
            db.refresh(brand)
            created_brands.append(brand.name)

        existing = (
            db.query(models.PriceListItem)
            .filter(models.PriceListItem.brand_id == brand.id)
            .filter(func.lower(models.PriceListItem.item_details) == item_details.lower())
            .first()
        )
        if existing:
            existing.model_no = (row.get("model_no") or "").strip() or None
            existing.serial_no = (row.get("serial_no") or "").strip() or None
            existing.imei = (row.get("imei") or "").strip() or None
            existing.total_stock = row.get("total_stock")
            existing.purchase_price = row.get("purchase_price")
            existing.msp = row.get("msp")
            existing.isp = row.get("isp")
            existing.updated_by_user_id = current_user.id
            existing.updated_by_username = current_user.username
            existing.source_file = row.get("source_file")
            existing.updated_date = datetime.utcnow()
            updated += 1
        else:
            db.add(models.PriceListItem(
                brand_id=brand.id,
                item_details=item_details,
                model_no=(row.get("model_no") or "").strip() or None,
                serial_no=(row.get("serial_no") or "").strip() or None,
                imei=(row.get("imei") or "").strip() or None,
                total_stock=row.get("total_stock"),
                purchase_price=row.get("purchase_price"),
                msp=row.get("msp"),
                isp=row.get("isp"),
                source_file=row.get("source_file"),
                updated_by_user_id=current_user.id,
                updated_by_username=current_user.username,
            ))
            inserted += 1

    db.commit()
    message = (
        "Price list read from image/PDF and uploaded. AI-read values can occasionally "
        "misread a digit or column - please spot-check the items below."
        if extraction_method == "ai_vision"
        else "Price list uploaded successfully"
    )
    return {
        "message": message,
        "method": extraction_method,
        "provider": vision_provider,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "brands_created": created_brands,
    }


@app.get("/manifest.webmanifest")
def manifest_file():
    return FileResponse("static/manifest.webmanifest", media_type="application/manifest+json")


@app.get("/sw.js")
def service_worker_file():
    return FileResponse("static/sw.js", media_type="application/javascript")


@app.get("/offline")
@app.get("/offline.html")
def offline_page():
    return serve_html("static/offline.html")


# ============================================================
# AUTH: SIGNUP / LOGIN / CURRENT USER
# ============================================================

@app.post("/auth/signup", response_model=schemas.Token)
def signup(user: schemas.UserSignup, db: Session = Depends(get_db)):
    # --------------------------------------------------------------
    # Invite-code gate: the signup page is public (anyone can reach it
    # once this app is on the Play/App Store). Admin uses a separate secure
    # code; every other role uses the universal code.
    #
    # Override any of these in your environment (Render dashboard ->
    # Environment, or a local .env file) without changing code:
    #   SIGNUP_CODE_ADMIN, SIGNUP_CODE_UNIVERSAL
    # --------------------------------------------------------------
    ADMIN_INVITE_CODE = os.getenv("SIGNUP_CODE_ADMIN", "Initiative@#%_-Admin")
    UNIVERSAL_INVITE_CODE = os.getenv("SIGNUP_CODE_UNIVERSAL", "Initiative@Universal")

    submitted_code = (user.invite_code or "").strip()
    expected_invite_code = ADMIN_INVITE_CODE if user.role == "Admin" else UNIVERSAL_INVITE_CODE
    if submitted_code != expected_invite_code:
        detail = "Invalid Admin invite code" if user.role == "Admin" else "Invalid universal invite code"
        raise HTTPException(status_code=403, detail=detail)

    if user.role == "CategoryManager":
        selected_categories = normalize_category_codes(user.category_codes or user.category_code)
        if not selected_categories:
            raise HTTPException(status_code=400, detail="Select at least one category")

    custom_brand_name = (user.brand_name_other or "").strip()
    if user.role in ("BrandManager", "BrandPartner"):
        if not user.brand_ids and not custom_brand_name:
            raise HTTPException(status_code=400, detail="Select at least one brand")
        if custom_brand_name and not 2 <= len(custom_brand_name) <= 100:
            raise HTTPException(status_code=400, detail="Brand name must contain 2 to 100 characters")

    normalized_email = (user.email or "").strip().lower()
    if not normalized_email:
        raise HTTPException(status_code=400, detail="Email is required")

    existing = (
        db.query(models.User)
        .filter(func.lower(models.User.email) == normalized_email)
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="Only one account can be created with this email address")

    if user.role not in VALID_ROLES:
        raise HTTPException(status_code=400, detail=f"Role must be one of {VALID_ROLES}")

    if user.role in {"ServiceManager", "ACTechnicianA", "ACTechnicianB"}:
        user.store_id = None
    elif user.store_id is not None:
        selected_store = db.query(models.Store).filter(models.Store.id == user.store_id).first()
        if selected_store is None or selected_store.status != "Active":
            raise HTTPException(status_code=400, detail="Select an active outlet")

    category_roles = {"CategoryManager"}
    db_user = models.User(
        username=user.username,
        email=normalized_email,
        password_hash=auth.hash_password(user.password),
        full_name=user.full_name,
        role=user.role,
        store_id=user.store_id,
        category_code=",".join(normalize_category_codes(user.category_codes or user.category_code)) if user.role in category_roles else None,
        status="Active",
    )
    db.add(db_user)
    try:
        db.commit()
    except IntegrityError:
        # The database uniqueness constraint also protects against two
        # simultaneous signup requests using the same email address.
        db.rollback()
        raise HTTPException(status_code=409, detail="Only one account can be created with this email address")
    db.refresh(db_user)

    if user.role in ("BrandManager", "BrandPartner", "CategoryManager"):
        assigned_brand_ids = list(user.brand_ids)
        if user.role in ("BrandManager", "BrandPartner") and custom_brand_name:
            custom_brand = db.query(models.Brand).filter(func.lower(models.Brand.name) == custom_brand_name.lower()).first()
            if custom_brand is None:
                custom_brand = models.Brand(name=custom_brand_name, subcategory_id=None, is_seeded_default=False)
                db.add(custom_brand)
                db.flush()
            if custom_brand.id not in assigned_brand_ids:
                assigned_brand_ids.append(custom_brand.id)
        for brand_id in assigned_brand_ids:
            db.add(models.UserBrand(user_id=db_user.id, brand_id=brand_id))
        db.commit()

    # Sign the newly created account in immediately. This avoids forcing the
    # user through a second form and a second password-hash operation.
    token = auth.create_access_token({"user_id": db_user.id, "role": db_user.role})
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": db_user.role,
        "username": db_user.username,
    }


@app.post("/auth/login", response_model=schemas.Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    identifier = (form_data.username or "").strip()
    user = (
        db.query(models.User)
        .filter((models.User.username == identifier) | (func.lower(models.User.email) == identifier.lower()))
        .first()
    )
    if not user or not auth.verify_password(form_data.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Username or password incorrect")
    if user.status != "Active":
        raise HTTPException(status_code=403, detail="This account is not active")

    token = auth.create_access_token({"user_id": user.id, "role": user.role})
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": user.role,
        "username": user.username,
    }


@app.post("/auth/password-reset/request")
def request_password_reset(payload: schemas.PasswordResetRequest, db: Session = Depends(get_db)):
    identifier = payload.identifier.strip()
    user = db.query(models.User).filter(
        (models.User.username == identifier) | (func.lower(models.User.email) == identifier.lower())
    ).first()
    # Return the same message for unknown accounts to prevent account discovery.
    if not user or not user.email or user.status != "Active":
        return {"message": "If the account exists, a verification code has been sent to its registered email."}
    smtp_password = os.getenv("SMTP_PASSWORD")
    if not smtp_password:
        raise HTTPException(status_code=503, detail="Password reset email is unavailable. Please contact your Admin.")
    code = f"{secrets.randbelow(1_000_000):06d}"
    user.reset_token = hashlib.sha256(code.encode("utf-8")).hexdigest()
    user.reset_token_expires = datetime.utcnow() + timedelta(minutes=15)
    db.commit()
    message = MIMEText(
        f"Your Initiative ERP password reset code is: {code}\n\n"
        "This code expires in 15 minutes. If you did not request it, ignore this email."
    )
    message["Subject"] = "Initiative ERP password reset code"
    message["From"] = os.getenv("SMTP_FROM", "initiative.lucknow@gmail.com")
    message["To"] = user.email
    try:
        with smtplib.SMTP(os.getenv("SMTP_HOST", "smtp.gmail.com"), int(os.getenv("SMTP_PORT", "587")), timeout=15) as server:
            server.starttls()
            server.login(os.getenv("SMTP_USER", "initiative.lucknow@gmail.com"), smtp_password)
            server.send_message(message)
    except Exception:
        user.reset_token = None
        user.reset_token_expires = None
        db.commit()
        raise HTTPException(status_code=503, detail="Unable to send the reset email. Please try again or contact your Admin.")
    return {"message": "If the account exists, a verification code has been sent to its registered email."}


@app.post("/auth/password-reset/confirm")
def confirm_password_reset(payload: schemas.PasswordResetConfirm, db: Session = Depends(get_db)):
    identifier = payload.identifier.strip()
    if len(payload.new_password) < 8:
        raise HTTPException(status_code=400, detail="Password must contain at least 8 characters")
    user = db.query(models.User).filter(
        (models.User.username == identifier) | (func.lower(models.User.email) == identifier.lower())
    ).first()
    code_hash = hashlib.sha256(payload.code.strip().encode("utf-8")).hexdigest()
    if (
        not user or not user.reset_token or not secrets.compare_digest(user.reset_token, code_hash)
        or not user.reset_token_expires or user.reset_token_expires < datetime.utcnow()
    ):
        raise HTTPException(status_code=400, detail="The verification code is invalid or expired")
    user.password_hash = auth.hash_password(payload.new_password)
    user.reset_token = None
    user.reset_token_expires = None
    db.commit()
    return {"message": "Password reset successfully. You can now sign in."}


@app.get("/me", response_model=schemas.UserOut)
def get_me(current_user: models.User = Depends(auth.get_current_user)):
    return current_user


WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")


@app.put("/api/me/weekoff", response_model=schemas.UserOut)
def update_my_weekoff(
    payload: schemas.WeekoffUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    selected_day = (payload.weekoff_day or "").strip().title()
    if selected_day not in WEEKDAYS:
        raise HTTPException(status_code=400, detail="Week Off must be Monday through Sunday")
    current_user.weekoff_day = selected_day
    db.commit()
    db.refresh(current_user)
    return current_user


@app.post("/auth/verify-password")
def verify_own_password(
    payload: schemas.PasswordVerify,
    current_user: models.User = Depends(auth.get_current_user),
):
    """Re-checks the logged-in user's own current password, without issuing
    a new token. Used as a re-auth gate before revealing something sensitive
    (e.g. the User Management table) to someone who's stepped away from an
    already-unlocked session."""
    if not auth.verify_password(payload.password, current_user.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect password")
    return {"valid": True}


# ============================================================
# ROLE-SCOPED DATA (each role only sees what it's allowed to)
# ============================================================

def get_sales_for_user(db: Session, current_user: models.User):
    if current_user.role in ("Admin", "Owner", "HR", "Accounts", "MISExecutive"):
        return db.query(models.Sale).all()
    if current_user.role == "StoreManager":
        return db.query(models.Sale).filter(models.Sale.store_id == current_user.store_id).all()
    if current_user.role == "CategoryManager":
        category_codes = category_codes_for_user(current_user)
        if not category_codes:
            return []
        return (
            db.query(models.Sale)
            .join(models.Category, models.Sale.category_id == models.Category.id)
            .filter(models.Category.code.in_(category_codes))
            .all()
        )
    if current_user.role in ("BrandManager", "BrandPartner"):
        brand_ids = [ub.brand_id for ub in current_user.brands]
        return db.query(models.Sale).filter(models.Sale.brand_id.in_(brand_ids)).all()
    return []


def get_claims_for_user(db: Session, current_user: models.User):
    if current_user.role in ("Admin", "Owner", "HR", "Accounts", "MISExecutive"):
        return db.query(models.ClaimHeader).all()

    if current_user.role in ("BrandManager", "BrandPartner"):
        brand_ids = [ub.brand_id for ub in current_user.brands]
        return (
            db.query(models.ClaimHeader)
            .join(models.Sale, models.ClaimHeader.sale_id == models.Sale.id)
            .filter(models.Sale.brand_id.in_(brand_ids))
            .all()
        )

    if current_user.role == "StoreManager":
        return (
            db.query(models.ClaimHeader)
            .join(models.Sale, models.ClaimHeader.sale_id == models.Sale.id)
            .filter(models.Sale.store_id == current_user.store_id)
            .all()
        )

    if current_user.role == "CategoryManager":
        category_codes = category_codes_for_user(current_user)
        if not category_codes:
            return []
        return (
            db.query(models.ClaimHeader)
            .join(models.Sale, models.ClaimHeader.sale_id == models.Sale.id)
            .join(models.Category, models.Sale.category_id == models.Category.id)
            .filter(models.Category.code.in_(category_codes))
            .all()
        )

    return []


def can_user_access_sale(db: Session, current_user: models.User, sale: models.Sale) -> bool:
    if current_user.role in ("Admin", "Owner", "HR", "Accounts", "MISExecutive"):
        return True

    if current_user.role == "StoreManager":
        return current_user.store_id is not None and sale.store_id == current_user.store_id

    if current_user.role == "CategoryManager":
        category_codes = category_codes_for_user(current_user)
        if not category_codes:
            return False
        sale_category = db.query(models.Category).filter(models.Category.id == sale.category_id).first()
        return bool(sale_category and sale_category.code in category_codes)

    if current_user.role in ("BrandManager", "BrandPartner"):
        brand_ids = [ub.brand_id for ub in current_user.brands]
        return sale.brand_id in brand_ids

    return False

@app.get("/my-scope/sales", response_model=List[schemas.SaleOut])
def my_scope_sales(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return get_sales_for_user(db, current_user)


@app.get("/my-scope/claims", response_model=List[schemas.ClaimOut])
def my_scope_claims(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return get_claims_for_user(db, current_user)


# ============================================================
# MASTERS (Branch / Category / Subcategory / Brand / Product / Variant)
# ============================================================

@app.post("/categories", response_model=schemas.CategoryOut)
def create_category(category: schemas.CategoryCreate, db: Session = Depends(get_db)):
    db_category = models.Category(code=category.code, name=category.name)
    db.add(db_category)
    db.commit()
    db.refresh(db_category)
    return db_category


@app.get("/categories", response_model=List[schemas.CategoryOut])
def list_categories(db: Session = Depends(get_db)):
    allowed_codes = ["HA", "HE", "MH", "IT", "ASC", "OTH"]
    categories = db.query(models.Category).filter(models.Category.code.in_(allowed_codes)).all()

    order_map = {"HA": 0, "HE": 1, "MH": 2, "IT": 3, "ASC": 4, "OTH": 5}

    def sort_key(item):
        code = (item.code or "").upper()
        return (order_map.get(code, 999), (item.name or "").lower())

    return sorted(categories, key=sort_key)


@app.post("/subcategories", response_model=schemas.SubCategoryOut)
def create_subcategory(subcategory: schemas.SubCategoryCreate, db: Session = Depends(get_db)):
    db_subcategory = models.SubCategory(category_id=subcategory.category_id, name=subcategory.name)
    db.add(db_subcategory)
    db.commit()
    db.refresh(db_subcategory)
    return db_subcategory


@app.get("/subcategories", response_model=List[schemas.SubCategoryOut])
def list_subcategories(category_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.SubCategory)
    if category_id is not None:
        query = query.filter(models.SubCategory.category_id == category_id)
    return query.order_by(models.SubCategory.name).all()


@app.post("/brands", response_model=schemas.BrandOut)
def create_brand(brand: schemas.BrandCreate, db: Session = Depends(get_db)):
    db_brand = models.Brand(name=brand.name, subcategory_id=brand.subcategory_id)
    db.add(db_brand)
    db.commit()
    db.refresh(db_brand)
    return db_brand


@app.get("/brands", response_model=List[schemas.BrandOut])
def list_brands(
    subcategory_id: Optional[int] = None,
    category_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    # Tablet-suffixed duplicates remain available to historical records, but
    # brand selectors expose only their canonical Samsung/Xiaomi brands.
    query = db.query(models.Brand).filter(
        ~func.lower(models.Brand.name).in_(["samsung tablet", "xiaomi tablet"])
    )
    if subcategory_id is not None:
        query = query.filter(models.Brand.subcategory_id == subcategory_id)
    elif category_id is not None:
        primary_ids = {
            row[0]
            for row in db.query(models.Brand.id)
            .join(models.SubCategory, models.Brand.subcategory_id == models.SubCategory.id)
            .filter(models.SubCategory.category_id == category_id)
            .all()
        }
        visible_ids = {
            row[0]
            for row in db.query(models.BrandCategoryVisibility.brand_id)
            .filter(models.BrandCategoryVisibility.category_id == category_id)
            .all()
        }
        all_ids = primary_ids | visible_ids
        if not all_ids:
            return []
        query = query.filter(models.Brand.id.in_(all_ids))
    return query.order_by(models.Brand.name).all()


@app.get("/api/brand-category-visibility", response_model=List[schemas.BrandCategoryVisibilityOut])
def list_brand_category_visibility(db: Session = Depends(get_db)):
    """All (brand, category) pairs — a brand can legitimately appear under
    more than one division (e.g. Samsung under both Mobiles and Home
    Entertainment) even though its `subcategory_id` only points to one."""
    return db.query(models.BrandCategoryVisibility).all()


@app.post("/products", response_model=schemas.ProductOut)
def create_product(product: schemas.ProductCreate, db: Session = Depends(get_db)):
    db_product = models.Product(brand_id=product.brand_id, name=product.name)
    db.add(db_product)
    db.commit()
    db.refresh(db_product)
    return db_product


@app.get("/products", response_model=List[schemas.ProductOut])
def list_products(brand_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.Product)
    if brand_id is not None:
        query = query.filter(models.Product.brand_id == brand_id)
    return query.order_by(models.Product.name).all()


@app.post("/variants", response_model=schemas.VariantOut)
def create_variant(variant: schemas.VariantCreate, db: Session = Depends(get_db)):
    db_variant = models.Variant(product_id=variant.product_id, name=variant.name)
    db.add(db_variant)
    db.commit()
    db.refresh(db_variant)
    return db_variant


@app.get("/variants", response_model=List[schemas.VariantOut])
def list_variants(product_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.Variant)
    if product_id is not None:
        query = query.filter(models.Variant.product_id == product_id)
    return query.order_by(models.Variant.name).all()


@app.post("/customers", response_model=schemas.CustomerOut)
def create_customer(customer: schemas.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = models.Customer(name=customer.name, phone=customer.phone, city=customer.city)
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer


@app.get("/customers", response_model=List[schemas.CustomerOut])
def list_customers(db: Session = Depends(get_db)):
    return db.query(models.Customer).order_by(models.Customer.name).all()


@app.post("/dealers", response_model=schemas.DealerOut)
def create_dealer(dealer: schemas.DealerCreate, db: Session = Depends(get_db)):
    db_dealer = models.Dealer(name=dealer.name, city=dealer.city, contact=dealer.contact)
    db.add(db_dealer)
    db.commit()
    db.refresh(db_dealer)
    return db_dealer


@app.get("/dealers", response_model=List[schemas.DealerOut])
def list_dealers(db: Session = Depends(get_db)):
    return db.query(models.Dealer).order_by(models.Dealer.name).all()


@app.post("/stores", response_model=schemas.StoreOut)
def create_store(store: schemas.StoreCreate, db: Session = Depends(get_db)):
    db_store = models.Store(
        name=store.name,
        code=store.code,
        city=store.city,
        status=store.status or "Active",
        latitude=store.latitude,
        longitude=store.longitude,
        geofence_radius_m=store.geofence_radius_m or 100,
    )
    db.add(db_store)
    db.commit()
    db.refresh(db_store)
    return db_store


@app.get("/stores", response_model=List[schemas.StoreOut])
def list_stores(db: Session = Depends(get_db)):
    return db.query(models.Store).order_by(models.Store.code, models.Store.name).all()


# ============================================================
# CURRENT USER PROFILE & ADMIN ASSIGNMENTS
# (Used to scope a Category Manager's Division and Brand choices to only
#  what's assigned to their account.)
# ============================================================

def serialize_user_with_brands(user: models.User, db: Optional[Session] = None) -> dict:
    brand_ids = [ub.brand_id for ub in user.brands]
    brand_names = []
    outlet = None
    if db is not None:
        if brand_ids:
            brand_names = [
                brand.name for brand in db.query(models.Brand)
                .filter(models.Brand.id.in_(brand_ids))
                .order_by(models.Brand.name).all()
            ]
        if user.store_id is not None:
            outlet = db.query(models.Store).filter(models.Store.id == user.store_id).first()
    return {
        "id": user.id,
        "username": user.username,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "store_id": user.store_id,
        "category_code": user.category_code,
        "brand_ids": brand_ids,
        "brand_names": brand_names,
        "outlet_name": outlet.name if outlet else None,
        "outlet_code": outlet.code if outlet else None,
        "status": user.status,
        "weekoff_day": user.weekoff_day,
        "created_date": user.created_date,
    }


@app.get("/api/me", response_model=schemas.MyProfileOut)
def get_my_profile(current_user: models.User = Depends(auth.get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "role": current_user.role,
        "store_id": current_user.store_id,
        "category_code": current_user.category_code,
        "brand_ids": [ub.brand_id for ub in current_user.brands],
        "weekoff_day": current_user.weekoff_day,
    }


def compress_attendance_selfie(data_url: str) -> str:
    """Normalize attendance selfies to a small JPEG before database storage."""
    if not data_url:
        return ""
    try:
        from PIL import Image, ImageOps
        header, encoded = data_url.split(",", 1)
        if not header.lower().startswith("data:image/"):
            raise ValueError("Selfie must be an image data URL")
        raw = base64.b64decode(encoded, validate=True)
        if len(raw) > 8 * 1024 * 1024:
            raise ValueError("Selfie image is too large")
        with Image.open(BytesIO(raw)) as source:
            if source.width * source.height > 20_000_000:
                raise ValueError("Selfie image dimensions are too large")
            image = ImageOps.exif_transpose(source).convert("RGB")
            resampling = getattr(Image, "Resampling", Image)
            image.thumbnail((480, 640), resampling.LANCZOS)
            compressed = BytesIO()
            image.save(compressed, format="JPEG", quality=55, optimize=True, progressive=True)
        return "data:image/jpeg;base64," + base64.b64encode(compressed.getvalue()).decode("ascii")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to process attendance selfie: {exc}")


@app.post("/api/attendance", response_model=schemas.AttendanceOut)
def save_attendance(
    attendance: schemas.AttendanceCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if attendance.action not in {"checkin", "checkout"}:
        raise HTTPException(status_code=400, detail="Action must be checkin or checkout")
    radius = 6371000
    radians = math.pi / 180
    def distance_to(candidate):
        lat_delta = (attendance.latitude - candidate.latitude) * radians
        lon_delta = (attendance.longitude - candidate.longitude) * radians
        value = math.sin(lat_delta / 2) ** 2 + math.cos(candidate.latitude * radians) * math.cos(attendance.latitude * radians) * math.sin(lon_delta / 2) ** 2
        return 2 * radius * math.atan2(math.sqrt(value), math.sqrt(1 - value))
    can_mark_from_anywhere = current_user.role in {"Admin", "HR", "ServiceManager", "ACTechnicianA", "ACTechnicianB"}
    if can_mark_from_anywhere:
        available_stores = db.query(models.Store).filter(
            models.Store.status == "Active",
            models.Store.latitude.isnot(None),
            models.Store.longitude.isnot(None),
        ).all()
        if not available_stores:
            raise HTTPException(status_code=400, detail="No outlet GPS coordinates are configured")
        store, actual_distance = min(((candidate, distance_to(candidate)) for candidate in available_stores), key=lambda pair: pair[1])
    else:
        if current_user.store_id is None:
            raise HTTPException(status_code=400, detail="No outlet is assigned to this account")
        store = db.query(models.Store).filter(models.Store.id == current_user.store_id).first()
        if not store or store.latitude is None or store.longitude is None:
            raise HTTPException(status_code=400, detail="Assigned outlet has no GPS coordinates")
        actual_distance = distance_to(store)
    allowed_radius = store.geofence_radius_m or 100
    if not can_mark_from_anywhere and actual_distance > allowed_radius:
        raise HTTPException(status_code=403, detail=f"Attendance blocked: you are {round(actual_distance)} m from {store.name}; maximum allowed distance is {round(allowed_radius)} m")
    # Never trust the phone clock as the saved punch time. The database always
    # receives authoritative server IST.
    captured_at = verified_attendance_time(attendance.captured_at)
    record = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id == current_user.id,
        models.AttendanceRecord.attendance_date == captured_at.date(),
    ).first()
    if not record:
        if attendance.action == "checkout":
            raise HTTPException(status_code=409, detail="Punch in is required before punch out")
        record = models.AttendanceRecord(
            user_id=current_user.id,
            store_id=store.id,
            attendance_date=captured_at.date(),
        )
        db.add(record)
    if attendance.action == "checkout" and record.checkin_at is None:
        raise HTTPException(status_code=409, detail="Punch in is required before punch out")
    prefix = "checkin" if attendance.action == "checkin" else "checkout"
    if getattr(record, f"{prefix}_at") is not None:
        raise HTTPException(status_code=409, detail=f"{attendance.action.title()} already recorded")
    setattr(record, f"{prefix}_at", captured_at)
    setattr(record, f"{prefix}_selfie", compress_attendance_selfie(attendance.selfie))
    setattr(record, f"{prefix}_latitude", attendance.latitude)
    setattr(record, f"{prefix}_longitude", attendance.longitude)
    # Store the distance calculated from the submitted coordinates and the
    # outlet coordinates, never the phone's claimed distance value.
    setattr(record, f"{prefix}_distance_m", actual_distance)
    setattr(record, f"{prefix}_accuracy_m", attendance.accuracy_m)
    db.commit()
    db.refresh(record)
    return record


@app.get("/api/attendance", response_model=List[schemas.AttendanceOut])
def list_attendance(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.AttendanceRecord)
    if not auth.has_admin_access(current_user):
        query = query.filter(models.AttendanceRecord.user_id == current_user.id)
    records = query.order_by(models.AttendanceRecord.attendance_date.desc()).all()
    users = {user.id: user.username for user in db.query(models.User).all()}
    stores = {store.id: store.name for store in db.query(models.Store).all()}
    return [
        {
            **{column.name: getattr(record, column.name) for column in models.AttendanceRecord.__table__.columns},
            "username": users.get(record.user_id),
            "outlet_name": stores.get(record.store_id),
        }
        for record in records
    ]


@app.get("/api/attendance/{record_id}/selfies")
def get_attendance_selfies(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Return the saved punch-in and punch-out selfies for one attendance record.

    Employees can only view their own record. Admin can view any employee's
    record. Keeping the images behind this small on-demand endpoint avoids
    loading every Base64 selfie just to render attendance tables.
    """
    record = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.id == record_id
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    if not auth.has_admin_access(current_user) and record.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="You are not allowed to view these attendance images")

    user = db.query(models.User).filter(models.User.id == record.user_id).first()
    store = db.query(models.Store).filter(models.Store.id == record.store_id).first() if record.store_id else None
    return {
        "id": record.id,
        "user_id": record.user_id,
        "username": user.username if user else None,
        "outlet_name": store.name if store else None,
        "attendance_date": record.attendance_date.isoformat(),
        "checkin_at": record.checkin_at,
        "second_punch_at": record.second_punch_at,
        "checkout_at": record.checkout_at,
        "checkin_selfie": record.checkin_selfie,
        "second_punch_selfie": record.second_punch_selfie,
        "checkout_selfie": record.checkout_selfie,
    }


@app.post("/api/attendance/location")
def save_attendance_location(
    point: schemas.AttendanceLocationCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    radius = 6371000
    radians = math.pi / 180
    def distance_to(candidate):
        lat_delta = (point.latitude - candidate.latitude) * radians
        lon_delta = (point.longitude - candidate.longitude) * radians
        value = math.sin(lat_delta / 2) ** 2 + math.cos(candidate.latitude * radians) * math.cos(point.latitude * radians) * math.sin(lon_delta / 2) ** 2
        return 2 * radius * math.atan2(math.sqrt(value), math.sqrt(1 - value))
    if current_user.role in {"Admin", "HR", "ServiceManager", "ACTechnicianA", "ACTechnicianB"}:
        available_stores = db.query(models.Store).filter(models.Store.status == "Active", models.Store.latitude.isnot(None), models.Store.longitude.isnot(None)).all()
        if not available_stores:
            raise HTTPException(status_code=400, detail="No outlet GPS coordinates are configured")
        store, actual_distance = min(((candidate, distance_to(candidate)) for candidate in available_stores), key=lambda pair: pair[1])
    else:
        if current_user.store_id is None:
            raise HTTPException(status_code=400, detail="No outlet is assigned to this account")
        store = db.query(models.Store).filter(models.Store.id == current_user.store_id).first()
        if not store or store.latitude is None or store.longitude is None:
            raise HTTPException(status_code=400, detail="Assigned outlet has no GPS coordinates")
        actual_distance = distance_to(store)
    # Location trails also use server IST so changing the phone clock cannot
    # move GPS points to another day or reorder the route history.
    captured_at = datetime.now(INDIA_TZ).replace(tzinfo=None)
    active_record = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id == current_user.id,
        models.AttendanceRecord.attendance_date == captured_at.date(),
        models.AttendanceRecord.checkin_at.isnot(None),
        models.AttendanceRecord.checkout_at.is_(None),
    ).first()
    if not active_record:
        raise HTTPException(status_code=409, detail="Location tracking is available only during working hours")
    previous = db.query(models.AttendanceLocationPoint).filter(
            models.AttendanceLocationPoint.user_id == current_user.id,
            models.AttendanceLocationPoint.captured_at >= datetime.combine(captured_at.date(), datetime.min.time()),
            models.AttendanceLocationPoint.captured_at < captured_at,
        ).order_by(models.AttendanceLocationPoint.captured_at.desc()).first()
    route_distance = 0
    if previous:
        radius = 6371000
        radians = math.pi / 180
        lat_delta = (point.latitude - previous.latitude) * radians
        lon_delta = (point.longitude - previous.longitude) * radians
        value = math.sin(lat_delta / 2) ** 2 + math.cos(previous.latitude * radians) * math.cos(point.latitude * radians) * math.sin(lon_delta / 2) ** 2
        route_distance = 2 * radius * math.atan2(math.sqrt(value), math.sqrt(1 - value))
    location = models.AttendanceLocationPoint(
        user_id=current_user.id, store_id=store.id, captured_at=captured_at,
        latitude=point.latitude, longitude=point.longitude,
        accuracy_m=point.accuracy_m, distance_from_store_m=actual_distance,
        route_distance_m=route_distance,
    )
    db.add(location)
    db.commit()
    return {"id": location.id, "route_distance_m": route_distance}


@app.get("/api/attendance/admin-summary")
def attendance_admin_summary(
    store_id: Optional[int] = Query(None),
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    weekoff_day: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    today = india_today()
    start_date = from_date or today
    end_date = to_date or today
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    single_day = start_date == end_date
    days_in_range = (end_date - start_date).days + 1

    user_query = db.query(models.User).filter(models.User.status == "Active")
    if store_id is not None:
        user_query = user_query.filter(models.User.store_id == store_id)
    if weekoff_day:
        normalized_weekoff = weekoff_day.strip().title()
        if normalized_weekoff not in WEEKDAYS:
            raise HTTPException(status_code=400, detail="Week Off must be Monday through Sunday")
        user_query = user_query.filter(models.User.weekoff_day == normalized_weekoff)
    users = user_query.all()
    stores_by_id = {
        store.id: store for store in db.query(models.Store).all()
    }
    brands_by_id = {
        brand.id: brand.name for brand in db.query(models.Brand).all()
    }
    outlet_abbreviations = {
        "hazratganj": "HZT",
        "alambagh": "ALM",
        "ashiyana": "ASH",
        "gomtinagar": "GNG",
        "gomti nagar": "GNG",
        "vikas nagar": "VKN",
        "vikasnagar": "VKN",
        "head office": "HO",
        "head-office": "HO",
        "warehouse": "MWH",
    }

    record_query = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.attendance_date >= start_date,
        models.AttendanceRecord.attendance_date <= end_date,
        models.AttendanceRecord.user_id.in_([user.id for user in users] or [-1]),
    )
    records_by_user = {}
    for record in record_query.all():
        records_by_user.setdefault(record.user_id, []).append(record)
    all_records_by_user = defaultdict(list)
    all_record_query = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id.in_([user.id for user in users] or [-1]),
        models.AttendanceRecord.checkin_at.isnot(None),
    ).order_by(models.AttendanceRecord.attendance_date.desc())
    for record in all_record_query.all():
        all_records_by_user[record.user_id].append(record)

    range_start_dt = datetime.combine(start_date, datetime.min.time())
    range_end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time())

    rows = []
    for user in users:
        promoter_brand_names = [
            brands_by_id[user_brand.brand_id]
            for user_brand in user.brands
            if user_brand.brand_id in brands_by_id
        ] if user.role == "BrandPartner" else []
        promoter_brand = ", ".join(promoter_brand_names)
        outlet = stores_by_id.get(user.store_id)
        outlet_name = outlet.name if outlet else None
        outlet_abbreviation = outlet_abbreviations.get(
            (outlet_name or "").strip().lower(),
            outlet.code if outlet and outlet.code else "—",
        )
        user_records = sorted(records_by_user.get(user.id, []), key=lambda r: r.attendance_date)
        present_days = sum(1 for r in user_records if r.checkin_at)
        if user.role in {"ServiceManager", "ACTechnicianA", "ACTechnicianB"} and user_records:
            attendance_outlet = stores_by_id.get(user_records[-1].store_id)
            if attendance_outlet:
                outlet = attendance_outlet
                outlet_name = outlet.name
                outlet_abbreviation = outlet_abbreviations.get(
                    outlet_name.strip().lower(), outlet.code or "—"
                )

        points = db.query(models.AttendanceLocationPoint).filter(
            models.AttendanceLocationPoint.user_id == user.id,
            models.AttendanceLocationPoint.captured_at >= range_start_dt,
            models.AttendanceLocationPoint.captured_at < range_end_dt,
        ).all()
        work_windows = [
            (
                record.checkin_at,
                record.checkout_at
                or (datetime.now(INDIA_TZ).replace(tzinfo=None) if record.attendance_date == today
                    else datetime.combine(record.attendance_date + timedelta(days=1), datetime.min.time())),
            )
            for record in user_records if record.checkin_at
        ]
        points = [
            point for point in points
            if any(start <= point.captured_at <= end for start, end in work_windows)
        ]
        latest_point = max(points, key=lambda point: point.captured_at) if points else None
        day_record = user_records[0] if single_day and user_records else None
        tracking_open = bool(day_record and day_record.checkin_at and not day_record.checkout_at)
        location_active = bool(
            tracking_open
            and start_date == today
            and latest_point
            and latest_point.captured_at >= datetime.now(INDIA_TZ).replace(tzinfo=None) - timedelta(minutes=3)
        )
        current_distance = (
            latest_point.distance_from_store_m if latest_point
            else day_record.checkin_distance_m if day_record and day_record.checkin_at
            else None
        )
        location_tracking_status = (
            "Active" if location_active
            else "Inactive" if tracking_open
            else "Completed" if day_record and day_record.checkout_at
            else "Not started"
        )
        max_point = max(points, key=lambda point: point.distance_from_store_m or 0) if points else None
        punch_distances = [
            (distance, timestamp)
            for record in user_records
            for distance, timestamp in (
                (record.checkin_distance_m, record.checkin_at),
                (record.checkout_distance_m, record.checkout_at),
            )
            if distance is not None
        ]
        max_punch = max(punch_distances, key=lambda item: item[0]) if punch_distances else (0, None)
        max_point_distance = max_point.distance_from_store_m if max_point and max_point.distance_from_store_m else 0
        if max_punch[0] >= max_point_distance:
            maximum_distance, maximum_distance_at = max_punch
        else:
            maximum_distance, maximum_distance_at = max_point_distance, max_point.captured_at
        # Send only lightweight record metadata with the dashboard. Selfies are
        # still fetched on demand for the date selected by the administrator.
        history = [
            {
                "id": record.id,
                "attendance_date": record.attendance_date.isoformat(),
                "checkin_at": record.checkin_at,
                "checkout_at": record.checkout_at,
                "outlet_name": stores_by_id.get(record.store_id).name
                if stores_by_id.get(record.store_id) else None,
            }
            for record in all_records_by_user[user.id]
        ]

        if single_day:
            record = user_records[0] if user_records else None
            is_weekoff = user.weekoff_day == start_date.strftime("%A")
            # When Admin explicitly filters by an assigned Week Off day, the
            # table is an assignment view: show Week Off rather than the
            # selected calendar date's attendance result. Without that filter,
            # normal daily attendance remains Present/Week Off/Absent.
            row_status = (
                "Week Off" if weekoff_day
                else "Present" if record and record.checkin_at
                else "Week Off" if is_weekoff
                else "Absent"
            )
            rows.append({
                "user_id": user.id, "username": user.username, "email": user.email, "role": user.role,
                "promoter_brand": promoter_brand, "outlet_id": user.store_id,
                "outlet_name": outlet_name, "outlet_abbreviation": outlet_abbreviation,
                "status": row_status, "weekoff_day": user.weekoff_day,
                "present_days": present_days, "days_in_range": days_in_range,
                "history": history,
                "checkin_at": record.checkin_at if record else None,
                "checkout_at": record.checkout_at if record else None,
                "route_distance_m": round(sum(point.route_distance_m or 0 for point in points)),
                "max_distance_from_store_m": round(maximum_distance),
                "max_distance_at": maximum_distance_at,
                "last_latitude": latest_point.latitude if latest_point else None,
                "last_longitude": latest_point.longitude if latest_point else None,
                "location_points": len(points),
                "current_distance_from_store_m": round(current_distance) if current_distance is not None else None,
                "location_tracking_status": location_tracking_status,
                "last_location_at": latest_point.captured_at if latest_point else None,
            })
        else:
            rows.append({
                "user_id": user.id, "username": user.username, "email": user.email, "role": user.role,
                "promoter_brand": promoter_brand, "outlet_id": user.store_id,
                "outlet_name": outlet_name, "outlet_abbreviation": outlet_abbreviation,
                "status": f"{present_days}/{days_in_range} Present", "weekoff_day": user.weekoff_day,
                "present_days": present_days, "days_in_range": days_in_range,
                "history": history,
                "checkin_at": None,
                "checkout_at": None,
                "route_distance_m": round(sum(point.route_distance_m or 0 for point in points)),
                "max_distance_from_store_m": round(maximum_distance),
                "max_distance_at": maximum_distance_at,
                "last_latitude": latest_point.latitude if latest_point else None,
                "last_longitude": latest_point.longitude if latest_point else None,
                "location_points": len(points),
                "current_distance_from_store_m": round(current_distance) if current_distance is not None else None,
                "location_tracking_status": location_tracking_status,
                "last_location_at": latest_point.captured_at if latest_point else None,
            })

    if single_day:
        # Keep the daily dashboard in punch order: the first employee to mark
        # attendance is first, the second is next, and so on. Employees who
        # have not checked in remain below them in a predictable name order.
        rows.sort(key=lambda row: (
            row["checkin_at"] is None,
            row["checkin_at"] or datetime.max,
            row["username"].casefold(),
        ))

    return {
        "from_date": start_date, "to_date": end_date, "single_day": single_day, "days_in_range": days_in_range,
        # This metric is a headcount, not the number of possible attendance
        # entries. It must remain 11 for 11 active users whether the selected
        # range is one day, one week, or one month.
        "total": len(users),
        # Present/Absent are also employee counts. Across a date range, an
        # employee is Present when they attended on at least one selected day;
        # otherwise they are Absent. These two values therefore always add up
        # to the active-user headcount instead of multiplying users by days.
        "present": sum(row["status"] == "Present" for row in rows) if single_day else sum(row["present_days"] > 0 for row in rows),
        "absent": sum(row["status"] == "Absent" for row in rows) if single_day else sum(row["present_days"] == 0 for row in rows),
        "weekoff": sum(row["status"] == "Week Off" for row in rows) if single_day else 0,
        "rows": rows,
    }


def build_daily_attendance_whatsapp_report(db: Session, report_date: date) -> str:
    """Build an outlet-wise headcount summary for one attendance date."""
    users = db.query(models.User).filter(models.User.status == "Active").all()
    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.attendance_date == report_date,
        models.AttendanceRecord.user_id.in_([user.id for user in users] or [-1]),
    ).all()
    records_by_user = {record.user_id: record for record in records}
    stores = {store.id: store for store in db.query(models.Store).all()}
    outlet_totals = defaultdict(lambda: {"total": 0, "present": 0, "absent": 0, "weekoff": 0})

    for user in users:
        record = records_by_user.get(user.id)
        store_id = record.store_id if record and record.store_id else user.store_id
        store = stores.get(store_id)
        outlet = (store.code or store.name).strip() if store else "Unassigned"
        status = (
            "present" if record and record.checkin_at
            else "weekoff" if user.weekoff_day == report_date.strftime("%A")
            else "absent"
        )
        outlet_totals[outlet]["total"] += 1
        outlet_totals[outlet][status] += 1

    overall = {
        field: sum(values[field] for values in outlet_totals.values())
        for field in ("total", "present", "absent", "weekoff")
    }
    lines = [
        "*Daily Attendance Report*",
        report_date.strftime("%d %B %Y"),
        "",
        f"*Overall* - Total Employees: {overall['total']} | Present: {overall['present']} | Absent: {overall['absent']} | Week Off: {overall['weekoff']}",
        "",
        "*Outlet-wise Summary*",
    ]
    for outlet in sorted(outlet_totals, key=str.casefold):
        values = outlet_totals[outlet]
        lines.append(
            f"{outlet}: Total {values['total']} | Present {values['present']} | "
            f"Absent {values['absent']} | Week Off {values['weekoff']}"
        )
    return "\n".join(lines)


def send_attendance_whatsapp_report(db: Session, report_date: date) -> dict:
    access_token = os.getenv("WHATSAPP_ACCESS_TOKEN")
    phone_number_id = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
    recipients = [
        value.strip() for value in os.getenv(
            "WHATSAPP_ATTENDANCE_RECIPIENTS", "917521956646"
        ).split(",") if value.strip()
    ]
    if not access_token or not phone_number_id:
        raise HTTPException(status_code=503, detail="WhatsApp Cloud API is not configured on Render.")
    if not recipients:
        raise HTTPException(status_code=503, detail="No attendance WhatsApp recipient is configured.")

    message = build_daily_attendance_whatsapp_report(db, report_date)
    api_version = os.getenv("WHATSAPP_API_VERSION", "v23.0")
    endpoint = f"https://graph.facebook.com/{api_version}/{phone_number_id}/messages"
    template_name = os.getenv("WHATSAPP_ATTENDANCE_TEMPLATE", "").strip()
    template_language = os.getenv("WHATSAPP_ATTENDANCE_TEMPLATE_LANGUAGE", "en_US").strip()
    failures = 0

    for recipient in recipients:
        if template_name:
            body = {
                "messaging_product": "whatsapp", "to": recipient, "type": "template",
                "template": {
                    "name": template_name,
                    "language": {"code": template_language},
                    "components": [{
                        "type": "body",
                        "parameters": [{"type": "text", "text": message.replace("\n", " | ")}],
                    }],
                },
            }
        else:
            body = {
                "messaging_product": "whatsapp", "to": recipient, "type": "text",
                "text": {"body": message},
            }
        api_request = URLRequest(
            endpoint, data=json.dumps(body).encode("utf-8"),
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"},
            method="POST",
        )
        try:
            with urlopen(api_request, timeout=15):
                pass
        except (HTTPError, URLError, TimeoutError):
            failures += 1

    if failures:
        raise HTTPException(status_code=502, detail=f"WhatsApp failed for {failures} recipient(s).")
    return {
        "message": f"Attendance report sent to {len(recipients)} WhatsApp recipient(s).",
        "report_date": report_date.isoformat(), "recipients": len(recipients),
    }


@app.post("/api/attendance/whatsapp-send")
def admin_send_attendance_whatsapp(
    attendance_date: Optional[date] = Query(None, alias="date"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    return send_attendance_whatsapp_report(db, attendance_date or india_today())


@app.post("/api/attendance/whatsapp-daily")
def scheduled_send_attendance_whatsapp(request: Request, db: Session = Depends(get_db)):
    expected_secret = os.getenv("ATTENDANCE_WHATSAPP_CRON_SECRET", "")
    supplied_secret = request.headers.get("X-Cron-Secret", "")
    if not expected_secret or not secrets.compare_digest(expected_secret, supplied_secret):
        raise HTTPException(status_code=401, detail="Invalid scheduler secret")
    return send_attendance_whatsapp_report(db, india_today())


@app.get("/api/attendance/admin-export")
def export_admin_attendance(
    export_format: str = Query("xlsx", alias="format"),
    attendance_date: date = Query(..., alias="date"),
    store_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    weekoff_day: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    """Download the selected day's admin attendance table as Excel or PDF."""
    if export_format not in {"xlsx", "pdf"}:
        raise HTTPException(status_code=400, detail="Format must be xlsx or pdf")
    user_query = db.query(models.User).filter(models.User.status == "Active")
    if store_id is not None:
        user_query = user_query.filter(models.User.store_id == store_id)
    if weekoff_day:
        normalized_weekoff = weekoff_day.strip().title()
        if normalized_weekoff not in WEEKDAYS:
            raise HTTPException(status_code=400, detail="Week Off must be Monday through Sunday")
        user_query = user_query.filter(models.User.weekoff_day == normalized_weekoff)
    users = user_query.order_by(models.User.username).all()
    user_ids = [user.id for user in users]
    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id.in_(user_ids or [-1]),
        models.AttendanceRecord.attendance_date == attendance_date,
    ).all()
    records_by_user = {record.user_id: record for record in records}
    stores = {store.id: store for store in db.query(models.Store).all()}
    start_dt = datetime.combine(attendance_date, datetime.min.time())
    end_dt = start_dt + timedelta(days=1)
    rows = []
    for user in users:
        record = records_by_user.get(user.id)
        is_weekoff = user.weekoff_day == attendance_date.strftime("%A")
        row_status = "Present" if record and record.checkin_at else ("Week Off" if is_weekoff else "Absent")
        if status in {"Present", "Absent", "Week Off"} and row_status != status:
            continue
        points = db.query(models.AttendanceLocationPoint.distance_from_store_m).filter(
            models.AttendanceLocationPoint.user_id == user.id,
            models.AttendanceLocationPoint.captured_at >= start_dt,
            models.AttendanceLocationPoint.captured_at < end_dt,
        ).all()
        distances = [value for (value,) in points if value is not None]
        if record:
            distances.extend(value for value in (record.checkin_distance_m, record.checkout_distance_m) if value is not None)
        store = stores.get(user.store_id)
        rows.append([
            user.username,
            row_status,
            record.checkin_at.strftime("%I:%M %p") if record and record.checkin_at else "-",
            record.checkout_at.strftime("%I:%M %p") if record and record.checkout_at else "-",
            store.name if store else "Unassigned",
            f"{round(max(distances, default=0))} m",
        ])
    headers = ["Employee", "Status", "Check-in", "Check-out", "Outlet", "Distance"]
    filename_base = f"attendance-{attendance_date.isoformat()}"
    if export_format == "xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Attendance Date", attendance_date.strftime("%d-%m-%Y")])
        sheet.append([])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        for cell in sheet[3]:
            cell.font = cell.font.copy(bold=True)
        for column, width in zip("ABCDEF", [26, 14, 15, 15, 22, 16]):
            sheet.column_dimensions[column].width = width
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'})

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    date_style = styles["Normal"].clone("AttendanceReportDate")
    date_style.alignment = 1
    date_style.spaceBefore = 4
    elements = [
        Paragraph("Attendance Report", styles["Title"]),
        Paragraph(attendance_date.strftime("%d %B %Y"), date_style),
        Spacer(1, 6 * mm),
    ]
    table = Table([headers] + rows, repeatRows=1, colWidths=[48 * mm, 28 * mm, 30 * mm, 30 * mm, 42 * mm, 30 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#155eef")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d0d5dd")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(table)
    document.build(elements)
    output.seek(0)
    return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'})


def _attendance_month_range(month: str):
    if not re.fullmatch(r"\d{4}-\d{2}", month or ""):
        raise HTTPException(status_code=400, detail="Month must use YYYY-MM format")
    try:
        year, month_number = map(int, month.split("-"))
        start = date(year, month_number, 1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month")
    return start, date(year, month_number, calendar.monthrange(year, month_number)[1])


def _build_monthly_attendance_workbook(db: Session, users: list, month: str):
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    month_start, month_end = _attendance_month_range(month)
    user_ids = [user.id for user in users]
    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id.in_(user_ids or [-1]),
        models.AttendanceRecord.attendance_date >= month_start,
        models.AttendanceRecord.attendance_date <= month_end,
    ).order_by(models.AttendanceRecord.attendance_date).all()
    records_per_user = defaultdict(int)
    for record in records:
        records_per_user[record.user_id] += 1
    unique_users = {}
    for user in users:
        display_key = (user.username or user.full_name or str(user.id)).strip().casefold()
        current = unique_users.get(display_key)
        user_rank = (records_per_user[user.id], user.store_id is not None, user.id)
        current_rank = (records_per_user[current.id], current.store_id is not None, current.id) if current else None
        if current is None or user_rank > current_rank:
            unique_users[display_key] = user
    users = sorted(unique_users.values(), key=lambda user: ((user.store_id or 0), (user.full_name or user.username or "").lower(), user.id))
    records_by_key = {(record.user_id, record.attendance_date): record for record in records}
    stores = {store.id: store for store in db.query(models.Store).all()}
    days = month_end.day
    today = india_today()

    book = Workbook()
    summary = book.active
    summary.title = "Monthly Summary"
    last_col = 2 + days + 4
    last_letter = get_column_letter(last_col)
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    summary["A1"] = "MONTHLY ATTENDANCE SUMMARY"
    summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    summary["A2"] = month_start.strftime("%B %Y")
    headers = ["Outlet", "Employee"] + list(range(1, days + 1)) + ["Present", "Absent", "Week Off", "Total"]
    summary.append([])
    summary.append(headers)

    navy, blue, white = "17365D", "4472C4", "FFFFFF"
    pale_blue, pale_green, pale_red, pale_orange, pale_gray = "D9EAF7", "D9EAD3", "F4CCCC", "F4B183", "E7E6E6"
    attendance_status_fills = {
        "P": PatternFill("solid", fgColor=pale_green),
        "A": PatternFill("solid", fgColor=pale_red),
        "WO": PatternFill("solid", fgColor=pale_orange),
        "-": PatternFill("solid", fgColor=pale_gray),
    }
    summary["A1"].fill = PatternFill("solid", fgColor=navy)
    summary["A1"].font = Font(color=white, bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary["A2"].fill = PatternFill("solid", fgColor=pale_blue)
    summary["A2"].font = Font(color=navy, bold=True, size=12)
    summary["A2"].alignment = Alignment(horizontal="center", vertical="center")
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    first_day_col, last_day_col = 3, 2 + days
    detail_rows = []
    for row_no, user in enumerate(users, 5):
        store = stores.get(user.store_id)
        display_name = user.full_name or user.username
        if store is None:
            recorded_stores = [stores.get(record.store_id) for record in records if record.user_id == user.id and stores.get(record.store_id)]
            store = recorded_stores[-1] if recorded_stores else None
        summary.cell(row_no, 1, store.name if store else "Unassigned")
        summary.cell(row_no, 2, display_name)
        for day_number in range(1, days + 1):
            current_date = date(month_start.year, month_start.month, day_number)
            record = records_by_key.get((user.id, current_date))
            before_joining = bool(user.created_date and current_date < user.created_date.date())
            if current_date > today or before_joining:
                status = "-"
            elif record and record.checkin_at:
                status = "P"
            elif user.weekoff_day == current_date.strftime("%A"):
                status = "WO"
            else:
                status = "A"
            cell = summary.cell(row_no, first_day_col + day_number - 1, status)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = attendance_status_fills[status]
            if record:
                record_store = stores.get(record.store_id) or store
                detail_rows.append([
                    user.id, record_store.name if record_store else "Unassigned", display_name, current_date,
                    current_date.strftime("%A"), status,
                    record.checkin_at, record.checkout_at,
                    round(record.checkin_distance_m or 0), round(record.checkout_distance_m or 0),
                    "Yes" if record.checkin_selfie else "No", "Yes" if record.checkout_selfie else "No",
                ])
        day_start, day_end = get_column_letter(first_day_col), get_column_letter(last_day_col)
        summary.cell(row_no, last_day_col + 1, f'=COUNTIF({day_start}{row_no}:{day_end}{row_no},"P")')
        summary.cell(row_no, last_day_col + 2, f'=COUNTIF({day_start}{row_no}:{day_end}{row_no},"A")')
        summary.cell(row_no, last_day_col + 3, f'=COUNTIF({day_start}{row_no}:{day_end}{row_no},"WO")')
        present_cell = f"{get_column_letter(last_day_col + 1)}{row_no}"
        summary.cell(row_no, last_day_col + 4, f'={present_cell}&"/{days}"')
        summary.cell(row_no, last_day_col + 4).alignment = Alignment(horizontal="center")

    data_end = max(5, 4 + len(users))
    day_range = f"{get_column_letter(first_day_col)}5:{get_column_letter(last_day_col)}{data_end}"
    for value, color in (("P", pale_green), ("A", pale_red), ("WO", pale_orange), ("-", pale_gray)):
        summary.conditional_formatting.add(day_range, CellIsRule(operator="equal", formula=[f'"{value}"'], fill=PatternFill("solid", fgColor=color)))
    thin = Side(style="thin", color="D0D5DD")
    for row in summary.iter_rows(min_row=4, max_row=data_end, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = Border(bottom=thin)
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 25
    for col in range(first_day_col, last_day_col + 1):
        summary.column_dimensions[get_column_letter(col)].width = 6
    for col in range(last_day_col + 1, last_col + 1):
        summary.column_dimensions[get_column_letter(col)].width = 12
    summary.freeze_panes = "A5"
    summary.row_dimensions[1].height = 34
    summary.row_dimensions[2].height = 25
    summary.row_dimensions[3].height = 10
    summary.row_dimensions[4].height = 30
    summary.sheet_view.showGridLines = False
    summary.sheet_view.zoomScale = 85
    summary.auto_filter.ref = f"A4:{last_letter}{data_end}"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    summary.print_title_rows = "1:4"

    detail = book.create_sheet("Attendance Detail")
    detail_headers = ["User ID", "Outlet", "Employee", "Date", "Day", "Status", "Punch In", "Punch Out", "Punch In Distance (m)", "Punch Out Distance (m)", "Punch In Photo", "Punch Out Photo"]
    detail.append(detail_headers)
    for row in detail_rows:
        detail.append(row)
    for cell in detail[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row_no in range(2, len(detail_rows) + 2):
        detail.cell(row_no, 4).number_format = "dd-mmm-yyyy"
        detail.cell(row_no, 6).fill = attendance_status_fills.get(detail.cell(row_no, 6).value, attendance_status_fills["-"])
        detail.cell(row_no, 6).alignment = Alignment(horizontal="center")
        detail.cell(row_no, 7).number_format = "hh:mm AM/PM"
        detail.cell(row_no, 8).number_format = "hh:mm AM/PM"
        detail.cell(row_no, 9).number_format = '0.0'
        detail.cell(row_no, 10).number_format = '0.0'
    widths = [10, 20, 25, 14, 14, 11, 18, 18, 22, 23, 17, 18]
    for index, width in enumerate(widths, 1):
        detail.column_dimensions[get_column_letter(index)].width = width
    detail.column_dimensions["A"].hidden = True
    detail.freeze_panes = "B2"
    detail.auto_filter.ref = f"A1:L{max(1, len(detail_rows) + 1)}"
    detail.sheet_view.showGridLines = False
    return book


@app.get("/api/attendance/monthly-export")
def export_monthly_attendance(
    month: str = Query(...),
    scope: str = Query("self"),
    store_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Export a monthly attendance matrix for oneself or, for admins, all active employees."""
    if scope not in {"self", "all"}:
        raise HTTPException(status_code=400, detail="Scope must be self or all")
    _attendance_month_range(month)
    if scope == "all":
        if current_user.role not in {"Admin", "Owner", "HR"}:
            raise HTTPException(status_code=403, detail="Only attendance administrators can export all employees")
        query = db.query(models.User).filter(models.User.status == "Active")
        if store_id is not None:
            query = query.filter(models.User.store_id == store_id)
        users = query.all()
        filename = f"all-outlets-attendance-{month}.xlsx" if store_id is None else f"outlet-attendance-{month}.xlsx"
    else:
        users = [current_user]
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "-", current_user.username).strip("-") or "employee"
        filename = f"{safe_name}-attendance-{month}.xlsx"
    workbook = _build_monthly_attendance_workbook(db, users, month)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/attendance/admin-user-export")
def export_admin_user_attendance(
    username: str = Query(...),
    export_format: str = Query("xlsx", alias="format"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    """Download one employee's complete attendance history as Excel or PDF."""
    if export_format not in {"xlsx", "pdf"}:
        raise HTTPException(status_code=400, detail="Format must be xlsx or pdf")
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id == user.id
    ).order_by(models.AttendanceRecord.attendance_date.desc()).all()
    stores = {store.id: store.name for store in db.query(models.Store).all()}
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Full Attendance"
    sheet.append(["Employee", user.username])
    sheet.append(["Generated at", datetime.now(INDIA_TZ).strftime("%d-%m-%Y %I:%M %p")])
    sheet.append([])
    headers = [
        "Date", "Day", "Status", "Punch In", "Punch Out", "Working Hours",
        "Outlet", "Punch In Distance", "Punch Out Distance", "Max Distance",
        "Punch In Photo", "Punch Out Photo",
    ]
    sheet.append(headers)
    export_rows = []
    for record in records:
        start_dt = datetime.combine(record.attendance_date, datetime.min.time())
        end_dt = start_dt + timedelta(days=1)
        tracked = db.query(models.AttendanceLocationPoint.distance_from_store_m).filter(
            models.AttendanceLocationPoint.user_id == user.id,
            models.AttendanceLocationPoint.captured_at >= start_dt,
            models.AttendanceLocationPoint.captured_at < end_dt,
        ).all()
        distances = [value for (value,) in tracked if value is not None]
        distances.extend(value for value in (record.checkin_distance_m, record.checkout_distance_m) if value is not None)
        if record.checkin_at and record.checkout_at:
            minutes = max(0, int((record.checkout_at - record.checkin_at).total_seconds() // 60))
            working_hours = f"{minutes // 60:02d}:{minutes % 60:02d}"
        elif record.checkin_at:
            working_hours = "In progress"
        else:
            working_hours = "-"
        row_values = [
            record.attendance_date.strftime("%d-%m-%Y"),
            record.attendance_date.strftime("%A"),
            "Present" if record.checkin_at else "Absent",
            record.checkin_at.strftime("%I:%M %p") if record.checkin_at else "-",
            record.checkout_at.strftime("%I:%M %p") if record.checkout_at else "-",
            working_hours,
            stores.get(record.store_id, "Unassigned"),
            f"{round(record.checkin_distance_m or 0)} m",
            f"{round(record.checkout_distance_m or 0)} m",
            f"{round(max(distances, default=0))} m",
            "Yes" if record.checkin_selfie else "No",
            "Yes" if record.checkout_selfie else "No",
        ]
        export_rows.append(row_values)
        sheet.append(row_values)
    safe_username = re.sub(r"[^A-Za-z0-9_-]+", "-", user.username).strip("-") or "employee"
    if export_format == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        output = BytesIO()
        document = SimpleDocTemplate(output, pagesize=landscape(A3), rightMargin=10 * mm, leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
        styles = getSampleStyleSheet()
        subtitle = styles["Normal"].clone("EmployeeAttendanceSubtitle")
        subtitle.alignment = 1
        elements = [
            Paragraph(f"Attendance Report - {user.username}", styles["Title"]),
            Paragraph("Complete attendance history", subtitle),
            Spacer(1, 5 * mm),
        ]
        table = Table([headers] + export_rows, repeatRows=1, colWidths=[25 * mm, 27 * mm, 20 * mm, 25 * mm, 25 * mm, 28 * mm, 36 * mm, 31 * mm, 32 * mm, 28 * mm, 27 * mm, 28 * mm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#155eef")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d0d5dd")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        document.build(elements)
        output.seek(0)
        return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{safe_username}-full-attendance.pdf"'})
    for cell in sheet[4]:
        cell.font = cell.font.copy(bold=True)
    widths = [14, 14, 12, 15, 15, 17, 22, 20, 21, 17, 18, 19]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A5"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_username}-full-attendance.xlsx"'},
    )


@app.get("/api/users", response_model=List[schemas.UserAdminOut])
def list_users_for_admin(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_user_management_admin),
):
    users = db.query(models.User).order_by(models.User.username).all()
    return [serialize_user_with_brands(u, db) for u in users]


@app.get("/api/users/count", response_model=schemas.UserCountOut)
def count_registered_users(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_user_management_admin),
):
    """Total number of accounts registered on this system, plus a
    breakdown by role, shown on the Admin dashboard."""
    rows = db.query(models.User.role, func.count(models.User.id)).group_by(models.User.role).all()
    by_role = {role: count for role, count in rows}
    return {"total": sum(by_role.values()), "by_role": by_role}


@app.post("/api/users/{user_id}/reset-password")
def admin_reset_user_password(
    user_id: int,
    payload: schemas.AdminPasswordReset,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_user_management_admin),
):
    """Admin-only account recovery: directly set a new password for any
    user. This replaces the old email-based forgot-password flow, since
    outbound account-recovery email isn't configured in this deployment -
    a user who's locked out should ask an Admin to reset their password
    here instead."""
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    new_password = (payload.new_password or "").strip()
    if len(new_password) < 4:
        raise HTTPException(status_code=400, detail="New password must be at least 4 characters")

    target_user.password_hash = auth.hash_password(new_password)
    target_user.reset_token = None
    target_user.reset_token_expires = None
    db.commit()

    return {"message": "Reset Password Successfully", "username": target_user.username}


@app.patch("/api/users/{user_id}/assignments", response_model=schemas.UserAdminOut)
def update_user_assignments(
    user_id: int,
    payload: schemas.UserAssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_user_management_admin),
):
    """Assign which store, category (division), and brands a user — typically
    a CategoryManager — can see on the Purchase Orders page."""
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if payload.store_id is not None:
        target_user.store_id = payload.store_id
    if payload.category_code is not None:
        target_user.category_code = ",".join(normalize_category_codes(payload.category_code)) or None
    if payload.brand_ids is not None:
        db.query(models.UserBrand).filter(models.UserBrand.user_id == target_user.id).delete()
        for brand_id in payload.brand_ids:
            db.add(models.UserBrand(user_id=target_user.id, brand_id=brand_id))

    db.commit()
    db.refresh(target_user)
    return serialize_user_with_brands(target_user, db)


@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_user_management_admin),
):
    """Admin/HR: permanently delete an allowed user account, regardless of what
    they've submitted/uploaded in the past. Purchase orders require an
    owning user (that column is NOT NULL), so any the deleted user
    submitted or approved are reassigned to the admin doing the deletion,
    with a note recording who originally submitted them - the order itself
    and its history are kept, just no longer tied to the removed login.
    Other historical references (uploaded scheme attachments, price list
    edits, etc.) are simply detached from the deleted user."""
    target_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")

    if target_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    if current_user.role == "HR" and target_user.role == "Admin":
        raise HTTPException(status_code=403, detail="HR cannot delete an Admin account.")

    deleted_username = target_user.username

    # Purchase orders must always have a submitter, so reassign any this
    # user submitted to the admin performing the deletion, and record who
    # originally submitted it in processing_notes for the audit trail.
    submitted_orders = (
        db.query(models.PurchaseOrder)
        .filter(models.PurchaseOrder.submitted_by_user_id == target_user.id)
        .all()
    )
    for po in submitted_orders:
        note = f"[Originally submitted by {deleted_username} - account deleted]"
        po.processing_notes = f"{po.processing_notes}\n{note}" if po.processing_notes else note
        po.submitted_by_user_id = current_user.id

    db.query(models.PurchaseOrder).filter(models.PurchaseOrder.approved_by_user_id == target_user.id).update(
        {"approved_by_user_id": None}
    )

    # Detach other, optional historical references so deleting the account
    # doesn't break foreign-key constraints on records that don't strictly
    # need an owning user (the record itself, and its data, stays intact).
    db.query(models.SchemeAttachment).filter(models.SchemeAttachment.uploaded_by_user_id == target_user.id).update(
        {"uploaded_by_user_id": None}
    )
    db.query(models.IntervalSaleUpload).filter(models.IntervalSaleUpload.uploaded_by == target_user.id).update(
        {"uploaded_by": None}
    )
    db.query(models.AnalyticsUpload).filter(models.AnalyticsUpload.uploaded_by == target_user.id).update(
        {"uploaded_by": None}
    )
    db.query(models.PriceListItem).filter(models.PriceListItem.updated_by_user_id == target_user.id).update(
        {"updated_by_user_id": None}
    )
    db.query(models.AgeingStockUpload).filter(models.AgeingStockUpload.uploaded_by == target_user.id).update(
        {"uploaded_by": None}
    )

    # Attendance belongs to the employee account itself.  These tables use
    # required foreign keys, so they cannot be detached like upload/audit
    # records.  Remove the location trail first, followed by daily attendance,
    # before deleting the login.  This makes Admin deletion work whether or
    # not the employee has punched in previously.
    db.query(models.AttendanceLocationPoint).filter(
        models.AttendanceLocationPoint.user_id == target_user.id
    ).delete(synchronize_session=False)
    db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.user_id == target_user.id
    ).delete(synchronize_session=False)
    db.query(models.UserBrand).filter(
        models.UserBrand.user_id == target_user.id
    ).delete(synchronize_session=False)

    try:
        db.delete(target_user)
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"Unable to delete {deleted_username}. A protected system record still references this account.",
        ) from exc

    return {"message": f"Account '{deleted_username}' deleted"}


# ============================================================
# PURCHASE ORDERS
# ============================================================

PURCHASE_ORDER_STATUSES = {"Requested", "Approved", "Rejected", "Ordered", "Cancelled"}

# Which roles are allowed to move a PO into which status. A Category
# Manager's request starts as "Requested". Only Admin can Approve or Reject
# it. Only once it's "Approved" can Admin/MIS move it on to "Ordered" (i.e.
# finalized and sent to the supplier) or "Cancelled".
ADMIN_ONLY_STATUSES = {"Approved", "Rejected", "Requested"}


def assert_status_transition_allowed(current_user: models.User, purchase_order: models.PurchaseOrder, new_status: str):
    if new_status in ADMIN_ONLY_STATUSES and not auth.has_admin_access(current_user):
        raise HTTPException(
            status_code=403,
            detail="Only Admin can approve, reject, or reopen a purchase order.",
        )
    if new_status == "Ordered" and purchase_order.status != "Approved":
        raise HTTPException(
            status_code=400,
            detail="This purchase order must be Approved by Admin before it can be marked Ordered.",
        )


def serialize_purchase_order(purchase_order: models.PurchaseOrder, notification_status: Optional[str] = None):
    return {
        "id": purchase_order.id,
        "request_no": purchase_order.request_no,
        "request_date": purchase_order.request_date,
        "division": purchase_order.division,
        "branch_id": purchase_order.branch_id,
        "brand_name": purchase_order.brand_name,
        "supplier_name": purchase_order.supplier_name,
        "supplier_email": purchase_order.supplier_email,
        "supplier_address": purchase_order.supplier_address,
        "supplier_gstin": purchase_order.supplier_gstin,
        "delivery_address": purchase_order.delivery_address,
        "remarks": purchase_order.remarks,
        "status": purchase_order.status,
        "busy_po_number": purchase_order.busy_po_number,
        "ordered_date": purchase_order.ordered_date,
        "processing_notes": purchase_order.processing_notes,
        "exported_to_busy": purchase_order.exported_to_busy,
        "exported_to_busy_at": purchase_order.exported_to_busy_at,
        "submitted_by_user_id": purchase_order.submitted_by_user_id,
        "submitted_by_username": purchase_order.submitted_by.username if purchase_order.submitted_by else None,
        "approved_by_username": (getattr(purchase_order, "approved_by", None).username if getattr(purchase_order, "approved_by", None) else None),
        "approved_date": getattr(purchase_order, "approved_date", None),
        "created_date": purchase_order.created_date,
        "updated_date": purchase_order.updated_date,
        "items": purchase_order.items,
        "notification_status": notification_status,
    }


def send_purchase_order_whatsapp_notification(purchase_order: models.PurchaseOrder) -> str:
    """Send a short MIS alert through WhatsApp Cloud API when configured."""
    access_token = os.getenv("WHATSAPP_ACCESS_TOKEN")
    phone_number_id = os.getenv("WHATSAPP_PHONE_NUMBER_ID")
    recipients = [value.strip() for value in os.getenv("WHATSAPP_MIS_RECIPIENTS", "").split(",") if value.strip()]
    if not (access_token and phone_number_id and recipients):
        return "Not sent: WhatsApp notification is not configured."

    requester = purchase_order.submitted_by.username if purchase_order.submitted_by else "Unknown user"
    message = (
        f"New PO request {purchase_order.request_no} from {requester}. "
        f"Branch: {purchase_order.branch_id or 'Not selected'} | "
        f"Items: {len(purchase_order.items)} | Status: {purchase_order.status}."
    )
    api_version = os.getenv("WHATSAPP_API_VERSION", "v23.0")
    endpoint = f"https://graph.facebook.com/{api_version}/{phone_number_id}/messages"
    failures = 0

    for recipient in recipients:
        payload = json.dumps({
            "messaging_product": "whatsapp",
            "to": recipient,
            "type": "text",
            "text": {"body": message},
        }).encode("utf-8")
        request = URLRequest(
            endpoint,
            data=payload,
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        try:
            with urlopen(request, timeout=10):
                pass
        except (HTTPError, URLError, TimeoutError):
            failures += 1

    if failures:
        return f"Saved, but WhatsApp delivery failed for {failures} recipient(s)."
    return f"WhatsApp notification sent to {len(recipients)} MIS recipient(s)."


def can_access_purchase_order(current_user: models.User, purchase_order: models.PurchaseOrder) -> bool:
    return auth.has_admin_access(current_user) or current_user.role == "MISExecutive" or purchase_order.submitted_by_user_id == current_user.id


# ============================================================
# BRAND SUPPLIER EMAIL BOOK
# ============================================================

DEFAULT_BRAND_SUPPLIER_EMAILS = {
    "Samsung": ["orders@samsung.com", "sales@samsung.com", "purchase@samsung.com", "support@samsung.com", "distributor@samsung.com"],
    "LG": ["orders@lg.com", "sales@lg.com", "purchase@lg.com", "support@lg.com", "distributor@lg.com"],
    "Haier": ["orders@haier.com", "sales@haier.com", "purchase@haier.com", "support@haier.com", "distributor@haier.com"],
    "Vivo": ["orders@vivo.com", "sales@vivo.com", "purchase@vivo.com", "support@vivo.com", "distributor@vivo.com"],
    "Oppo": ["orders@oppo.com", "sales@oppo.com", "purchase@oppo.com", "support@oppo.com", "distributor@oppo.com"],
    "Redmi": ["orders@mi.com", "sales@mi.com", "purchase@mi.com", "support@mi.com", "distributor@mi.com"],
}


def seed_default_brand_supplier_emails(db: Session):
    existing_count = db.query(models.BrandSupplierEmail).count()
    if existing_count:
        return
    for brand_name, emails in DEFAULT_BRAND_SUPPLIER_EMAILS.items():
        for email in emails:
            db.add(models.BrandSupplierEmail(brand_name=brand_name, email=email))
    db.commit()


with SessionLocal() as _db:
    seed_default_brand_supplier_emails(_db)


@app.get("/api/brand-emails", response_model=List[schemas.BrandSupplierEmailOut])
def list_brand_emails(
    brand: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.BrandSupplierEmail)
    if brand:
        query = query.filter(models.BrandSupplierEmail.brand_name == brand)
    return query.order_by(models.BrandSupplierEmail.brand_name, models.BrandSupplierEmail.id).all()


@app.post("/api/brand-emails", response_model=schemas.BrandSupplierEmailOut)
def add_brand_email(
    payload: schemas.BrandSupplierEmailCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    brand_name = payload.brand_name.strip()
    email = payload.email.strip().lower()
    if not brand_name or not email:
        raise HTTPException(status_code=400, detail="Brand and email are required")
    existing = (
        db.query(models.BrandSupplierEmail)
        .filter(models.BrandSupplierEmail.brand_name == brand_name, models.BrandSupplierEmail.email == email)
        .first()
    )
    if existing:
        return existing
    row = models.BrandSupplierEmail(brand_name=brand_name, email=email)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.delete("/api/brand-emails/{email_id}")
def delete_brand_email(
    email_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    row = db.query(models.BrandSupplierEmail).filter(models.BrandSupplierEmail.id == email_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Email not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


# ============================================================
# SUPPLIER PROFILE (per supplier name)
# Address, GSTIN, and every email entered for a supplier are remembered
# under that supplier's name, so the next purchase order for the same
# supplier (on any brand/division) can auto-fill instead of retyping.
# ============================================================

def upsert_supplier_profile(db: Session, supplier_name: Optional[str], supplier_address: Optional[str], supplier_gstin: Optional[str]):
    """Save/update the address and GSTIN on file for this supplier name.
    Only overwrites a field when a non-blank value was actually provided,
    so clearing one field on one PO doesn't blank it out for every other
    request that shares the same supplier."""
    name = (supplier_name or "").strip()
    if not name:
        return
    profile = (
        db.query(models.SupplierProfile)
        .filter(func.lower(models.SupplierProfile.supplier_name) == name.lower())
        .first()
    )
    if not profile:
        profile = models.SupplierProfile(supplier_name=name)
        db.add(profile)
    if supplier_address and supplier_address.strip():
        profile.supplier_address = supplier_address.strip()
    if supplier_gstin and supplier_gstin.strip():
        profile.supplier_gstin = supplier_gstin.strip()
    db.commit()


def upsert_supplier_email(db: Session, supplier_name: Optional[str], email: Optional[str]):
    """Remember this email under the supplier's name (no duplicates)."""
    name = (supplier_name or "").strip()
    email = (email or "").strip().lower()
    if not name or not email:
        return
    existing = (
        db.query(models.SupplierEmail)
        .filter(func.lower(models.SupplierEmail.supplier_name) == name.lower(), models.SupplierEmail.email == email)
        .first()
    )
    if existing:
        return
    db.add(models.SupplierEmail(supplier_name=name, email=email))
    db.commit()


@app.get("/api/supplier-profile", response_model=schemas.SupplierProfileOut)
def get_supplier_profile(
    supplier_name: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    name = supplier_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="supplier_name is required")

    profile = (
        db.query(models.SupplierProfile)
        .filter(func.lower(models.SupplierProfile.supplier_name) == name.lower())
        .first()
    )
    email_rows = (
        db.query(models.SupplierEmail)
        .filter(func.lower(models.SupplierEmail.supplier_name) == name.lower())
        .order_by(models.SupplierEmail.id)
        .all()
    )
    return {
        "supplier_name": profile.supplier_name if profile else name,
        "supplier_address": profile.supplier_address if profile else None,
        "supplier_gstin": profile.supplier_gstin if profile else None,
        "emails": email_rows,
    }


@app.post("/api/supplier-emails", response_model=schemas.SupplierEmailEntry)
def add_supplier_email(
    payload: schemas.SupplierEmailCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    supplier_name = payload.supplier_name.strip()
    email = payload.email.strip().lower()
    if not supplier_name or not email:
        raise HTTPException(status_code=400, detail="Supplier name and email are required")
    existing = (
        db.query(models.SupplierEmail)
        .filter(func.lower(models.SupplierEmail.supplier_name) == supplier_name.lower(), models.SupplierEmail.email == email)
        .first()
    )
    if existing:
        return existing
    row = models.SupplierEmail(supplier_name=supplier_name, email=email)
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


@app.delete("/api/supplier-emails/{email_id}")
def delete_supplier_email(
    email_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    row = db.query(models.SupplierEmail).filter(models.SupplierEmail.id == email_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="Email not found")
    db.delete(row)
    db.commit()
    return {"deleted": True}


def send_purchase_order_email(purchase_order: models.PurchaseOrder, recipients: List[str]) -> str:
    """Email the finalized PO to every address on file for the brand (plus
    the request's own supplier_email if set) in a single send. Host, user,
    port, and from-address all default to the company Gmail mailbox
    (initiative.lucknow@gmail.com), so on Render the only secret you need
    to set is SMTP_PASSWORD (a Gmail App Password) - see README for setup."""
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_user = os.getenv("SMTP_USER", "initiative.lucknow@gmail.com")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM", "initiative.lucknow@gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    if not (smtp_host and smtp_user and smtp_password and recipients):
        print("[PO email] Not sent: SMTP_PASSWORD not set (or no recipients).")
        return "Not sent: SMTP is not configured (set SMTP_PASSWORD on the host)."

    lines = [
        f"Purchase Order: {purchase_order.request_no}",
        f"Date: {purchase_order.request_date}",
        f"Brand: {purchase_order.brand_name or '-'}",
        f"Division: {purchase_order.division or '-'}",
        f"Delivery address: {purchase_order.delivery_address or '-'}",
        "",
        "Items:",
    ]
    for item in purchase_order.items:
        variant = f" ({item.variant})" if item.variant else ""
        lines.append(f"  - {item.product_name}{variant} x {item.quantity} {item.unit or 'Nos'}")
    if purchase_order.remarks:
        lines.append("")
        lines.append(f"Remarks: {purchase_order.remarks}")
    body = "\n".join(lines)

    message = MIMEText(body)
    message["Subject"] = f"Purchase Order {purchase_order.request_no} - {purchase_order.brand_name or ''}"
    message["From"] = smtp_from
    message["To"] = ", ".join(recipients)

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_from, recipients, message.as_string())
    except Exception as exc:  # noqa: BLE001 - surface any SMTP failure to the caller
        print(f"[PO email] SMTP send failed ({type(exc).__name__}): {exc}")
        return f"Not sent: email delivery failed ({type(exc).__name__}: {exc})."

    return f"Emailed to {len(recipients)} recipient(s)."


@app.post("/api/purchase-orders/{purchase_order_id}/send-email", response_model=schemas.SendPurchaseOrderEmailResult)
def send_purchase_order_email_endpoint(
    purchase_order_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    purchase_order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == purchase_order_id).first()
    if not purchase_order:
        raise HTTPException(status_code=404, detail="Purchase order request not found")
    if purchase_order.status not in {"Approved", "Ordered"}:
        raise HTTPException(
            status_code=400,
            detail="This purchase order must be Approved by Admin before it can be sent to the supplier.",
        )

    recipients = set()
    if purchase_order.brand_name:
        brand_rows = db.query(models.BrandSupplierEmail).filter(models.BrandSupplierEmail.brand_name == purchase_order.brand_name).all()
        recipients.update(row.email for row in brand_rows)
    if purchase_order.supplier_name:
        supplier_rows = (
            db.query(models.SupplierEmail)
            .filter(func.lower(models.SupplierEmail.supplier_name) == purchase_order.supplier_name.strip().lower())
            .all()
        )
        recipients.update(row.email for row in supplier_rows)
    if purchase_order.supplier_email:
        recipients.add(purchase_order.supplier_email.strip().lower())
    recipients = sorted(r for r in recipients if r)

    if not recipients:
        raise HTTPException(status_code=400, detail="No supplier emails on file for this brand or supplier yet. Add at least one first.")

    notification_status = send_purchase_order_email(purchase_order, recipients)
    return {"sent_to": recipients, "notification_status": notification_status}


@app.post("/api/purchase-orders", response_model=schemas.PurchaseOrderOut)
def create_purchase_order(
    payload: schemas.PurchaseOrderCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Add at least one purchase item")
    if any(not item.product_name.strip() or item.quantity <= 0 for item in payload.items):
        raise HTTPException(status_code=400, detail="Every item needs a product name and quantity greater than zero")

    purchase_order = models.PurchaseOrder(
        request_no=f"REQ-{payload.request_date.strftime('%Y%m%d')}-{uuid4().hex[:6].upper()}",
        request_date=payload.request_date,
        division=payload.division,
        branch_id=payload.branch_id,
        brand_name=payload.brand_name,
        supplier_name=payload.supplier_name,
        supplier_email=payload.supplier_email,
        supplier_address=payload.supplier_address,
        supplier_gstin=payload.supplier_gstin,
        delivery_address=payload.delivery_address,
        remarks=payload.remarks,
        status="Requested",
        submitted_by_user_id=current_user.id,
    )
    purchase_order.items = [models.PurchaseOrderItem(**item.dict()) for item in payload.items]
    db.add(purchase_order)
    db.commit()
    db.refresh(purchase_order)

    if payload.supplier_name:
        upsert_supplier_profile(db, payload.supplier_name, payload.supplier_address, payload.supplier_gstin)
        for email in (payload.supplier_emails or ([payload.supplier_email] if payload.supplier_email else [])):
            upsert_supplier_email(db, payload.supplier_name, email)

    notification_status = send_purchase_order_whatsapp_notification(purchase_order)
    return serialize_purchase_order(purchase_order, notification_status)


@app.get("/api/purchase-orders", response_model=List[schemas.PurchaseOrderOut])
def list_purchase_orders(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.PurchaseOrder)
    if not auth.has_admin_access(current_user) and current_user.role != "MISExecutive":
        query = query.filter(models.PurchaseOrder.submitted_by_user_id == current_user.id)
    purchase_orders = query.order_by(models.PurchaseOrder.created_date.desc()).all()
    return [serialize_purchase_order(item) for item in purchase_orders]


@app.get("/api/purchase-orders/{purchase_order_id}", response_model=schemas.PurchaseOrderOut)
def get_purchase_order(
    purchase_order_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    purchase_order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == purchase_order_id).first()
    if not purchase_order:
        raise HTTPException(status_code=404, detail="Purchase order request not found")
    if not can_access_purchase_order(current_user, purchase_order):
        raise HTTPException(status_code=403, detail="You can only view your own purchase requests")
    return serialize_purchase_order(purchase_order)


@app.patch("/api/purchase-orders/{purchase_order_id}/status", response_model=schemas.PurchaseOrderOut)
def update_purchase_order_status(
    purchase_order_id: int,
    payload: schemas.PurchaseOrderStatusUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    status_value = (payload.status or "").strip()
    if status_value not in PURCHASE_ORDER_STATUSES:
        raise HTTPException(status_code=400, detail=f"Status must be one of: {sorted(PURCHASE_ORDER_STATUSES)}")

    purchase_order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == purchase_order_id).first()
    if not purchase_order:
        raise HTTPException(status_code=404, detail="Purchase order request not found")

    if status_value != purchase_order.status:
        assert_status_transition_allowed(current_user, purchase_order, status_value)

    if status_value == "Approved" and purchase_order.status != "Approved":
        purchase_order.approved_by_user_id = current_user.id
        purchase_order.approved_date = datetime.utcnow()
    elif status_value in {"Rejected", "Requested"}:
        # Sent back for changes or turned down - clear any prior approval so
        # it has to go through Admin again before it can be Ordered.
        purchase_order.approved_by_user_id = None
        purchase_order.approved_date = None

    purchase_order.status = status_value
    purchase_order.busy_po_number = (payload.busy_po_number or "").strip() or None
    purchase_order.ordered_date = payload.ordered_date
    purchase_order.processing_notes = (payload.processing_notes or "").strip() or None

    # Admin/MIS can fill in or correct procurement details while processing
    # a category manager's request. Only touch fields that were actually sent.
    if payload.division is not None:
        purchase_order.division = payload.division or None
    if payload.branch_id is not None:
        purchase_order.branch_id = payload.branch_id
    if payload.brand_name is not None:
        purchase_order.brand_name = payload.brand_name or None
    if payload.supplier_name is not None:
        purchase_order.supplier_name = payload.supplier_name or None
    if payload.supplier_email is not None:
        purchase_order.supplier_email = payload.supplier_email or None
    if payload.supplier_address is not None:
        purchase_order.supplier_address = payload.supplier_address or None
    if payload.supplier_gstin is not None:
        purchase_order.supplier_gstin = payload.supplier_gstin or None
    if payload.delivery_address is not None:
        purchase_order.delivery_address = payload.delivery_address or None
    if payload.remarks is not None:
        purchase_order.remarks = payload.remarks or None

    if payload.items is not None:
        if not payload.items:
            raise HTTPException(status_code=400, detail="A purchase order needs at least one item")
        if any(not item.product_name.strip() or item.quantity <= 0 for item in payload.items):
            raise HTTPException(status_code=400, detail="Every item needs a product name and quantity greater than zero")
        purchase_order.items = [models.PurchaseOrderItem(**item.dict()) for item in payload.items]

    db.commit()
    db.refresh(purchase_order)

    if purchase_order.supplier_name:
        upsert_supplier_profile(db, purchase_order.supplier_name, purchase_order.supplier_address, purchase_order.supplier_gstin)
        for email in (payload.supplier_emails or ([purchase_order.supplier_email] if purchase_order.supplier_email else [])):
            upsert_supplier_email(db, purchase_order.supplier_name, email)

    return serialize_purchase_order(purchase_order)


@app.delete("/api/purchase-orders/{purchase_order_id}")
def delete_purchase_order(
    purchase_order_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    purchase_order = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.id == purchase_order_id).first()
    if not purchase_order:
        raise HTTPException(status_code=404, detail="Purchase order request not found")
    db.delete(purchase_order)
    db.commit()
    return {"message": "Purchase order request deleted"}


@app.post("/api/purchase-orders/mark-exported-to-busy")
def mark_purchase_orders_exported_to_busy(
    payload: schemas.MarkExportedToBusyRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
):
    """Called after downloading a batch export file for Busy's Import
    Vouchers feature, so the same requests aren't included in next time's
    export. Purely a bookkeeping flag on this side — it does not talk to
    Busy directly."""
    if not payload.purchase_order_ids:
        return {"updated": 0}
    now = datetime.utcnow()
    updated = (
        db.query(models.PurchaseOrder)
        .filter(models.PurchaseOrder.id.in_(payload.purchase_order_ids))
        .update({"exported_to_busy": True, "exported_to_busy_at": now}, synchronize_session=False)
    )
    db.commit()
    return {"updated": updated}


# ============================================================
# SCHEMES
# ============================================================

@app.post("/schemes", response_model=schemas.SchemeOut)
def create_scheme(
    scheme: schemas.SchemeCreate,
    db: Session = Depends(get_db),
    # Brand promoters/managers no longer get manual create rights - their
    # only path into the scheme table is "Attach Scheme Document"
    # (POST /schemes/upload-document), which always lands as a Draft for
    # an Admin to review.
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    scheme_code = (scheme.scheme_code or "").strip()
    if not scheme_code:
        # Auto-generate a unique code since the Scheme Maintenance form no
        # longer asks for one. Format: SCH-<epoch-milliseconds>.
        scheme_code = f"SCH-{int(datetime.utcnow().timestamp() * 1000)}"
    else:
        existing = (
            db.query(models.Scheme)
            .filter(models.Scheme.scheme_code == scheme_code)
            .first()
        )
        if existing:
            raise HTTPException(status_code=400, detail="Scheme code already exists")

    normalized_reward_type = normalize_reward_type(scheme.reward_type)
    normalized_offer_type = normalize_offer_type(scheme.offer_type)

    db_scheme = models.Scheme(
        scheme_code=scheme_code,
        scheme_name=scheme.scheme_name,
        brand_id=scheme.brand_id,
        category_id=scheme.category_id,
        subcategory_id=scheme.subcategory_id,
        product_id=scheme.product_id,
        variant_id=scheme.variant_id,
        offer_type=normalized_offer_type,
        offer_value=scheme.offer_value,
        calculation_method=scheme.calculation_method,
        start_date=scheme.start_date,
        end_date=scheme.end_date,
        min_qty=scheme.min_qty,
        max_qty=scheme.max_qty,
        applicable_branch_id=scheme.applicable_branch_id,
        applicable_customer=scheme.applicable_customer,
        applicable_dealer=scheme.applicable_dealer,
        circular_number=scheme.circular_number,
        remarks=scheme.remarks,
        reward_type=normalized_reward_type,
        reward_value=scheme.reward_value,
        reward_type_other=(scheme.reward_type_other or "").strip() or None,
        status=scheme.status,
    )
    db.add(db_scheme)
    db.commit()
    db.refresh(db_scheme)

    for cond in scheme.conditions:
        db.add(
            models.SchemeCondition(
                scheme_id=db_scheme.id,
                field_name=cond.field_name,
                operator=cond.operator,
                value=cond.value,
            )
        )

    for slab in scheme.slabs:
        db.add(
            models.SchemeSlab(
                scheme_id=db_scheme.id,
                min_quantity=slab.min_quantity,
                reward_per_unit=slab.reward_per_unit,
            )
        )

    db.commit()
    db.refresh(db_scheme)
    return db_scheme


@app.get("/schemes", response_model=List[schemas.SchemeOut])
def list_schemes(
    status: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(models.Scheme)
    if status:
        query = query.filter(models.Scheme.status == status)
    if not auth.has_admin_access(current_user):
        # Draft schemes hold whatever a document upload extracted and
        # haven't been reviewed yet. Only Admin should see those fields -
        # everyone else (including the promoter who attached the
        # document) only sees schemes once they're Active/Paused.
        query = query.filter(models.Scheme.status != "Draft")
    return query.all()


@app.get("/schemes/my-attachments")
def list_my_scheme_attachments(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Upload history for whoever attached the document. Deliberately
    excludes the extracted scheme fields (offer, reward value, target,
    etc.) - those stay hidden until an Admin reviews and activates the
    Draft. Non-admins only ever see their own uploads."""
    query = db.query(models.SchemeAttachment).order_by(models.SchemeAttachment.id.desc())
    if not auth.has_admin_access(current_user):
        query = query.filter(models.SchemeAttachment.uploaded_by_user_id == current_user.id)

    rows = []
    for attachment in query.limit(200).all():
        scheme = attachment.scheme
        brand = (
            db.query(models.Brand).filter(models.Brand.id == scheme.brand_id).first()
            if scheme and scheme.brand_id
            else None
        )
        uploader = (
            db.query(models.User).filter(models.User.id == attachment.uploaded_by_user_id).first()
            if attachment.uploaded_by_user_id
            else None
        )
        rows.append({
            "id": attachment.id,
            "scheme_id": attachment.scheme_id,
            "filename": attachment.original_filename,
            "brand": brand.name if brand else "Not matched yet",
            "uploaded_by": uploader.username if uploader else "",
            "uploaded_date": attachment.created_date.isoformat() if attachment.created_date else None,
            "review_status": "Reviewed" if scheme and scheme.status != "Draft" else "Pending review",
        })
    return rows


def apply_scheme_extraction(db: Session, db_scheme: "models.Scheme", attachment: "models.SchemeAttachment", extraction: dict) -> None:
    """Applies a Claude extraction result onto a Draft scheme + its
    attachment record. Shared by the upload endpoint (Admin uploads, which
    still extract immediately) and the Admin-triggered
    POST /schemes/{id}/extract endpoint (used for promoter uploads, which
    are deferred until an Admin runs OCR)."""
    attachment.extraction_status = extraction["status"]
    attachment.extraction_error = extraction.get("error")
    attachment.extraction_raw_json = json.dumps(extraction.get("data") or {})

    if extraction["status"] == "Extracted":
        data = extraction["data"]
        try:
            if data.get("brand_name") and db_scheme.brand_id is None:
                match = (
                    db.query(models.Brand)
                    .filter(func.lower(models.Brand.name) == str(data["brand_name"]).strip().lower())
                    .first()
                )
                if match:
                    db_scheme.brand_id = match.id

            if data.get("product_name"):
                product_match = (
                    db.query(models.Product)
                    .filter(func.lower(models.Product.name) == str(data["product_name"]).strip().lower())
                    .first()
                )
                if product_match:
                    db_scheme.product_id = product_match.id

            if data.get("scheme_name"):
                db_scheme.scheme_name = str(data["scheme_name"]).strip()[:200] or db_scheme.scheme_name
            if data.get("start_date"):
                db_scheme.start_date = parse_date_value(data["start_date"])
            if data.get("end_date"):
                db_scheme.end_date = parse_date_value(data["end_date"])
            if data.get("reward_type"):
                db_scheme.reward_type = normalize_reward_type(data["reward_type"])
            if data.get("reward_value") is not None:
                db_scheme.reward_value = parse_float_value(data.get("reward_value"), fallback=0.0)
            if data.get("min_qty") is not None:
                db_scheme.min_qty = int(parse_float_value(data.get("min_qty"), fallback=0.0))
            if data.get("max_qty"):
                db_scheme.max_qty = int(parse_float_value(data.get("max_qty"), fallback=0.0))
            if data.get("offer_type"):
                db_scheme.offer_type = normalize_offer_type(data["offer_type"])
            if data.get("circular_number"):
                db_scheme.circular_number = str(data["circular_number"])[:50]

            extracted_remarks = str(data.get("remarks") or "").strip()
            db_scheme.remarks = (
                (extracted_remarks or "Extracted from attached scheme document.")
                + " (Draft - review before activating.)"
            )[:255]

            for slab in data.get("slabs") or []:
                try:
                    db.add(models.SchemeSlab(
                        scheme_id=db_scheme.id,
                        min_quantity=int(slab.get("min_quantity") or 0),
                        reward_per_unit=float(slab.get("reward_per_unit") or 0),
                    ))
                except Exception:
                    continue
        except Exception as exc:
            attachment.extraction_status = "Failed"
            attachment.extraction_error = f"Extracted data didn't fit the scheme form: {exc}"


@app.post("/schemes/upload-document")
def upload_scheme_document(
    file: UploadFile = File(...),
    brand_id: Optional[int] = Query(None),
    scheme_name: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Any logged-in user can attach a scheme circular (image, PDF, or
    Excel). This creates a Draft scheme and saves the document. When an
    Admin uploads directly, Claude reads it immediately (Admin already
    reviews everything). When anyone else uploads, extraction
    is deferred - the document just sits in "Draft Schemes Pending
    Review" until an Admin clicks "Extract (OCR)" there. Either way, Admin
    reviews and hits Activate (existing PUT /schemes/{id}/activate) when
    it looks right."""
    filename = file.filename or "scheme_document"
    raw_bytes = file.file.read()
    if not raw_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    if current_user.role in {"BrandManager", "BrandPartner"}:
        allowed_brand_ids = [ub.brand_id for ub in current_user.brands]
        if not allowed_brand_ids:
            raise HTTPException(status_code=403, detail="Your account has no brand assigned yet. Ask an Admin to assign one.")
        if brand_id is None:
            brand_id = allowed_brand_ids[0]
        elif brand_id not in allowed_brand_ids:
            raise HTTPException(status_code=403, detail="You can only attach documents for your own assigned brand(s).")

    today = date.today()
    db_scheme = models.Scheme(
        scheme_code=f"SCH-{int(datetime.utcnow().timestamp() * 1000)}",
        scheme_name=(scheme_name or "").strip() or f"Pending review - {filename}",
        brand_id=brand_id,
        start_date=today,
        end_date=today,
        status="Draft",
        reward_type="Fixed",
        reward_value=0,
        offer_type="Backend",
        calculation_method="Fixed Amount",
        remarks="Awaiting extraction from attached document.",
    )
    db.add(db_scheme)
    db.commit()
    db.refresh(db_scheme)

    content_type = file.content_type or ""
    attachment = models.SchemeAttachment(
        scheme_id=db_scheme.id,
        original_filename=filename,
        content_type=content_type,
        file_size=len(raw_bytes),
        file_data=raw_bytes,
        uploaded_by_user_id=current_user.id,
        extraction_status="Pending",
    )
    db.add(attachment)
    db.commit()

    if auth.has_admin_access(current_user):
        # Admin uploads still extract right away, same as before.
        extraction = extract_scheme_from_document(db, filename, content_type, raw_bytes)
        apply_scheme_extraction(db, db_scheme, attachment, extraction)
        db.commit()
        db.refresh(db_scheme)
        db.refresh(attachment)

    return {
        "scheme_id": db_scheme.id,
        "scheme_code": db_scheme.scheme_code,
        "status": db_scheme.status,
        "extraction_status": attachment.extraction_status,
        "extraction_error": attachment.extraction_error,
        "message": (
            "Document attached and scheme fields pre-filled. Review and Activate when ready."
            if attachment.extraction_status == "Extracted"
            else "Document attached. An Admin will review it shortly."
            if not auth.has_admin_access(current_user)
            else "Document attached, but automatic extraction did not complete - fill the scheme fields manually before activating."
        ),
    }


@app.post("/schemes/{scheme_id}/extract")
def extract_scheme_document(
    scheme_id: int,
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    """Admin-triggered OCR/extraction for a Draft scheme's most recently
    attached document - used for promoter/brand-manager uploads, which no
    longer auto-extract at upload time."""
    db_scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
    if not db_scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")

    attachment = (
        db.query(models.SchemeAttachment)
        .filter(models.SchemeAttachment.scheme_id == scheme_id)
        .order_by(models.SchemeAttachment.id.desc())
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="No document attached to this scheme")

    extraction = extract_scheme_from_document(
        db, attachment.original_filename, attachment.content_type or "", attachment.file_data
    )
    apply_scheme_extraction(db, db_scheme, attachment, extraction)
    db.commit()
    db.refresh(db_scheme)
    db.refresh(attachment)

    return {
        "scheme_id": db_scheme.id,
        "extraction_status": attachment.extraction_status,
        "extraction_error": attachment.extraction_error,
        "message": (
            "Extraction complete. Review the fields and Activate when ready."
            if attachment.extraction_status == "Extracted"
            else "Automatic extraction did not complete - fill the scheme fields manually before activating."
        ),
    }


@app.get("/schemes/drafts")
def list_draft_schemes(
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    """Admin-only feed for the "Draft Schemes Pending Review" table -
    includes attachment/extraction status so the UI can decide whether to
    show "Extract (OCR)" or the normal Edit/Activate actions."""
    drafts = (
        db.query(models.Scheme)
        .filter(models.Scheme.status == "Draft")
        .order_by(models.Scheme.id.desc())
        .all()
    )

    rows = []
    for scheme in drafts:
        brand = db.query(models.Brand).filter(models.Brand.id == scheme.brand_id).first() if scheme.brand_id else None
        attachment = (
            db.query(models.SchemeAttachment)
            .filter(models.SchemeAttachment.scheme_id == scheme.id)
            .order_by(models.SchemeAttachment.id.desc())
            .first()
        )
        rows.append({
            "id": scheme.id,
            "scheme_name": scheme.scheme_name,
            "brand": brand.name if brand else "Not matched",
            "start_date": str(scheme.start_date) if scheme.start_date else "",
            "end_date": str(scheme.end_date) if scheme.end_date else "",
            "reward_type": scheme.reward_type,
            "reward_value": scheme.reward_value,
            "extraction_status": attachment.extraction_status if attachment else "Pending",
            "extraction_error": attachment.extraction_error if attachment else None,
            "filename": attachment.original_filename if attachment else None,
        })
    return rows


@app.get("/schemes/{scheme_id}/attachment")
def download_scheme_attachment(
    scheme_id: int,
    current_user: models.User = Depends(auth.require_roles("Admin", "BrandManager", "BrandPartner")),
    db: Session = Depends(get_db),
):
    attachment = (
        db.query(models.SchemeAttachment)
        .filter(models.SchemeAttachment.scheme_id == scheme_id)
        .order_by(models.SchemeAttachment.id.desc())
        .first()
    )
    if not attachment:
        raise HTTPException(status_code=404, detail="No document attached to this scheme")

    if current_user.role in {"BrandManager", "BrandPartner"}:
        scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
        allowed_brand_ids = [ub.brand_id for ub in current_user.brands]
        if not scheme or scheme.brand_id not in allowed_brand_ids:
            raise HTTPException(status_code=403, detail="You can only view documents for your own brand's schemes.")

    return Response(
        content=attachment.file_data,
        media_type=attachment.content_type or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{attachment.original_filename}"'},
    )


@app.put("/schemes/{scheme_id}", response_model=schemas.SchemeOut)
def update_scheme(
    scheme_id: int,
    scheme: schemas.SchemeCreate,
    db: Session = Depends(get_db),
    # Editing (including Draft review/correction) is Admin-only. Brand
    # promoters/managers only attach documents; they don't see or touch
    # the extracted fields.
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    db_scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
    if not db_scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")

    normalized_reward_type = normalize_reward_type(scheme.reward_type)
    normalized_offer_type = normalize_offer_type(scheme.offer_type)

    db_scheme.scheme_name = scheme.scheme_name
    db_scheme.brand_id = scheme.brand_id
    db_scheme.category_id = scheme.category_id
    db_scheme.subcategory_id = scheme.subcategory_id
    db_scheme.product_id = scheme.product_id
    db_scheme.variant_id = scheme.variant_id
    db_scheme.offer_type = normalized_offer_type
    db_scheme.offer_value = scheme.offer_value
    db_scheme.calculation_method = scheme.calculation_method
    db_scheme.start_date = scheme.start_date
    db_scheme.end_date = scheme.end_date
    db_scheme.min_qty = scheme.min_qty
    db_scheme.max_qty = scheme.max_qty
    db_scheme.applicable_branch_id = scheme.applicable_branch_id
    db_scheme.applicable_customer = scheme.applicable_customer
    db_scheme.applicable_dealer = scheme.applicable_dealer
    db_scheme.circular_number = scheme.circular_number
    db_scheme.remarks = scheme.remarks
    db_scheme.reward_type = normalized_reward_type
    db_scheme.reward_value = scheme.reward_value
    db_scheme.reward_type_other = (scheme.reward_type_other or "").strip() or None
    db_scheme.status = scheme.status

    # Replace conditions/slabs entirely with whatever was submitted.
    db_scheme.conditions.clear()
    db_scheme.slabs.clear()
    db.flush()

    for cond in scheme.conditions:
        db.add(
            models.SchemeCondition(
                scheme_id=db_scheme.id,
                field_name=cond.field_name,
                operator=cond.operator,
                value=cond.value,
            )
        )

    for slab in scheme.slabs:
        db.add(
            models.SchemeSlab(
                scheme_id=db_scheme.id,
                min_quantity=slab.min_quantity,
                reward_per_unit=slab.reward_per_unit,
            )
        )

    db.commit()
    db.refresh(db_scheme)
    return db_scheme


@app.put("/schemes/{scheme_id}/pause")
def pause_scheme(
    scheme_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
    if not scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")
    scheme.status = "Paused"
    db.commit()
    return {"message": f"Scheme {scheme_id} paused"}


@app.put("/schemes/{scheme_id}/activate")
def activate_scheme(
    scheme_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
    if not scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")
    scheme.status = "Active"
    db.commit()
    return {"message": f"Scheme {scheme_id} activated"}


@app.delete("/schemes/{scheme_id}")
def delete_scheme(
    scheme_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    scheme = db.query(models.Scheme).filter(models.Scheme.id == scheme_id).first()
    if not scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")

    existing_claims = (
        db.query(models.ClaimHeader.id)
        .filter(models.ClaimHeader.scheme_id == scheme_id)
        .count()
    )
    if existing_claims:
        raise HTTPException(
            status_code=400,
            detail="This scheme has claims linked to it and can't be deleted. Pause it instead.",
        )

    # Delete child rows explicitly and in this order (attachments/slabs/
    # conditions all have a NOT NULL scheme_id). We don't rely on the
    # relationship cascade="all, delete-orphan" alone here — depending on
    # what's already loaded in the session, SQLAlchemy can try to detach a
    # child by setting its scheme_id to NULL instead of deleting the row,
    # which fails against the NOT NULL constraint. Bulk-deleting by query
    # issues a direct DELETE for each table and avoids that entirely.
    db.query(models.SchemeAttachment).filter(models.SchemeAttachment.scheme_id == scheme_id).delete(synchronize_session=False)
    db.query(models.SchemeSlab).filter(models.SchemeSlab.scheme_id == scheme_id).delete(synchronize_session=False)
    db.query(models.SchemeCondition).filter(models.SchemeCondition.scheme_id == scheme_id).delete(synchronize_session=False)
    db.delete(scheme)
    db.commit()
    return {"message": f"Scheme {scheme_id} deleted"}


# ============================================================
# SALES (this is what TRIGGERS the scheme engine)
# ============================================================

@app.post("/sales", response_model=schemas.SaleOut)
def create_sale(
    sale: schemas.SaleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    allowed_roles = {"Admin", "Owner", "HR", "StoreManager", "CategoryManager", "BrandManager", "BrandPartner", "Super Admin", "Management", "Branch Manager", "Sales Executive", "Scheme Manager"}
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="You are not allowed to create sales")

    if current_user.role in {"StoreManager", "Branch Manager"}:
        if current_user.store_id is None or sale.store_id != current_user.store_id:
            raise HTTPException(status_code=403, detail="You can only create sales for your assigned branch")

    if current_user.role == "CategoryManager":
        category_codes = category_codes_for_user(current_user)
        if not category_codes:
            raise HTTPException(status_code=403, detail="You are not assigned to a category")
        sale_category = db.query(models.Category).filter(models.Category.id == sale.category_id).first()
        if not sale_category or sale_category.code not in category_codes:
            raise HTTPException(status_code=403, detail="You can only create sales for your assigned category")

    if current_user.role in {"BrandManager", "BrandPartner", "Scheme Manager"}:
        brand_ids = [ub.brand_id for ub in current_user.brands]
        if sale.brand_id not in brand_ids:
            raise HTTPException(status_code=403, detail="You can only create sales for your assigned brands")

    existing = (
        db.query(models.Sale)
        .filter(models.Sale.invoice_no == sale.invoice_no)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Invoice number already exists")

    def quantize_money(raw_text: Optional[str], numeric_value: Optional[float], field_name: str):
        if raw_text is not None and str(raw_text).strip() != "":
            cleaned = str(raw_text).strip()
            try:
                dec = Decimal(cleaned)
            except Exception:
                raise HTTPException(status_code=400, detail=f"{field_name} must be a valid number")
            dec = dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return float(dec), cleaned

        if numeric_value is None:
            return None, None

        dec = Decimal(str(numeric_value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        fallback_text = format(dec, "f").rstrip("0").rstrip(".")
        if not fallback_text:
            fallback_text = "0"
        return float(dec), fallback_text

    # Keep money fields stable at 2 decimals so users see the same values they entered.
    sale_value_num, sale_value_exact = quantize_money(sale.sale_value_exact, sale.sale_value, "sale_value")
    scheme_amount_num, scheme_amount_exact = quantize_money(sale.scheme_amount_exact, sale.scheme_amount, "scheme_amount")
    upi_amount_num, upi_amount_exact = quantize_money(sale.upi_scheme_amount_exact, sale.upi_scheme_amount, "upi_scheme_amount")
    backend_amount_num, backend_amount_exact = quantize_money(sale.backend_scheme_amount_exact, sale.backend_scheme_amount, "backend_scheme_amount")

    sale.sale_value = sale_value_num
    sale.sale_value_exact = sale_value_exact
    sale.scheme_amount = scheme_amount_num
    sale.scheme_amount_exact = scheme_amount_exact
    sale.upi_scheme_amount = upi_amount_num
    sale.upi_scheme_amount_exact = upi_amount_exact
    sale.backend_scheme_amount = backend_amount_num
    sale.backend_scheme_amount_exact = backend_amount_exact

    db_sale = models.Sale(**sale.dict())
    db.add(db_sale)
    db.commit()
    db.refresh(db_sale)

    scheme_engine.evaluate_sale_against_schemes(db, db_sale)

    return db_sale


@app.put("/sales/{sale_id}", response_model=schemas.SaleOut)
def update_sale(
    sale_id: int,
    sale: schemas.SaleCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if not auth.has_admin_access(current_user) and current_user.role not in ("MISExecutive", "CategoryManager"):
        raise HTTPException(status_code=403, detail="You are not allowed to edit sales")

    db_sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not db_sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    # Category Manager can only edit sales already inside their own assigned
    # category (same scope rule used for "Sales in your scope" viewing and
    # for delete), and can't use an edit to move a sale into someone else's
    # category either.
    if current_user.role == "CategoryManager":
        if not can_user_access_sale(db, current_user, db_sale):
            raise HTTPException(status_code=403, detail="You can only edit sales in your access scope")
        sale_category = db.query(models.Category).filter(models.Category.id == sale.category_id).first()
        if not sale_category or sale_category.code not in category_codes_for_user(current_user):
            raise HTTPException(status_code=403, detail="You can only edit sales into your assigned category")

    duplicate_invoice = (
        db.query(models.Sale)
        .filter(models.Sale.invoice_no == sale.invoice_no, models.Sale.id != sale_id)
        .first()
    )
    if duplicate_invoice:
        raise HTTPException(status_code=400, detail="Invoice number already exists")

    def quantize_money(raw_text: Optional[str], numeric_value: Optional[float], field_name: str):
        if raw_text is not None and str(raw_text).strip() != "":
            cleaned = str(raw_text).strip()
            try:
                dec = Decimal(cleaned)
            except Exception:
                raise HTTPException(status_code=400, detail=f"{field_name} must be a valid number")
            dec = dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return float(dec), cleaned

        if numeric_value is None:
            return None, None

        dec = Decimal(str(numeric_value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        fallback_text = format(dec, "f").rstrip("0").rstrip(".")
        if not fallback_text:
            fallback_text = "0"
        return float(dec), fallback_text

    sale_value_num, sale_value_exact = quantize_money(sale.sale_value_exact, sale.sale_value, "sale_value")
    scheme_amount_num, scheme_amount_exact = quantize_money(sale.scheme_amount_exact, sale.scheme_amount, "scheme_amount")
    upi_amount_num, upi_amount_exact = quantize_money(sale.upi_scheme_amount_exact, sale.upi_scheme_amount, "upi_scheme_amount")
    backend_amount_num, backend_amount_exact = quantize_money(sale.backend_scheme_amount_exact, sale.backend_scheme_amount, "backend_scheme_amount")

    sale.sale_value = sale_value_num
    sale.sale_value_exact = sale_value_exact
    sale.scheme_amount = scheme_amount_num
    sale.scheme_amount_exact = scheme_amount_exact
    sale.upi_scheme_amount = upi_amount_num
    sale.upi_scheme_amount_exact = upi_amount_exact
    sale.backend_scheme_amount = backend_amount_num
    sale.backend_scheme_amount_exact = backend_amount_exact

    for field_name, value in sale.dict().items():
        setattr(db_sale, field_name, value)

    db.commit()
    db.refresh(db_sale)
    return db_sale


@app.get("/sales", response_model=List[schemas.SaleOut])
def list_sales(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return get_sales_for_user(db, current_user)


@app.delete("/sales/{sale_id}")
def delete_sale(
    sale_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    if not auth.has_admin_access(current_user):
        raise HTTPException(status_code=403, detail="You are not allowed to delete sales")

    sale = db.query(models.Sale).filter(models.Sale.id == sale_id).first()
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")

    if not can_user_access_sale(db, current_user, sale):
        raise HTTPException(status_code=403, detail="You can only delete sales in your access scope")

    claim_ids = [cid for (cid,) in db.query(models.ClaimHeader.id).filter(models.ClaimHeader.sale_id == sale.id).all()]
    if claim_ids:
        db.query(models.ClaimStatusHistory).filter(models.ClaimStatusHistory.claim_id.in_(claim_ids)).delete(synchronize_session=False)
        db.query(models.ClaimHeader).filter(models.ClaimHeader.id.in_(claim_ids)).delete(synchronize_session=False)

    db.delete(sale)
    db.commit()
    return {"message": f"Sale {sale_id} deleted successfully"}


@app.get("/dashboard-stats")
def dashboard_stats(db: Session = Depends(get_db)):
    from datetime import date
    today = date.today()

    active_schemes = db.query(models.Scheme).filter(models.Scheme.status == "Active").count()
    expired_schemes = db.query(models.Scheme).filter(models.Scheme.status != "Active").count()
    todays_sales = db.query(models.Sale).filter(models.Sale.sale_date == today).count()
    pending_claims = db.query(models.ClaimHeader).filter(models.ClaimHeader.status == "Draft").count()
    approved_claims = db.query(models.ClaimHeader).filter(models.ClaimHeader.status == "Approved").count()
    rejected_claims = db.query(models.ClaimHeader).filter(models.ClaimHeader.status == "Rejected").count()
    received_claims = db.query(models.ClaimHeader).filter(models.ClaimHeader.status == "Received").count()
    total_claim_amount = sum(claim.claim_amount for claim in db.query(models.ClaimHeader).all())
    pending_amount = sum(claim.claim_amount for claim in db.query(models.ClaimHeader).filter(models.ClaimHeader.status.in_(["Draft", "Pending", "Submitted"])).all())

    return {
        "active_schemes": active_schemes,
        "expired_schemes": expired_schemes,
        "todays_sales": todays_sales,
        "eligible_sales": db.query(models.ClaimHeader).count(),
        "total_claim_amount": round(total_claim_amount, 2),
        "pending_claims": pending_claims,
        "approved_claims": approved_claims,
        "rejected_claims": rejected_claims,
        "received_claims": received_claims,
        "pending_amount": round(pending_amount, 2),
    }


@app.post("/admin/interval-sales/upload")
def upload_interval_sales_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
    db: Session = Depends(get_db),
):
    filename = file.filename or "uploaded_file"
    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""
    raw_content = file.file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    parsed_rows = parse_tabular_rows(ext, raw_content)
    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No readable sales rows found. Ensure columns include Date, Vch No, Item, Qty, Sales Amt, Cost Amt, Profit/Loss, Profit %")

    # First pass: parse every row's fields (without inserting yet) so we know
    # the full date range covered by this file up front.
    prepared_rows = []
    skipped_rows = []
    for idx, row in enumerate(parsed_rows, start=1):
        try:
            sale_date = parse_date_value(row.get("sale_date"))
            vch_no = str(row.get("vch_no") or "").strip() or None
            account = str(row.get("account") or "").strip() or None
            item = str(row.get("item") or "").strip() or None
            qty = parse_float_value(row.get("qty"), fallback=0.0)
            unit = str(row.get("unit") or "").strip() or None
            sales_amt = parse_float_value(row.get("sales_amt"), fallback=0.0)
            cost_amt = parse_float_value(row.get("cost_amt"), fallback=0.0)
            profit_loss = parse_float_value(row.get("profit_loss"), fallback=sales_amt - cost_amt)
            profit_percent = parse_float_value(
                row.get("profit_percent"),
                fallback=((profit_loss / sales_amt) * 100.0 if sales_amt else 0.0),
            )
            prepared_rows.append({
                "sale_date": sale_date, "vch_no": vch_no, "account": account, "item": item,
                "qty": qty, "unit": unit, "sales_amt": sales_amt, "cost_amt": cost_amt,
                "profit_loss": profit_loss, "profit_percent": profit_percent,
            })
        except Exception as exc:
            skipped_rows.append({"row": idx, "reason": str(exc)})

    # Re-uploading a file that covers dates already in the system (e.g. a
    # corrected export re-sent after fixing a cost error) used to just pile
    # the new rows on top of the old ones with no de-duplication, silently
    # double-counting/merging stale + corrected line items in every report
    # that reads this table (Daily Profitability, Scheme-Matched Sales,
    # Interval Sales Analytics). Replace only the dates this file actually
    # covers - other dates already stored are left untouched.
    dates_in_file = [r["sale_date"] for r in prepared_rows if r["sale_date"]]
    replaced_count = 0
    if dates_in_file:
        replaced_count = (
            db.query(models.IntervalSaleUpload)
            .filter(models.IntervalSaleUpload.sale_date >= min(dates_in_file))
            .filter(models.IntervalSaleUpload.sale_date <= max(dates_in_file))
            .delete()
        )

    inserted_count = 0
    for r in prepared_rows:
        db.add(
            models.IntervalSaleUpload(
                sale_date=r["sale_date"],
                vch_no=r["vch_no"],
                account=r["account"],
                item=r["item"],
                qty=r["qty"],
                unit=r["unit"],
                sales_amt=r["sales_amt"],
                cost_amt=r["cost_amt"],
                profit_loss=r["profit_loss"],
                profit_percent=r["profit_percent"],
                source_file=filename,
                uploaded_by=current_user.id,
            )
        )
        inserted_count += 1

    db.commit()

    return {
        "message": "File processed successfully",
        "file_name": filename,
        "inserted": inserted_count,
        "skipped": len(skipped_rows),
        "replaced": replaced_count,
        "errors_preview": skipped_rows[:10],
    }


@app.delete("/admin/interval-sales/clear")
def clear_interval_sales_data(
    current_user: models.User = Depends(auth.require_roles("Admin", "MISExecutive")),
    db: Session = Depends(get_db),
):
    deleted_count = db.query(models.IntervalSaleUpload).delete()
    db.commit()
    return {"message": "All uploaded interval sales data cleared", "deleted": deleted_count}


@app.get("/admin/interval-sales/summary")
def interval_sales_summary(
    interval: str = Query("daily", pattern="^(daily|weekly|monthly|custom)$"),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    query = db.query(models.IntervalSaleUpload)
    if start_date:
        query = query.filter(models.IntervalSaleUpload.sale_date >= start_date)
    if end_date:
        query = query.filter(models.IntervalSaleUpload.sale_date <= end_date)

    rows = query.order_by(models.IntervalSaleUpload.sale_date.asc()).all()
    data = build_interval_analytics(rows, interval="daily" if interval == "custom" else interval)
    data["filters"] = {
        "start_date": start_date.isoformat() if start_date else None,
        "end_date": end_date.isoformat() if end_date else None,
        "requested_interval": interval,
    }
    data["sources"] = [
        {
            "file": file_name,
            "rows": count,
        }
        for file_name, count in (
            db.query(models.IntervalSaleUpload.source_file, text("COUNT(*)"))
            .group_by(models.IntervalSaleUpload.source_file)
            .order_by(text("COUNT(*) DESC"))
            .limit(5)
            .all()
        )
    ]
    return data


@app.get("/admin/interval-sales/records")
def interval_sales_records(
    limit: int = Query(200, ge=1, le=2000),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    query = db.query(models.IntervalSaleUpload)
    if start_date:
        query = query.filter(models.IntervalSaleUpload.sale_date >= start_date)
    if end_date:
        query = query.filter(models.IntervalSaleUpload.sale_date <= end_date)

    rows = (
        query.order_by(models.IntervalSaleUpload.sale_date.desc(), models.IntervalSaleUpload.id.desc())
        .limit(limit)
        .all()
    )

    return [
        {
            "id": row.id,
            "sale_date": row.sale_date,
            "vch_no": row.vch_no,
            "account": row.account,
            "item": row.item,
            "qty": row.qty,
            "unit": row.unit,
            "sales_amt": row.sales_amt,
            "cost_amt": row.cost_amt,
            "profit_loss": row.profit_loss,
            "profit_percent": row.profit_percent,
            "source_file": row.source_file,
            "created_date": row.created_date,
        }
        for row in rows
    ]


@app.get("/admin/interval-sales/scheme-matches")
def interval_sales_scheme_matches(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: models.User = Depends(auth.require_roles("Admin", "Accounts", "MISExecutive")),
    db: Session = Depends(get_db),
):
    """Feeds 'Sales in your scope': out of the uploaded profitability report
    (Interval Sales Analytics Upload), return only the rows whose item and
    sale date fall inside an Active scheme from Scheme Maintenance, with the
    backend claim amount computed for each - so Admin can see exactly which
    Busy sales are scheme-eligible without re-typing anything."""
    filters = {
        "start_date": start_date.isoformat() if start_date else None,
        "end_date": end_date.isoformat() if end_date else None,
    }
    empty_result = {"matches": [], "totals": {"records": 0, "qty": 0, "sales_amt": 0, "backend_amount": 0}, "filters": filters}

    sales_query = db.query(models.IntervalSaleUpload)
    if start_date:
        sales_query = sales_query.filter(models.IntervalSaleUpload.sale_date >= start_date)
    if end_date:
        sales_query = sales_query.filter(models.IntervalSaleUpload.sale_date <= end_date)
    sale_rows = sales_query.all()
    if not sale_rows:
        return empty_result

    active_schemes = db.query(models.Scheme).filter(models.Scheme.status == "Active").all()
    if not active_schemes:
        return empty_result

    # For each scheme, the exact product-name set it applies to: its own
    # product if one is set, else every product under its brand.
    scheme_product_names = {}
    for scheme in active_schemes:
        names = set()
        if scheme.product_id:
            product = db.query(models.Product).filter(models.Product.id == scheme.product_id).first()
            if product:
                names.add(product.name.strip().lower())
        elif scheme.brand_id:
            for product in db.query(models.Product).filter(models.Product.brand_id == scheme.brand_id).all():
                names.add(product.name.strip().lower())
        scheme_product_names[scheme.id] = names

    matches = []
    for row in sale_rows:
        item_name = (row.item or "").strip().lower()
        if not item_name:
            continue
        for scheme in active_schemes:
            if not (scheme.start_date <= row.sale_date <= scheme.end_date):
                continue
            if item_name not in scheme_product_names.get(scheme.id, set()):
                continue

            backend_amount = _calculate_reward_for_interval_row(scheme, row)
            if backend_amount <= 0:
                continue

            matches.append({
                "sale_id": row.id,
                "sale_date": row.sale_date.isoformat(),
                "vch_no": row.vch_no,
                "account": row.account,
                "item": row.item,
                "qty": row.qty,
                "unit": row.unit,
                "sales_amt": row.sales_amt,
                "scheme_id": scheme.id,
                "scheme_code": scheme.scheme_code,
                "scheme_name": scheme.scheme_name,
                "reward_type": scheme.reward_type,
                "backend_amount": round(backend_amount, 2),
            })
            break  # a sale counts once, against its first matching scheme

    totals = {
        "records": len(matches),
        "qty": round(sum(m["qty"] or 0 for m in matches), 2),
        "sales_amt": round(sum(m["sales_amt"] or 0 for m in matches), 2),
        "backend_amount": round(sum(m["backend_amount"] for m in matches), 2),
    }

    return {
        "matches": sorted(matches, key=lambda m: m["sale_date"], reverse=True),
        "totals": totals,
        "filters": filters,
    }




def build_daily_profitability_workbook(merged_items: list, period_label: str) -> bytes:
    """Builds the styled, formula-driven 'Sales Control Report' workbook
    (same layout as the company's DAILY_PROFITIABILITY.xlsx template - one
    section per category with Store/Item/Sale/Purchase/Margin/PL% columns
    and a Gross Profit subtotal), plus a Dashboard sheet with KPIs, a
    category chart, and Top Profitable / Loss-Making Items. Returns the
    .xlsx file as bytes, ready to stream back in an HTTP response."""
    import openpyxl as _openpyxl
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment, Border as _Border, Side as _Side
    from openpyxl.chart import BarChart as _BarChart, Reference as _Reference
    from collections import defaultdict as _defaultdict

    COMPANY = "Initiative Data Systems Pvt Ltd"
    FONT_TITLE = _Font(name='Times New Roman', size=18, bold=True)
    FONT_SECTION = _Font(name='Times New Roman', size=14, bold=True)
    FONT_HEADER = _Font(name='Calibri', size=11, bold=True, underline='single')
    FONT_DATA = _Font(name='Calibri', size=11, bold=True)
    FONT_DATA_LOSS = _Font(name='Calibri', size=11, bold=True, color='C0392B')
    FILL_YELLOW = _PatternFill('solid', fgColor='FFFFFF00')
    THIN = _Side(style='thin')
    BORDER_ALL = _Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUMFMT_ACC = '_ * #,##0_ ;_ * \\-#,##0_ ;_ * "-"??_ ;_ @_ '
    NUMFMT_PCT = '0.0%'
    COL_WIDTHS = {'A': 5.55, 'B': 20.0, 'C': 42.89, 'D': 12.33, 'E': 18.11,
                  'F': 9.0, 'G': 8.44, 'H': 9.0, 'I': 11.33, 'J': 9.44}
    HEADERS = ['S. No', 'Vch. No.', 'Item Name', 'Sale Amount', 'Purchase Price',
               'Upfront Margin', 'Backend', 'Margin', 'PL %', 'Narration']

    by_cat = _defaultdict(list)
    for m in merged_items:
        by_cat[m['category']].append(m)

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Report'
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    r = 1
    gross_profit_rows = []
    for cat_label in DP_CATEGORIES:
        items = sorted(by_cat.get(cat_label, []), key=lambda x: x['item'].lower())
        if not items:
            continue

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1, value=COMPANY)
        c.font = FONT_TITLE
        c.alignment = _Alignment(horizontal='center')
        ws.row_dimensions[r].height = 22.8
        r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        c = ws.cell(row=r, column=1, value=f"{cat_label} Sales Control Report {period_label}")
        c.font = FONT_SECTION
        c.alignment = _Alignment(horizontal='center')
        ws.row_dimensions[r].height = 17.4
        r += 1

        ws.row_dimensions[r].height = 28.8
        for ci, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = FONT_HEADER
            cell.fill = FILL_YELLOW
            cell.border = BORDER_ALL
            if ci in (4, 5):
                cell.alignment = _Alignment(horizontal='right', vertical='center')
            elif ci in (6, 7, 8, 9):
                cell.alignment = _Alignment(horizontal='center', vertical='center', wrap_text=True)
            elif ci == 10:
                cell.alignment = _Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = _Alignment(horizontal='left', vertical='center')
        r += 1

        first_data_row = r
        for i, m in enumerate(items, start=1):
            row_font = FONT_DATA_LOSS if m.get('margin', m['sale'] - m['cost']) < 0 else FONT_DATA
            ws.cell(row=r, column=1, value=i).font = row_font
            cell = ws.cell(row=r, column=2, value=m['vch']); cell.font = row_font; cell.alignment = _Alignment(horizontal='center')
            cell = ws.cell(row=r, column=3, value=m['item']); cell.font = row_font; cell.alignment = _Alignment(horizontal='left')
            cell = ws.cell(row=r, column=4, value=round(m['sale'], 2)); cell.font = row_font; cell.alignment = _Alignment(horizontal='right'); cell.number_format = NUMFMT_ACC
            cell = ws.cell(row=r, column=5, value=round(m['cost'], 2)); cell.font = row_font; cell.alignment = _Alignment(horizontal='right'); cell.number_format = NUMFMT_ACC
            cell = ws.cell(row=r, column=6, value=f'=+D{r}-E{r}'); cell.font = row_font; cell.number_format = NUMFMT_ACC
            ws.cell(row=r, column=7).number_format = NUMFMT_ACC
            cell = ws.cell(row=r, column=8, value=f'=+G{r}+F{r}'); cell.font = row_font; cell.number_format = NUMFMT_ACC
            cell = ws.cell(row=r, column=9, value=f'=+H{r}/E{r}'); cell.font = row_font; cell.number_format = NUMFMT_PCT
            cell = ws.cell(row=r, column=10, value=m.get('note') or ''); cell.font = row_font; cell.alignment = _Alignment(horizontal='left')
            for ci in range(1, 11):
                ws.cell(row=r, column=ci).border = BORDER_ALL
            r += 1
        last_data_row = r - 1

        ws.cell(row=r, column=3, value=f'Gross Profit {cat_label}')
        ws.cell(row=r, column=4, value=f'=SUM(D{first_data_row}:D{last_data_row})')
        ws.cell(row=r, column=5, value=f'=SUM(E{first_data_row}:E{last_data_row})')
        ws.cell(row=r, column=6, value=f'=SUM(F{first_data_row}:F{last_data_row})')
        ws.cell(row=r, column=8, value=f'=SUM(H{first_data_row}:H{last_data_row})')
        ws.cell(row=r, column=9, value=f'=+H{r}/E{r}')
        for ci in range(1, 11):
            cell = ws.cell(row=r, column=ci)
            cell.font = FONT_DATA
            cell.fill = FILL_YELLOW
            if ci == 3:
                cell.number_format = '@'
            elif ci == 9:
                cell.number_format = NUMFMT_PCT
            elif ci in (4, 5, 6, 7, 8):
                cell.number_format = NUMFMT_ACC
        gross_profit_rows.append((cat_label, r))
        r += 2

    if gross_profit_rows:
        ws.cell(row=r, column=3, value='Total Gross Profit')
        for col_idx, col_letter in ((4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'), (8, 'H')):
            formula = '+' + '+'.join(f'{col_letter}{gr}' for _, gr in gross_profit_rows)
            ws.cell(row=r, column=col_idx, value=f'={formula}')
        ws.cell(row=r, column=9, value=f'=+H{r}/E{r}')
        for ci in range(1, 11):
            cell = ws.cell(row=r, column=ci)
            cell.font = FONT_DATA
            if ci == 3:
                cell.number_format = '@'
            elif ci == 9:
                cell.number_format = NUMFMT_PCT
            elif ci in (4, 5, 6, 7, 8):
                cell.number_format = NUMFMT_ACC

    # ---- Dashboard sheet ----
    dash = wb.create_sheet('Dashboard', 0)
    FONT_KPI_LABEL = _Font(name='Calibri', size=10, bold=True, color='555555')
    FONT_KPI_VALUE = _Font(name='Calibri', size=16, bold=True)
    FONT_D_HEADER = _Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    FONT_D_DATA = _Font(name='Calibri', size=10)
    FONT_D_LOSS = _Font(name='Calibri', size=10, color='C0392B', bold=True)
    FILL_HEADER = _PatternFill('solid', fgColor='1F2D3F')
    FILL_KPI = _PatternFill('solid', fgColor='F5F7FA')
    FILL_LOSS_HEADER = _PatternFill('solid', fgColor='C0392B')
    FILL_PROFIT_HEADER = _PatternFill('solid', fgColor='027A48')
    THIN2 = _Side(style='thin', color='D6D6D6')
    BORDER2 = _Border(left=THIN2, right=THIN2, top=THIN2, bottom=THIN2)

    for col, w in {'A': 4, 'B': 18, 'C': 42, 'D': 14, 'E': 14, 'F': 14, 'G': 4,
                    'H': 10, 'I': 40, 'J': 14, 'K': 14, 'L': 14}.items():
        dash.column_dimensions[col].width = w

    rr = 1
    dash.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    c = dash.cell(row=rr, column=1, value=COMPANY); c.font = FONT_TITLE; c.alignment = _Alignment(horizontal='center')
    rr += 1
    dash.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    c = dash.cell(row=rr, column=1, value=f'Daily Profitability Dashboard - {period_label}')
    c.font = FONT_SECTION; c.alignment = _Alignment(horizontal='center')
    rr += 2

    total_sale = sum(m['sale'] for m in merged_items)
    total_cost = sum(m['cost'] for m in merged_items)
    total_margin = total_sale - total_cost
    margin_pct = (total_margin / total_cost) if total_cost else 0
    loss_items_all = [m for m in merged_items if m['margin'] < 0]

    kpis = [
        ('Total Sales', total_sale, NUMFMT_ACC), ('Total Cost', total_cost, NUMFMT_ACC),
        ('Total Margin', total_margin, NUMFMT_ACC), ('Margin %', margin_pct, NUMFMT_PCT),
        ('Line Items', len(merged_items), '0'), ('Loss-Making Items', len(loss_items_all), '0'),
    ]
    kpi_col = 1
    kpi_row = rr
    for label, val, fmt in kpis:
        cl = dash.cell(row=kpi_row, column=kpi_col, value=label); cl.font = FONT_KPI_LABEL; cl.fill = FILL_KPI; cl.alignment = _Alignment(horizontal='center')
        cv = dash.cell(row=kpi_row + 1, column=kpi_col, value=val); cv.font = FONT_KPI_VALUE; cv.fill = FILL_KPI; cv.number_format = fmt; cv.alignment = _Alignment(horizontal='center')
        for rrr in (kpi_row, kpi_row + 1):
            dash.cell(row=rrr, column=kpi_col).border = BORDER2
        kpi_col += 2
    rr = kpi_row + 3

    by_cat = _defaultdict(lambda: {'sale': 0, 'cost': 0, 'margin': 0})
    for m in merged_items:
        by_cat[m['category']]['sale'] += m['sale']
        by_cat[m['category']]['cost'] += m['cost']
        by_cat[m['category']]['margin'] += m['margin']

    cat_table_row = rr
    dash.cell(row=rr, column=1, value='Category').font = FONT_D_HEADER; dash.cell(row=rr, column=1).fill = FILL_HEADER
    dash.cell(row=rr, column=2, value='Sales').font = FONT_D_HEADER; dash.cell(row=rr, column=2).fill = FILL_HEADER
    dash.cell(row=rr, column=3, value='Margin').font = FONT_D_HEADER; dash.cell(row=rr, column=3).fill = FILL_HEADER
    rr += 1
    cat_first = rr
    for cat in DP_CATEGORIES:
        d = by_cat.get(cat, {'sale': 0, 'cost': 0, 'margin': 0})
        dash.cell(row=rr, column=1, value=cat).font = FONT_D_DATA
        dash.cell(row=rr, column=2, value=round(d['sale'], 2)).number_format = NUMFMT_ACC
        dash.cell(row=rr, column=3, value=round(d['margin'], 2)).number_format = NUMFMT_ACC
        for ci in (1, 2, 3):
            dash.cell(row=rr, column=ci).border = BORDER2
        rr += 1
    cat_last = rr - 1

    chart = _BarChart()
    chart.type = 'col'
    chart.title = 'Sales & Margin by Category'
    chart.y_axis.title = 'Amount (Rs.)'
    chart.height = 7
    chart.width = 14
    data = _Reference(dash, min_col=2, max_col=3, min_row=cat_table_row, max_row=cat_last)
    cats = _Reference(dash, min_col=1, min_row=cat_first, max_row=cat_last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    dash.add_chart(chart, f'E{cat_table_row}')

    rr = cat_last + 12

    def write_item_table(start_row, title, item_rows, header_fill, is_loss):
        rrr = start_row
        dash.merge_cells(start_row=rrr, start_column=1, end_row=rrr, end_column=6)
        c = dash.cell(row=rrr, column=1, value=title); c.font = _Font(name='Calibri', size=12, bold=True)
        rrr += 1
        for ci, h in enumerate(['#', 'Vch. No.', 'Item Name', 'Sale Amount', 'Margin', 'PL %'], start=1):
            cell = dash.cell(row=rrr, column=ci, value=h)
            cell.font = FONT_D_HEADER; cell.fill = header_fill; cell.border = BORDER2
            cell.alignment = _Alignment(horizontal='center')
        rrr += 1
        for i, m in enumerate(item_rows, start=1):
            dash.cell(row=rrr, column=1, value=i).font = FONT_D_DATA
            dash.cell(row=rrr, column=2, value=m['vch']).font = FONT_D_DATA
            dash.cell(row=rrr, column=3, value=m['item']).font = FONT_D_DATA
            cell = dash.cell(row=rrr, column=4, value=round(m['sale'], 2)); cell.font = FONT_D_DATA; cell.number_format = NUMFMT_ACC
            cell = dash.cell(row=rrr, column=5, value=round(m['margin'], 2))
            cell.font = FONT_D_LOSS if is_loss else FONT_D_DATA
            cell.number_format = NUMFMT_ACC
            cell = dash.cell(row=rrr, column=6, value=m['pl_pct'])
            cell.font = FONT_D_LOSS if is_loss else FONT_D_DATA
            cell.number_format = NUMFMT_PCT
            for ci in range(1, 7):
                dash.cell(row=rrr, column=ci).border = BORDER2
            rrr += 1
        return rrr + 2

    top_profitable = sorted(merged_items, key=lambda m: -m['margin'])[:10]
    loss_sorted = sorted(loss_items_all, key=lambda m: m['margin'])[:10]
    rr = write_item_table(rr, f'Top 10 Profitable Items ({period_label})', top_profitable, FILL_PROFIT_HEADER, False)
    if loss_sorted:
        rr = write_item_table(rr, f'Loss-Making Items ({period_label})', loss_sorted, FILL_LOSS_HEADER, True)
    else:
        dash.cell(row=rr, column=1, value='No loss-making line items in this period.').font = _Font(bold=True, color='027A48')

    wb.active = 0
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_division_outlet_daily_report(merged_items: list, all_outlets: Optional[List[str]] = None) -> bytes:
    """Build a date-wise sales matrix with divisions on rows and outlets on columns."""
    import openpyxl as _openpyxl
    from openpyxl.styles import Alignment as _Alignment, Border as _Border, Font as _Font, PatternFill as _PatternFill, Side as _Side
    from openpyxl.utils import get_column_letter as _get_column_letter

    division_labels = {
        "HA": "HA", "HE": "HE", "Mobile": "MOBILE", "Computer": "COMPUTER",
        "Digital Camera": "CAMERA", "Accessories": "ACCESSORY",
    }
    divisions = [division for division in DP_CATEGORIES if any(item["category"] == division for item in merged_items)]
    outlets = sorted(set(all_outlets or []) | {str(item.get("store") or "UNK").upper() for item in merged_items})
    dates = sorted({item["date"] for item in merged_items if item.get("date")})
    sales = defaultdict(float)
    for item in merged_items:
        sales[(item.get("date"), item["category"], str(item.get("store") or "UNK").upper())] += item["sale"]

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws.sheet_view.showGridLines = False
    thin = _Side(style="thin", color="000000")
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = _PatternFill("solid", fgColor="F8CBAD")
    blue_fill = _PatternFill("solid", fgColor="9DC3E6")
    header_font = _Font(name="Arial", size=11, bold=True)
    body_font = _Font(name="Arial", size=10, bold=True)
    number_format = '#,##0;[Red]-#,##0;-'
    last_col = 3 + len(outlets)
    total_col_letter = _get_column_letter(last_col)
    row_no = 1
    first_header_row = 2

    for report_date in dates:
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=last_col)
        title = ws.cell(row_no, 1, f"Division Wise Sale Report {report_date.strftime('%d %b %Y').upper()}")
        title.font = _Font(name="Arial", size=14, bold=True)
        title.alignment = _Alignment(horizontal="center", vertical="center")
        for cell in ws[row_no][:last_col]:
            cell.fill = title_fill
            cell.border = border
        ws.row_dimensions[row_no].height = 24
        row_no += 1

        headers = ["SL no.", "Division", *outlets, "TOTAL VALUE"]
        for col_no, header in enumerate(headers, 1):
            cell = ws.cell(row_no, col_no, header)
            cell.fill = blue_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = _Alignment(horizontal="center", vertical="center")
        row_no += 1
        first_division_row = row_no

        for serial, division in enumerate(divisions, 1):
            ws.cell(row_no, 1, serial)
            ws.cell(row_no, 2, division_labels.get(division, division.upper()))
            for outlet_index, outlet in enumerate(outlets, 3):
                value = round(sales[(report_date, division, outlet)])
                ws.cell(row_no, outlet_index, value if value else 0)
            first_outlet_letter = _get_column_letter(3)
            last_outlet_letter = _get_column_letter(2 + len(outlets))
            ws.cell(row_no, last_col, f"=SUM({first_outlet_letter}{row_no}:{last_outlet_letter}{row_no})")
            for cell in ws[row_no][:last_col]:
                cell.border = border
                cell.font = body_font
                cell.alignment = _Alignment(horizontal="right" if cell.column >= 3 else "center" if cell.column == 1 else "left")
                if cell.column >= 3:
                    cell.number_format = number_format
            row_no += 1

        total_row = row_no
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        ws.cell(total_row, 1, "Total")
        for col_no in range(3, last_col + 1):
            col_letter = _get_column_letter(col_no)
            ws.cell(total_row, col_no, f"=SUM({col_letter}{first_division_row}:{col_letter}{total_row - 1})")
        for cell in ws[total_row][:last_col]:
            cell.fill = blue_fill
            cell.border = border
            cell.font = header_font
            cell.alignment = _Alignment(horizontal="center" if cell.column <= 2 else "right", vertical="center")
            if cell.column >= 3:
                cell.number_format = number_format
        row_no += 2

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    for col_no in range(3, last_col):
        ws.column_dimensions[_get_column_letter(col_no)].width = 14
    ws.column_dimensions[total_col_letter].width = 18
    if len(dates) == 1:
        ws.auto_filter.ref = f"A{first_header_row}:{total_col_letter}{row_no - 2}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================
# DAILY PROFITABILITY REPORT (home-page "Daily Profitability" tile)
# ============================================================
# Builds a store/category-wise profitability report and dashboard directly
# from the same data already uploaded via "Interval Sales Analytics Upload"
# (Sales in your scope -> Date/Vch No/Account/Item/Qty/Sales Amt/Cost Amt
# export from Busy). No separate upload flow: this reads whatever has been
# uploaded there for the requested date range and reshapes it.
#
# Busy exports a split AC as multiple line items under one voucher (an
# indoor-unit line carrying the sale value, plus outdoor-unit/panel lines
# with zero sale value but real cost). Left as-is those zero-sale lines
# would each show as a 100% loss, so they are merged into the priced line
# from the same voucher before any report or dashboard figure is computed.
# When a voucher has more than one priced line (several different products
# bought together) an unpriced line is attributed to whichever priced item
# shares the most name-tokens with it, and that attribution is listed in
# "review_notes" so Admin can sanity-check anything non-obvious.

DP_CATEGORIES = ["HA", "HE", "Computer", "Mobile", "Digital Camera", "Accessories"]

# Busy's Bill-wise Profitability export is GST-exclusive. Every Sale Amount
# and Purchase Price in the Daily Profitability report/dashboard is grossed
# up by this rate so the figures match real (GST-inclusive) invoice amounts.
# Margin scales with it accordingly; PL% (margin/cost) is unaffected since
# both sides of that ratio scale by the same factor.
DP_GST_RATE = 0.18
DP_GST_FACTOR = 1 + DP_GST_RATE


def dp_extract_store(vch_no: Optional[str]) -> str:
    if not vch_no or "/" not in vch_no:
        return (vch_no or "UNK").strip().upper() or "UNK"
    return vch_no.split("/")[0].strip().upper() or "UNK"


def dp_item_similarity(a: Optional[str], b: Optional[str]) -> float:
    """Case-insensitive whole-name similarity (0-1) between two item names,
    used to match an outdoor-unit/zero-value line to the indoor-unit/priced
    line it belongs to on the same voucher.

    A paired indoor/outdoor unit's model numbers are usually near-identical
    strings - e.g. indoor 'ASGG18CEAC-B' pairs with outdoor 'AOGG18CEAC-B',
    only two characters apart - so whole-string similarity is far more
    reliable here than splitting into tokens and requiring an exact token
    match, which treats 'ASGG18CEAC' and 'AOGG18CEAC' as two completely
    unrelated tokens even though they're the same model family, and can tie
    (and then silently mismatch) against an unrelated pair like '18CPWA-B'.
    """
    return difflib.SequenceMatcher(None, (a or "").upper(), (b or "").upper()).ratio()


# Accessories that ride along with a Mobile-category sale but are not a
# phone/tablet themselves - adapters, cables, converters, speakers, etc.
# These are pulled out into "Accessories" so product-brand checks cannot
# sweep them into Computer/Mobile and the report keeps accessories separate
# from genuinely unclassified "Other" items.
DP_ACCESSORY_KEYWORDS = (
    "ADAPTER", "ADAPTOR", "CABLE", "CONVERTER", "CONVERTOR", "CHARGER",
    "HDMI", "POWER BANK", "POWERBANK", "EARPHONE", "HEADPHONE",
    "EARBUD", "EARBUDS", "BUDS", "BATTERY",
    "NECK BAND", "NECKBAND", "NECK-BAND",
    "SMART WATCH", "SMARTWATCH", "TEMPERED GLASS", "SCREEN GUARD",
    "MOBILE COVER", "BACK COVER", "PENDRIVE", "PEN DRIVE", "FLASH DRIVE",
    "USB DRIVE", "MEMORY CARD", "OTG",
    "KEYBOARD", "MOUSE",
    # Small laptop/desktop accessory, not a computer unit itself (e.g.
    # "Laptop Fan", "HP Laptop Fan Cooler") - without this it falls through
    # to the laptop_keywords check below, which matches bare "LAPTOP" as a
    # substring and wrongly counts it as a full Computer-category sale.
    "LAPTOP FAN",
)


def dp_categorize(item_name: Optional[str]) -> str:
    n = (item_name or "").upper()
    padded = f" {n} "

    # Payout and other uncategorized non-PW lines are included in Accessories;
    # PW vouchers themselves are excluded from profitability entirely.
    if "PAYOUT" in n:
        return "Accessories"

    if "INVERTER EB 1100" in n:
        return "HA"

    # Faber built-in kitchen appliances. These model lines previously fell
    # through to the Accessories fallback because the profitability
    # classifier had no HOB/HOOD product-type rule.
    faber_ha_models = (
        "FABER HOB COOKTOP SUPERIA HT904 BR AI N",
        "FABER HOOD EVEREST 3D IN HCSCFLLG60",
    )
    if any(model in n for model in faber_ha_models):
        return "HA"

    # --- Mobile phones, first, before anything else can steal them ---
    # ASUS/dual-purpose brands make both laptops and phones, so a phone
    # model name (ROG Phone, Zenfone) must win over the laptop check below.
    phone_brand_overrides = ("ROG PHONE", "ZENFONE")
    if any(k in n for k in phone_brand_overrides):
        return "Mobile"

    # --- Home Entertainment (TVs, soundbars) ---
    # "LED" alone (exact word) still counts, but so do OLED/QLED/Neo QLED/
    # Mini-LED and other current TV panel tech that contain "LED" as a
    # substring rather than a standalone word - the old check required an
    # exact "LED" token and silently missed all of these.
    he_keywords = (
        "LED", "OLED", "QLED", "MINI LED", "MINILED", "SOUNDBAR",
        # "SOUND BAR" (as two words, e.g. "JBL Sound Bar SB180") - the
        # no-space "SOUNDBAR" form above doesn't catch this, and any
        # sound bar (any brand/model) belongs in Home Entertainment.
        "SOUND BAR",
        "TELEVISION", "SMART TV", "HOME THEATRE", "HOME THEATER",
        # Standalone speakers (any brand) - JBL Party Box, Bluetooth
        # speakers, etc. are their own HE product, not a phone accessory.
        "SPEAKER", "PARTY BOX", "PARTYBOX",
        # Gaming consoles (any brand) - e.g. "Sony PS5 CFI-2116 Std E
        # Chassis ARV", Xbox Series X/S, Nintendo Switch.
        "PLAYSTATION", "PS5", "PS4", "PS3", "GAMING CONSOLE",
        "XBOX", "NINTENDO SWITCH", "NINTENDO",
    )
    if any(k in n for k in he_keywords) or re.search(r"\bTV\b", n):
        return "HE"

    # Accessories/peripherals are excluded before any brand-based check below,
    # so e.g. an HP keyboard or a Lenovo mouse never gets swept into Computer
    # just because the brand also makes laptops.
    if any(k in n for k in DP_ACCESSORY_KEYWORDS):
        return "Accessories"

    laptop_keywords = (
        "DELL", "LAPTOP", "BACK PACK", "BACKPACK", "LENOVO", "ASUS", "ACER",
        "MSI", "MACBOOK", "IMAC", "THINKPAD", "IDEAPAD", "VIVOBOOK",
        "CHROMEBOOK", "NOTEBOOK", "GIGABYTE", "DESKTOP", "CPU CABINET",
        "MONITOR", "ALL IN ONE PC",
        # Printers (any brand) - PRINTER as a substring catches Laserjet/
        # Inkjet/MFP model names too (e.g. "HP Printer Laserjet MFP...").
        "PRINTER", "LASERJET", "INKJET",
    )
    if any(k in n for k in laptop_keywords) or re.search(r"\bHP\b", n) or re.search(r"\bMBA\b", n):
        return "Computer"

    ha_keywords = (
        "SAC ", "WAC ", "CASSETTE AC", "AC 3T", " AC ", " TON ", "SPLIT AC",
        "WINDOW AC", "INVERTER AC", "PANEL", "CHEST FREEZER", "DEEP FREEZER",
        " REF ", "REF EON", "REF RD", "REF HRD", "REF SJ", "REFRIGERATOR",
        "FRIDGE", "COOLER", "WATER PURIFIER", "WATER HEATER", "GEYSER",
        "EXCELL PART", "GARMENT STEAMER", " MW ", "MICROWAVE", "MWO",
        " FAN ", " WM ", "WASHING MACHINE", "MIXER GRINDER", "INDUCTION",
        "IN ICT", "CHIMNEY", "DISHWASHER", "VACUUM CLEANER", "AIR PURIFIER",
        "ROOM HEATER", "IRON BOX", "STEAM IRON", "AQUAGUARD", " RO ",
        # Drinking-water and small kitchen appliances. These explicit product
        # types must win regardless of brand (e.g. Voltas/Bajaj).
        "WATER DISPENSER", "SANDWICH TOASTER",
        # Bare "IRON" (any brand - dry iron, steam iron, curling iron,
        # garment iron, etc.) as a standalone word, not just the two
        # specific compounds above.
        " IRON ",
        # Kitchen appliances that weren't covered by "MIXER GRINDER" alone
        # (e.g. "Faber Juicer Mixer FSJ 200BK" has neither word next to
        # "GRINDER", so it fell through to Other before this).
        " MIXER ", "JUICER", "BLENDER", "FOOD PROCESSOR", "SANDWICH MAKER",
        "RICE COOKER", "PRESSURE COOKER", " OTG ", "HAND BLENDER",
        # Fryers (any brand) - e.g. "Philips Air Fryer NA130", single-pot
        # fryers, deep fryers. Bare "FRYER" catches all of these regardless
        # of what comes before it, so any fryer lands in Home Appliances.
        "FRYER",
        # Hair care appliances (any brand) - e.g. "Dyson Corrale Straightener
        # BC/BN", "Dyson Airstrait HT01 BN/BC", hair dryers/curlers.
        "STRAIGHTENER", "CORRALE", "AIRSTRAIT", "HAIR DRYER", "HAIR STYLER",
        "HAIR CURLER", "AIRWRAP",
    )
    if any(k in padded for k in ha_keywords):
        return "HA"

    # --- Digital Camera (any brand) ---
    dc_keywords = (
        "CAMERA", "DSLR", "MIRRORLESS", "GOPRO", "CAMCORDER",
        "ACTION CAM", "GIMBAL", "GO PRO",
    )
    if any(k in n for k in dc_keywords) or re.search(r"\bLENS\b", n):
        return "Digital Camera"

    # Phones/tablets consistently carry a RAM+Storage config in the name
    # (e.g. "8+128", "4 + 64", "6+256") regardless of brand - this catches
    # new/unlisted phone brands without needing a brand keyword for each one.
    # High-end foldables sometimes list storage in TB rather than GB (e.g.
    # "16+1TB"), so a single low-order digit followed by TB counts too.
    if re.search(r"\b\d{1,2}\s*\+\s*\d{2,3}\b", n) or re.search(r"\b\d{1,2}\s*\+\s*\d{1,2}\s*TB\b", n):
        return "Mobile"

    mobile_keywords = (
        "IPHONE", "VIVO", "OPPO", "REALME", "REDMI", "MOTOROLA", "ONEPLUS",
        "POCO", "NOTHING", " TAB ", " PAD ", "SAMSUNG Z FOLD", "PIXEL",
        "IQOO", "SAMSUNG GALAXY",
        # Foldable phones (any brand) - Fold/Flip model names.
        " FOLD ", " FLIP ",
        # iPad (any model - Air/Pro/Mini/base) - grouped under Mobile along
        # with every other Tab/Pad device, same as the rest of the app.
        # Bare "IPAD" (no padding) since "iPad" normalizes to one token
        # with no internal space (e.g. "iPad Air 11" M3 WiFi MC9YAHN/A
        # 128GB Sta").
        "IPAD",
    )
    if any(k in padded for k in mobile_keywords):
        return "Mobile"
    if "SAMSUNG" in n and re.search(r"\bF\d{2,3}[A-Z]?\b", n):
        return "Mobile"
    if "SAMSUNG" in n and re.search(r"\b[AMS]\d{2,3}[A-Z]?\b", n):
        return "Mobile"
    return "Accessories"


def dp_is_ac_od_component(item_name: Optional[str]) -> bool:
    """True for AC outdoor-unit line items (e.g. 'Daikin SAC OD RKC 50
    XV16UKA', 'Daikin Cassette AC OD RGVF48CRY16') whose cost gets merged
    into the indoor unit but should no longer be spelled out by name in
    the 'incl. cost of:' narration - their cost is already reflected
    correctly in the merged figures without needing to be named."""
    return "AC OD" in (item_name or "").upper()


def dp_merge_rows(rows: List["models.IntervalSaleUpload"]) -> tuple:
    """Groups raw Busy line items by voucher, merges zero-sale-value lines
    (OD units, panels, bundled parts) into the priced line(s) of the same
    voucher, then tags each merged item with its store code and category.
    Returns (merged_items, review_notes)."""
    groups = defaultdict(list)
    for r in rows:
        key = r.vch_no or f"__no_vch_{r.id}"
        if "PW" in str(key).upper():
            continue
        groups[key].append(r)

    merged = []
    review_notes = []

    for vch, grp in groups.items():
        priced = [r for r in grp if (r.sales_amt or 0) > 0]
        zero = [r for r in grp if not (r.sales_amt or 0) > 0]
        store = dp_extract_store(vch if not str(vch).startswith("__no_vch_") else None)

        if not priced:
            for r in grp:
                merged.append({
                    "vch": vch, "item": r.item or "Unknown Item", "sale": r.sales_amt or 0,
                    "cost": r.cost_amt or 0, "store": store, "date": r.sale_date,
                    "note": "Standalone (no sale value recorded)",
                })
            continue

        if len(priced) == 1:
            p = priced[0]
            extra_cost = sum((r.cost_amt or 0) for r in zero)
            components = ", ".join(
                (r.item or "").strip() for r in zero
                if r.item and not dp_is_ac_od_component(r.item)
            )
            note = f"incl. cost of: {components}" if components else ""
            merged.append({
                "vch": vch, "item": p.item or "Unknown Item", "sale": p.sales_amt or 0,
                "cost": (p.cost_amt or 0) + extra_cost, "store": store, "date": p.sale_date,
                "note": note,
            })
            continue

        # multiple priced lines in one voucher: attribute each zero-sale line
        # to the priced line whose name is the closest overall match (see
        # dp_item_similarity - paired indoor/outdoor model numbers are
        # near-identical strings, not necessarily sharing whole tokens).
        bucket = {id(p): {"item": p.item, "sale": p.sales_amt or 0, "cost": p.cost_amt or 0, "date": p.sale_date, "components": []} for p in priced}
        for z in zero:
            best, best_score = None, -1.0
            for p in priced:
                score = dp_item_similarity(z.item, p.item)
                if score > best_score:
                    best, best_score = p, score
            if best_score < 0.5:
                review_notes.append(
                    f"Vch {vch}: could not confidently match zero-value line "
                    f"'{z.item}' (cost {z.cost_amt or 0:.2f}) to a product in the "
                    f"same voucher - attributed to '{priced[0].item}' by default."
                )
                best = priced[0]
            else:
                review_notes.append(
                    f"Vch {vch}: cost {z.cost_amt or 0:.2f} from '{z.item}' "
                    f"attributed to '{best.item}' (matched by name similarity)."
                )
            bucket[id(best)]["cost"] += (z.cost_amt or 0)
            if z.item and not dp_is_ac_od_component(z.item):
                bucket[id(best)]["components"].append(z.item.strip())

        for p in priced:
            u = bucket[id(p)]
            note = f"incl. cost of: {', '.join(u['components'])}" if u["components"] else ""
            merged.append({
                "vch": vch, "item": u["item"] or "Unknown Item", "sale": u["sale"],
                "cost": u["cost"], "store": store, "date": u["date"], "note": note,
            })

    for m in merged:
        m["category"] = dp_categorize(m["item"])
        m["sale"] = m["sale"] * DP_GST_FACTOR
        m["cost"] = m["cost"] * DP_GST_FACTOR
        m["margin"] = m["sale"] - m["cost"]
        m["pl_pct"] = (m["margin"] / m["cost"]) if m["cost"] else 0.0

    return merged, review_notes


def dp_filter_rows(db: Session, start_date: Optional[date], end_date: Optional[date]):
    query = db.query(models.IntervalSaleUpload)
    if start_date:
        query = query.filter(models.IntervalSaleUpload.sale_date >= start_date)
    if end_date:
        query = query.filter(models.IntervalSaleUpload.sale_date <= end_date)
    return query.order_by(models.IntervalSaleUpload.sale_date.asc(), models.IntervalSaleUpload.id.asc()).all()


def dp_apply_filters(merged: List[dict], category: Optional[str], store: Optional[str]) -> List[dict]:
    out = merged
    if category and category.upper() != "ALL":
        out = [m for m in out if m["category"].upper() == category.upper()]
    if store and store.upper() != "ALL":
        out = [m for m in out if m["store"].upper() == store.upper()]
    return out


def dp_serialize_item(m: dict) -> dict:
    return {
        "store": m["store"], "vch_no": m["vch"], "item": m["item"], "category": m["category"],
        "sale": round(m["sale"], 2), "cost": round(m["cost"], 2), "margin": round(m["margin"], 2),
        "pl_pct": round(m["pl_pct"] * 100, 2), "note": m["note"],
    }


@app.get("/api/daily-profitability/meta")
def daily_profitability_meta(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    date_bounds = db.query(
        func.min(models.IntervalSaleUpload.sale_date),
        func.max(models.IntervalSaleUpload.sale_date),
    ).first()
    has_data = bool(date_bounds and date_bounds[0])
    vch_numbers = [row[0] for row in db.query(models.IntervalSaleUpload.vch_no).distinct().all()]
    stores = sorted({dp_extract_store(v) for v in vch_numbers if v})

    return {
        "has_data": has_data,
        "date_from": date_bounds[0] if has_data else None,
        "date_to": date_bounds[1] if has_data else None,
        "categories": DP_CATEGORIES,
        "stores": stores,
        "can_upload": auth.has_admin_access(current_user) or current_user.role == "MISExecutive",
    }


@app.get("/api/daily-profitability/dashboard")
def daily_profitability_dashboard(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    rows = dp_filter_rows(db, start_date, end_date)
    if not rows:
        return {"has_data": False}

    merged, review_notes = dp_merge_rows(rows)
    merged = dp_apply_filters(merged, category, store)
    if not merged:
        return {"has_data": False}

    total_sale = sum(m["sale"] for m in merged)
    total_cost = sum(m["cost"] for m in merged)
    total_margin = total_sale - total_cost

    by_cat = defaultdict(lambda: {"sale": 0.0, "cost": 0.0, "margin": 0.0, "count": 0})
    by_store = defaultdict(lambda: {"sale": 0.0, "cost": 0.0, "margin": 0.0, "count": 0})
    for m in merged:
        c = by_cat[m["category"]]
        c["sale"] += m["sale"]; c["cost"] += m["cost"]; c["margin"] += m["margin"]; c["count"] += 1
        s = by_store[m["store"]]
        s["sale"] += m["sale"]; s["cost"] += m["cost"]; s["margin"] += m["margin"]; s["count"] += 1

    loss_all = [m for m in merged if m["margin"] < 0]
    top_profitable = sorted(merged, key=lambda m: -m["margin"])[:10]
    loss_making = sorted(loss_all, key=lambda m: m["margin"])[:15]

    return {
        "has_data": True,
        "kpis": {
            "total_sale": round(total_sale, 2),
            "total_cost": round(total_cost, 2),
            "total_margin": round(total_margin, 2),
            "margin_pct": round((total_margin / total_cost * 100) if total_cost else 0.0, 2),
            "line_items": len(merged),
            "loss_items": len(loss_all),
            "stores": len(by_store),
        },
        "by_category": [
            {"category": c, "sale": round(d["sale"], 2), "cost": round(d["cost"], 2),
             "margin": round(d["margin"], 2), "count": d["count"]}
            for c, d in sorted(by_cat.items())
        ],
        "by_store": [
            {"store": s, "sale": round(d["sale"], 2), "cost": round(d["cost"], 2),
             "margin": round(d["margin"], 2), "count": d["count"]}
            for s, d in sorted(by_store.items())
        ],
        "top_profitable": [dp_serialize_item(m) for m in top_profitable],
        "loss_making": [dp_serialize_item(m) for m in loss_making],
        "review_notes": review_notes[:20],
        "filters": {
            "start_date": start_date.isoformat() if start_date else None,
            "end_date": end_date.isoformat() if end_date else None,
            "category": category or "ALL",
            "store": store or "ALL",
        },
    }


@app.get("/api/daily-profitability/items")
def daily_profitability_items(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Full (not top-10) list of line items for the current filters - backs
    the "View Items" button so Admin can see every sale in a category/store,
    not just the highlights on the main dashboard."""
    rows = dp_filter_rows(db, start_date, end_date)
    if not rows:
        return {"has_data": False, "items": [], "count": 0}

    merged, _ = dp_merge_rows(rows)
    merged = dp_apply_filters(merged, category, store)
    if not merged:
        return {"has_data": False, "items": [], "count": 0}

    items_sorted = sorted(merged, key=lambda m: m['item'].lower())
    return {
        "has_data": True,
        "count": len(items_sorted),
        "items": [dp_serialize_item(m) for m in items_sorted],
    }


@app.get("/api/daily-profitability/download")
def daily_profitability_download(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    rows = dp_filter_rows(db, start_date, end_date)
    if not rows:
        raise HTTPException(status_code=404, detail="No data available for the selected range. Upload a Busy export first (Sales in your scope -> Interval Sales Analytics Upload).")

    merged, _ = dp_merge_rows(rows)
    merged = dp_apply_filters(merged, category, store)
    if not merged:
        raise HTTPException(status_code=404, detail="No rows match the selected filters.")

    if start_date and end_date:
        label = f"{start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}" if start_date != end_date else start_date.strftime("%d-%b-%Y")
    else:
        dates = [m["date"] for m in merged if m.get("date")]
        label = f"{min(dates).strftime('%d-%b-%Y')} to {max(dates).strftime('%d-%b-%Y')}" if dates else "All Dates"

    workbook_bytes = build_daily_profitability_workbook(merged, label)

    fname_bit = f"{start_date}_{end_date}" if (start_date or end_date) else "all"
    filename = f"Daily_Profitability_{fname_bit}.xlsx"
    return Response(
        content=workbook_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/daily-profitability/daily-report")
def daily_profitability_daily_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    category: Optional[str] = Query(None),
    store: Optional[str] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    rows = dp_filter_rows(db, start_date, end_date)
    if not rows:
        raise HTTPException(status_code=404, detail="No data available for the selected range.")
    merged, _ = dp_merge_rows(rows)
    merged = dp_apply_filters(merged, category, store)
    if not merged:
        raise HTTPException(status_code=404, detail="No rows match the selected filters.")
    # The standalone division-wise Daily Report must show Busy's original
    # GST-exclusive values. dp_merge_rows adds estimated GST for the main
    # profitability dashboard/report, so reverse that uplift only for this
    # endpoint and leave every other profitability view unchanged.
    for item in merged:
        item["sale"] /= DP_GST_FACTOR
        item["cost"] /= DP_GST_FACTOR
        item["margin"] = item["sale"] - item["cost"]
        item["pl_pct"] = (item["margin"] / item["cost"]) if item["cost"] else 0.0
    all_outlets = sorted({
        dp_extract_store(voucher)
        for (voucher,) in db.query(models.IntervalSaleUpload.vch_no).distinct().all()
        if voucher and "PW" not in voucher.upper()
    })
    workbook_bytes = build_division_outlet_daily_report(merged, all_outlets)
    filename_date = start_date.isoformat() if start_date and start_date == end_date else f"{start_date or 'first'}_{end_date or 'latest'}"
    return Response(
        content=workbook_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Division_Wise_Daily_Report_{filename_date}.xlsx"'},
    )


@app.get("/daily-profitability")
@app.get("/daily-profitability.html")
def daily_profitability_page():
    return serve_html("static/daily_profitability.html")




# ============================================================
# CLAIMS
# ============================================================

@app.get("/claims", response_model=List[schemas.ClaimOut])
def list_claims(current_user: models.User = Depends(auth.get_current_user), db: Session = Depends(get_db)):
    return get_claims_for_user(db, current_user)


@app.put("/claims/{claim_id}/status")
def update_claim_status(
    claim_id: int, update: schemas.ClaimStatusUpdate, db: Session = Depends(get_db)
):
    claim = db.query(models.ClaimHeader).filter(models.ClaimHeader.id == claim_id).first()
    if not claim:
        raise HTTPException(status_code=404, detail="Claim not found")

    old_status = claim.status
    claim.status = update.new_status
    db.commit()

    history = models.ClaimStatusHistory(
        claim_id=claim.id,
        old_status=old_status,
        new_status=update.new_status,
        remarks=update.remarks,
    )
    db.add(history)
    db.commit()

    return {
        "message": f"Claim {claim_id} status changed from {old_status} to {update.new_status}"
    }

# ============================================================
# AI ANALYSIS DASHBOARD (home-page "AI Analysis" tile)
# ============================================================
# Upload any sales/profitability Excel or CSV export (Date, Item, Sales Amt,
# Cost Amt, Profit/Loss, and optionally Division/Qty - column names are
# matched loosely and out of order). Every sheet in a workbook is read, so a
# single file with one tab per financial year works in one upload. The
# previous dataset is replaced each time a new file is uploaded - this is a
# "current snapshot" dashboard, not a historical archive.
#
# All KPIs, rankings, and the recommendations panel are computed directly
# from the uploaded numbers with plain arithmetic/aggregation below - no
# external AI call is made, so it stays fast, free, and every recommendation
# can be traced back to a real figure in the data.

ANALYTICS_HEADER_ALIASES = {
    "sale_date": {"date", "invoicedate", "saledate", "billdate"},
    "vch_no": {"vchno", "voucherno", "invoiceno", "billno"},
    "item": {"item", "product", "itemname", "description"},
    "qty": {"qty", "quantity"},
    "sales_amt": {"salesamt", "salesamount", "salevalue", "amount", "invoicevalue"},
    "cost_amt": {"costamt", "costamount", "cost"},
    "profit_loss": {"profitloss", "grossprofit"},
    "division": {"division", "div", "segment"},
}


def analytics_canonical_column(normalized_header: str) -> Optional[str]:
    for canonical, aliases in ANALYTICS_HEADER_ALIASES.items():
        if normalized_header in aliases:
            return canonical
    return None


def find_analytics_header_row(table_rows: List[List]) -> tuple:
    scan_limit = min(len(table_rows), 15)
    best_index = 0
    best_score = -1
    for row_index in range(scan_limit):
        row = table_rows[row_index] or []
        normalized = [normalize_header_name(cell) for cell in row]
        canonicals = {analytics_canonical_column(name) for name in normalized if analytics_canonical_column(name)}
        canonicals.discard(None)
        score = len(canonicals)
        if "sale_date" in canonicals:
            score += 2
        if "sales_amt" in canonicals:
            score += 2
        if "item" in canonicals:
            score += 1
        if score > best_score:
            best_score = score
            best_index = row_index
    return best_index, best_score


def parse_analytics_file(filename: str, content: bytes) -> List[dict]:
    ext = "." + filename.lower().split(".")[-1] if "." in filename else ""
    rows_out: List[dict] = []

    def build_row(row_dict: dict, source_sheet: Optional[str]) -> Optional[dict]:
        if not any(str(v or "").strip() for v in row_dict.values()):
            return None
        try:
            sale_date = parse_date_value(row_dict.get("sale_date"))
        except Exception:
            return None
        item = str(row_dict.get("item") or "").strip() or "Unknown Item"
        division_raw = str(row_dict.get("division") or "").strip()
        division = division_raw.upper() if division_raw else "UNCATEGORIZED"
        vch_no = str(row_dict.get("vch_no") or "").strip() or None
        qty_raw = row_dict.get("qty")
        qty = parse_float_value(qty_raw, fallback=None) if str(qty_raw or "").strip() else None
        sales_amt = parse_float_value(row_dict.get("sales_amt"), fallback=0.0)
        cost_amt = parse_float_value(row_dict.get("cost_amt"), fallback=0.0)
        profit_loss = parse_float_value(row_dict.get("profit_loss"), fallback=sales_amt - cost_amt)
        return {
            "sale_date": sale_date,
            "item": item,
            "division": division,
            "division_from_file": division_raw.upper() if division_raw else None,
            "vch_no": vch_no,
            "qty": qty,
            "sales_amt": sales_amt,
            "cost_amt": cost_amt,
            "profit_loss": profit_loss,
            "source_sheet": source_sheet,
            "source_file": filename,
        }

    if ext in {".xlsx", ".xls"}:
        try:
            openpyxl_module = importlib.import_module("openpyxl")
            load_workbook = openpyxl_module.load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=400, detail="Excel upload requires openpyxl package. Install: pip install openpyxl") from exc

        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        for worksheet in workbook.worksheets:
            raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not raw_rows:
                continue
            header_index, score = find_analytics_header_row(raw_rows)
            if score < 3:
                continue
            headers = [normalize_header_name(cell) for cell in raw_rows[header_index]]
            canonical_headers = [analytics_canonical_column(h) for h in headers]
            if "sale_date" not in canonical_headers or "sales_amt" not in canonical_headers:
                continue

            for row in raw_rows[header_index + 1:]:
                row_dict = {}
                for idx, canonical in enumerate(canonical_headers):
                    if canonical:
                        row_dict[canonical] = row[idx] if idx < len(row) else None
                built = build_row(row_dict, worksheet.title)
                if built:
                    rows_out.append(built)
        return rows_out

    if ext == ".csv":
        decoded = content.decode("utf-8-sig", errors="replace")
        raw_lines = [line for line in decoded.splitlines() if line.strip()]
        if not raw_lines:
            return []
        preview_rows = [next(csv.reader([line])) for line in raw_lines[:15]]
        header_index, score = find_analytics_header_row(preview_rows)
        if score < 3:
            return []

        data_lines = raw_lines[header_index:]
        reader = csv.DictReader(data_lines)
        for input_row in reader:
            row_dict = {}
            for key, value in input_row.items():
                canonical = analytics_canonical_column(normalize_header_name(key))
                if canonical:
                    row_dict[canonical] = value
            built = build_row(row_dict, None)
            if built:
                rows_out.append(built)
        return rows_out

    raise HTTPException(status_code=400, detail="Unsupported file format. Upload an Excel (.xlsx/.xls) or CSV file.")


def analytics_fiscal_year_label(d: date) -> str:
    start_year = d.year if d.month >= 4 else d.year - 1
    end_year = start_year + 1
    return f"{start_year}-{str(end_year)[-2:]}"


def build_analytics_dashboard(rows: List[models.AnalyticsSalesRow]) -> dict:
    if not rows:
        return {"has_data": False}

    total_sales = 0.0
    total_cost = 0.0
    total_profit = 0.0
    dates = []

    item_stats = defaultdict(lambda: {"sales": 0.0, "cost": 0.0, "profit": 0.0, "qty": 0.0, "qty_known": False, "count": 0})
    division_stats = defaultdict(lambda: {"sales": 0.0, "cost": 0.0, "profit": 0.0, "count": 0})
    brand_stats = defaultdict(lambda: {"sales": 0.0, "cost": 0.0, "profit": 0.0, "count": 0})
    monthly_stats = defaultdict(lambda: {"sales": 0.0, "cost": 0.0, "profit": 0.0})
    yearly_stats = defaultdict(lambda: {"sales": 0.0, "cost": 0.0, "profit": 0.0})
    division_year_profit = defaultdict(float)
    month_number_profit = defaultdict(float)  # 1-12, across all years, for seasonality

    for row in rows:
        total_sales += row.sales_amt or 0.0
        total_cost += row.cost_amt or 0.0
        total_profit += row.profit_loss or 0.0
        dates.append(row.sale_date)

        item = row.item or "Unknown Item"
        istat = item_stats[item]
        istat["sales"] += row.sales_amt or 0.0
        istat["cost"] += row.cost_amt or 0.0
        istat["profit"] += row.profit_loss or 0.0
        istat["count"] += 1
        if row.qty is not None:
            istat["qty"] += row.qty
            istat["qty_known"] = True

        division = row.division or "UNCATEGORIZED"
        dstat = division_stats[division]
        dstat["sales"] += row.sales_amt or 0.0
        dstat["cost"] += row.cost_amt or 0.0
        dstat["profit"] += row.profit_loss or 0.0
        dstat["count"] += 1

        brand = row.brand or "Unknown Brand"
        bstat = brand_stats[brand]
        bstat["sales"] += row.sales_amt or 0.0
        bstat["cost"] += row.cost_amt or 0.0
        bstat["profit"] += row.profit_loss or 0.0
        bstat["count"] += 1

        ym = row.sale_date.strftime("%Y-%m")
        mstat = monthly_stats[ym]
        mstat["sales"] += row.sales_amt or 0.0
        mstat["cost"] += row.cost_amt or 0.0
        mstat["profit"] += row.profit_loss or 0.0

        fy = analytics_fiscal_year_label(row.sale_date)
        ystat = yearly_stats[fy]
        ystat["sales"] += row.sales_amt or 0.0
        ystat["cost"] += row.cost_amt or 0.0
        ystat["profit"] += row.profit_loss or 0.0

        division_year_profit[(division, fy)] += row.profit_loss or 0.0
        month_number_profit[row.sale_date.month] += row.profit_loss or 0.0

    def margin(sales, profit):
        return round((profit / sales) * 100, 2) if sales else 0.0

    overall_margin = margin(total_sales, total_profit)

    def item_out(name, stats):
        return {
            "item": name,
            "sales": round(stats["sales"], 2),
            "cost": round(stats["cost"], 2),
            "profit": round(stats["profit"], 2),
            "margin_percent": margin(stats["sales"], stats["profit"]),
            "qty": round(stats["qty"], 2) if stats["qty_known"] else None,
            "transactions": stats["count"],
        }

    all_items = [item_out(name, stats) for name, stats in item_stats.items()]
    top_profit_items = sorted(all_items, key=lambda x: x["profit"], reverse=True)[:10]
    top_revenue_items = sorted(all_items, key=lambda x: x["sales"], reverse=True)[:10]
    loss_items = sorted([i for i in all_items if i["profit"] < 0], key=lambda x: x["profit"])[:10]
    qty_known_items = [i for i in all_items if i["qty"] is not None]
    top_qty_items = sorted(qty_known_items, key=lambda x: x["qty"], reverse=True)[:10]

    # Margin leaders/laggards - only among items with enough transactions to
    # be meaningful (avoids one lucky/unlucky single sale skewing the list).
    margin_eligible = [i for i in all_items if i["transactions"] >= 3 and i["sales"] > 0]
    top_margin_items = sorted(margin_eligible, key=lambda x: x["margin_percent"], reverse=True)[:10]
    bottom_margin_items = sorted(margin_eligible, key=lambda x: x["margin_percent"])[:10]

    division_breakdown = []
    for name, stats in division_stats.items():
        division_breakdown.append({
            "division": name,
            "sales": round(stats["sales"], 2),
            "cost": round(stats["cost"], 2),
            "profit": round(stats["profit"], 2),
            "margin_percent": margin(stats["sales"], stats["profit"]),
            "transactions": stats["count"],
            "profit_share_percent": round((stats["profit"] / total_profit) * 100, 2) if total_profit else 0.0,
        })
    division_breakdown.sort(key=lambda x: x["profit"], reverse=True)

    brand_breakdown = []
    for name, stats in brand_stats.items():
        brand_breakdown.append({
            "brand": name,
            "sales": round(stats["sales"], 2),
            "cost": round(stats["cost"], 2),
            "profit": round(stats["profit"], 2),
            "margin_percent": margin(stats["sales"], stats["profit"]),
            "transactions": stats["count"],
        })
    brand_breakdown.sort(key=lambda x: x["profit"], reverse=True)
    brand_breakdown = brand_breakdown[:15]

    monthly_trend = [
        {"period": ym, "sales": round(s["sales"], 2), "cost": round(s["cost"], 2), "profit": round(s["profit"], 2)}
        for ym, s in sorted(monthly_stats.items())
    ]
    yearly_trend = [
        {"period": fy, "sales": round(s["sales"], 2), "cost": round(s["cost"], 2), "profit": round(s["profit"], 2), "margin_percent": margin(s["sales"], s["profit"])}
        for fy, s in sorted(yearly_stats.items())
    ]

    # ---------------- Recommendations (rule-based, computed from the
    # aggregates above - every number quoted is real, nothing is invented) ----
    recommendations = []

    if division_breakdown:
        leader = division_breakdown[0]
        if leader["profit"] > 0:
            recommendations.append({
                "type": "opportunity",
                "priority": "high",
                "title": f"{leader['division']} is your leading profit driver",
                "detail": f"It contributed ₹{leader['profit']:,.0f} in profit, {leader['profit_share_percent']:.1f}% of total profit across the uploaded data. Prioritize stock availability and scheme/promotional support here to protect this contribution.",
            })

    # YoY growth/decline per division, comparing the two most recent fiscal
    # years that division has data for.
    division_years = defaultdict(dict)
    for (division, fy), profit in division_year_profit.items():
        division_years[division][fy] = profit
    for division, year_map in division_years.items():
        years_sorted = sorted(year_map.keys())
        if len(years_sorted) < 2:
            continue
        prev_fy, latest_fy = years_sorted[-2], years_sorted[-1]
        prev_profit, latest_profit = year_map[prev_fy], year_map[latest_fy]
        if prev_profit == 0:
            continue
        change_pct = ((latest_profit - prev_profit) / abs(prev_profit)) * 100
        if change_pct <= -15:
            recommendations.append({
                "type": "decline",
                "priority": "high",
                "title": f"{division} profit declined {abs(change_pct):.0f}% year-on-year",
                "detail": f"FY {prev_fy} → FY {latest_fy}: ₹{prev_profit:,.0f} → ₹{latest_profit:,.0f}. Consider a fresh scheme, a pricing/cost review, or a promotional push to reverse the trend.",
            })
        elif change_pct >= 15:
            recommendations.append({
                "type": "growth",
                "priority": "medium",
                "title": f"{division} profit grew {change_pct:.0f}% year-on-year",
                "detail": f"FY {prev_fy} → FY {latest_fy}: ₹{prev_profit:,.0f} → ₹{latest_profit:,.0f}. Increase inventory allocation and marketing focus here to capture the momentum.",
            })

    if loss_items:
        names = ", ".join(f"{i['item']} (₹{i['profit']:,.0f})" for i in loss_items[:3])
        recommendations.append({
            "type": "loss",
            "priority": "high",
            "title": f"{len(loss_items)} item(s) are being sold at a net loss",
            "detail": f"Worst offenders: {names}. Review purchase cost, scheme support, or selling price for these lines.",
        })

    if margin_eligible:
        high_revenue_cutoff = sorted([i["sales"] for i in margin_eligible], reverse=True)
        cutoff_value = high_revenue_cutoff[max(0, len(high_revenue_cutoff) // 4 - 1)] if len(high_revenue_cutoff) >= 4 else high_revenue_cutoff[0]
        candidates = [
            i for i in margin_eligible
            if i["sales"] >= cutoff_value and i["margin_percent"] < overall_margin and i["margin_percent"] >= 0
        ]
        candidates.sort(key=lambda x: x["sales"], reverse=True)
        if candidates:
            top = candidates[0]
            recommendations.append({
                "type": "opportunity",
                "priority": "medium",
                "title": f"{top['item']} sells well but at a thin margin",
                "detail": f"₹{top['sales']:,.0f} in revenue at only {top['margin_percent']:.1f}% margin, below your overall {overall_margin:.1f}% average. Consider bundling with a scheme, or renegotiating cost, to lift profitability on this high-volume line.",
            })

    if month_number_profit:
        month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        best_month_num = max(month_number_profit, key=month_number_profit.get)
        worst_month_num = min(month_number_profit, key=month_number_profit.get)
        if best_month_num != worst_month_num:
            recommendations.append({
                "type": "seasonal",
                "priority": "low",
                "title": f"{month_names[best_month_num]} is historically your strongest month",
                "detail": f"Across the uploaded years, {month_names[best_month_num]} generated the most profit and {month_names[worst_month_num]} the least. Plan stock and staffing ahead of {month_names[best_month_num]}, and consider a targeted scheme in {month_names[worst_month_num]} to offset the seasonal dip.",
            })

    if total_sales > 0:
        if overall_margin < 8:
            recommendations.append({
                "type": "decline",
                "priority": "medium",
                "title": f"Overall margin is {overall_margin:.1f}%",
                "detail": "This is below a typical general-retail benchmark of ~10-12%. A broad cost or pricing review across top-selling lines may help.",
            })
        else:
            recommendations.append({
                "type": "growth",
                "priority": "low",
                "title": f"Overall margin of {overall_margin:.1f}% is healthy",
                "detail": "Maintain current pricing and scheme discipline; use the division and item breakdowns above to reinforce what's already working.",
            })

    priority_rank = {"high": 0, "medium": 1, "low": 2}
    recommendations.sort(key=lambda r: priority_rank.get(r["priority"], 3))

    return {
        "has_data": True,
        "kpis": {
            "total_sales": round(total_sales, 2),
            "total_cost": round(total_cost, 2),
            "total_profit": round(total_profit, 2),
            "margin_percent": overall_margin,
            "transactions": len(rows),
            "unique_items": len(item_stats),
            "divisions": len(division_stats),
            "date_from": min(dates).isoformat() if dates else None,
            "date_to": max(dates).isoformat() if dates else None,
        },
        "top_profit_items": top_profit_items,
        "top_revenue_items": top_revenue_items,
        "loss_items": loss_items,
        "top_margin_items": top_margin_items,
        "bottom_margin_items": bottom_margin_items,
        "top_qty_items": top_qty_items,
        "division_breakdown": division_breakdown,
        "brand_breakdown": brand_breakdown,
        "monthly_trend": monthly_trend,
        "yearly_trend": yearly_trend,
        "recommendations": recommendations,
    }



# ------------------------------------------------------------------
# DIVISION / BRAND AUTO-DETECTION + AC INDOOR/OUTDOOR MERGE
# ------------------------------------------------------------------
# Keyword rules below mirror the reference mapping the business uses
# (Home Appliance / Home Entertainment / Mobile / Computer / Digital
# Camera). Category keywords are checked before brand names, since a
# brand like Samsung or LG sells across every division and only the
# item text (AC/TV/Fridge/etc.) tells us which one a given row is.
DIVISION_KEYWORD_RULES = [
    ("HA", [
        # Cooling
        r"\bs\.?a\.?c\b", r"\bw\.?a\.?c\b", r"\bac\b", r"split\s*ac", r"window\s*ac",
        r"\brefrigerator\b", r"\brefregerator\b", r"\brefrigirator\b", r"\bfridge\b",
        r"\bcooler\b", r"air\s*cooler", r"chest\s*fre+zer", r"deep\s*fre+zer", r"\bfre+zer\b",
        # Fans & ventilation
        r"\bfan\b", r"ceiling\s*fan", r"table\s*fan", r"pedestal\s*fan", r"exhaust\s*fan",
        r"tower\s*fan", r"wall\s*fan",
        # Washing / cleaning
        r"washing\s*machine", r"\bwm\b", r"dish\s*washer", r"vacuum\s*cleaner",
        # Kitchen appliances
        r"\bmicrowave\b", r"\bmwo\b", r"\boven\b", r"\botg\b", r"induction\s*cook(top|er)?",
        r"\bcooktop\b", r"gas\s*stove", r"\bhob\b", r"\bchimney\b", r"mixer\s*grinder",
        r"\bmixer\b", r"\bjuicer\b", r"\bblender\b", r"food\s*processor", r"\btoaster\b",
        r"sandwich\s*maker", r"rice\s*cooker", r"pressure\s*cooker", r"\bkettle\b",
        r"electric\s*kettle",
        # Water
        r"water\s*purifier", r"\bro\s*purifier\b", r"water\s*dispenser", r"\bwater\s*d\b",
        r"\bgeyser\b", r"water\s*heater", r"immersion\s*rod", r"air\s*purifier",
        # Personal care / other small appliances
        r"hair\s*dryer", r"\btrimmer\b", r"\bshaver\b",
        r"\biron\b", r"dry\s*iron", r"steam\s*iron", r"\biron\s*box\b",
        r"garment\s*steamer", r"\bsteamer\b",
        r"hair\s*straightener", r"hair\s*curler", r"hair\s*styler", r"\bairwrap\b",
        r"\bepilator\b", r"foot\s*spa", r"foot\s*massager", r"\bmassager\b",
        r"curling\s*iron", r"\bcurler\b",
        # Comfort / power
        r"room\s*heater", r"\bheater\b", r"\bblower\b",
        r"\binverter\b", r"\bups\b", r"stabili[sz]er", r"\bgenerator\b",
    ]),
    ("HE", [
        r"\btv\b", r"\bled\b", r"sound\s*bar", r"soundbar", r"\bspeaker\b",
        r"home\s*theat(er|re)",
    ]),
    ("IT", [
        r"\bcomputer\b", r"\bdesktop\b", r"\blaptop\b", r"\bprinter\b", r"\bcpu\b",
    ]),
    ("DC", [
        r"\bcamera\b", r"\blens\b", r"\bdslr\b", r"\bgopro\b",
    ]),
    ("MH", [
        r"\bvivo\b", r"\boppo\b", r"\brealme\b", r"\bredmi\b", r"\biphone\b", r"\bapple\b",
        r"\bsamsung\b", r"\bmotorola\b", r"\bmoto\b", r"\biqoo\b", r"\bnothing\b",
        r"google\s*pixel", r"\bpixel\b", r"\bxiaomi\b", r"\bmi\b",
    ]),
]
DIVISION_NAMES = {
    "HA": "Home Appliance",
    "HE": "Home Entertainment",
    "MH": "Mobile",
    "IT": "Computer/IT",
    "DC": "Digital Camera",
    "UNCATEGORIZED": "Uncategorized",
}

BRAND_LIST = [
    "Daikin", "Blue Star", "O General", "Voltas", "Hitachi", "LG", "Samsung",
    "Godrej", "Haier", "Panasonic", "Sony", "Mitashi", "IFB", "Bosch",
    "Carrier", "Lloyd", "Bajaj","Havells", "Kent", 
    "Faber", "V-Guard", "Philips", "Luminous", "Dyson", "Vivo", "Oppo", "Realme", "Redmi", "Xiaomi", "iPhone", "Apple", "Motorola", "IQOO",
    "Nothing", "Google Pixel", "Google", "OnePlus", "Poco", "Tecno", "Infinix", "Itel", "Honor",
    "HP", "Dell", "Lenovo", "Acer", "Nikon"
]
# Longest names first, so e.g. "Google Pixel" matches before bare "Google".
BRAND_LIST.sort(key=len, reverse=True)

AC_ITEM_RE = re.compile(r"\b(s\.?a\.?c|w\.?a\.?c|ac)\b", re.IGNORECASE)
AC_ROLE_RE = re.compile(r"\b(ID|OD|INDOOR|OUTDOOR)\b", re.IGNORECASE)


def detect_division_code(item_text: str) -> str:
    for code, patterns in DIVISION_KEYWORD_RULES:
        for pattern in patterns:
            if re.search(pattern, item_text, re.IGNORECASE):
                return code
    return "UNCATEGORIZED"


def detect_brand(item_text: str) -> Optional[str]:
    for brand in BRAND_LIST:
        if re.search(r"\b" + re.escape(brand) + r"\b", item_text, re.IGNORECASE):
            return brand
    first_word = re.match(r"[A-Za-z0-9]+", item_text.strip())
    return first_word.group(0).title() if first_word else None


def detect_ac_role(item_text: str) -> Optional[str]:
    if not AC_ITEM_RE.search(item_text):
        return None
    match = AC_ROLE_RE.search(item_text)
    if not match:
        return None
    token = match.group(1).upper()
    return "ID" if token in ("ID", "INDOOR") else "OD"


def ac_merge_key(row: dict) -> tuple:
    """Groups an AC indoor row with its matching outdoor row. Prefers the
    invoice/voucher number (most reliable - both halves of one AC sale are
    almost always billed on the same voucher). Falls back to the item text
    with the ID/OD marker and the model code right after it stripped out,
    since that code is the one thing that legitimately differs between an
    indoor and outdoor unit of the same sale."""
    if row.get("vch_no"):
        return ("VCH", row["vch_no"].strip().upper(), row["sale_date"].isoformat())

    tokens = row["item"].split()
    role_idx = None
    for i, tok in enumerate(tokens):
        if AC_ROLE_RE.fullmatch(tok.strip(".,")):
            role_idx = i
            break
    if role_idx is not None:
        remove_indexes = {role_idx}
        if role_idx + 1 < len(tokens):
            remove_indexes.add(role_idx + 1)
        tokens = [t for i, t in enumerate(tokens) if i not in remove_indexes]
    base = " ".join(tokens).upper().strip()
    return ("ITEM", base, row["sale_date"].isoformat())


def merge_ac_pairs(rows: List[dict]) -> List[dict]:
    groups = defaultdict(list)
    passthrough = []
    for row in rows:
        if row.get("ac_role"):
            groups[ac_merge_key(row)].append(row)
        else:
            passthrough.append(row)

    merged_out = list(passthrough)
    for key, group in groups.items():
        ids = [r for r in group if r["ac_role"] == "ID"]
        ods = [r for r in group if r["ac_role"] == "OD"]
        if len(ids) == 1 and len(ods) == 1:
            id_row, od_row = ids[0], ods[0]
            combined = dict(id_row)
            combined["sales_amt"] = (id_row.get("sales_amt") or 0.0) + (od_row.get("sales_amt") or 0.0)
            combined["cost_amt"] = (id_row.get("cost_amt") or 0.0) + (od_row.get("cost_amt") or 0.0)
            combined["profit_loss"] = combined["sales_amt"] - combined["cost_amt"]
            combined["qty"] = id_row.get("qty") if id_row.get("qty") else od_row.get("qty")
            combined["item"] = f"{id_row['item']}  +  {od_row['item']} (ID+OD combined)"
            combined["merged"] = True
            combined["note"] = "Indoor + outdoor unit combined from 2 rows of the same AC sale."
            merged_out.append(combined)
        elif len(ids) == 1 and not ods:
            ids[0]["note"] = "Indoor unit only - no matching outdoor row found, shown individually."
            merged_out.append(ids[0])
        elif len(ods) == 1 and not ids:
            ods[0]["note"] = "Outdoor unit only - no matching indoor row found, shown individually."
            merged_out.append(ods[0])
        else:
            for r in group:
                r["note"] = "AC indoor/outdoor row could not be uniquely auto-matched - review manually."
                merged_out.append(r)
    return merged_out


def build_staged_rows(parsed_rows: List[dict]) -> List[dict]:
    staged = []
    for r in parsed_rows:
        row = dict(r)
        row["division"] = detect_division_code(row["item"])
        row["brand"] = detect_brand(row["item"])
        row["ac_role"] = detect_ac_role(row["item"])
        row["merged"] = False
        row["note"] = None
        staged.append(row)
    merged = merge_ac_pairs(staged)
    for i, row in enumerate(merged):
        row["row_id"] = i
    return merged


def serialize_staged_row(row: dict) -> dict:
    return {
        "row_id": row["row_id"],
        "sale_date": row["sale_date"].isoformat() if row.get("sale_date") else None,
        "item": row["item"],
        "division": row["division"],
        "division_name": DIVISION_NAMES.get(row["division"], row["division"]),
        "brand": row.get("brand"),
        "qty": row.get("qty"),
        "sales_amt": round(row.get("sales_amt") or 0.0, 2),
        "cost_amt": round(row.get("cost_amt") or 0.0, 2),
        "profit_loss": round(row.get("profit_loss") or 0.0, 2),
        "ac_role": row.get("ac_role"),
        "merged": bool(row.get("merged")),
        "note": row.get("note"),
        "source_sheet": row.get("source_sheet"),
    }


def build_staging_summary(rows: List[dict]) -> dict:
    by_division = defaultdict(lambda: {"count": 0, "sales": 0.0})
    merged_count = 0
    flagged_count = 0
    brand_counts = defaultdict(int)
    for row in rows:
        stat = by_division[row["division"]]
        stat["count"] += 1
        stat["sales"] += row.get("sales_amt") or 0.0
        if row.get("merged"):
            merged_count += 1
        elif row.get("note"):
            flagged_count += 1
        if row.get("brand"):
            brand_counts[row["brand"]] += 1

    division_summary = [
        {"division": code, "division_name": DIVISION_NAMES.get(code, code), "count": s["count"], "sales": round(s["sales"], 2)}
        for code, s in sorted(by_division.items(), key=lambda x: -x[1]["sales"])
    ]
    return {
        "total_rows": len(rows),
        "uncategorized_count": by_division.get("UNCATEGORIZED", {}).get("count", 0),
        "merged_ac_rows": merged_count,
        "flagged_ac_rows": flagged_count,
        "division_summary": division_summary,
        "brands_detected": len(brand_counts),
        "top_brands": [{"brand": b, "count": c} for b, c in sorted(brand_counts.items(), key=lambda x: -x[1])[:12]],
    }


# In-memory staging area: a file that's been parsed + auto-classified but not
# yet written to the live dashboard tables, so the Admin can review, bulk
# re-assign divisions, download the cleaned file, and only then commit it (or
# discard it and try a different file). Keyed by a random token; capped so a
# few abandoned uploads in a row don't grow this unbounded.
ANALYTICS_STAGING: dict = {}
ANALYTICS_STAGING_LIMIT = 5


def _staging_get_or_404(token: str) -> dict:
    entry = ANALYTICS_STAGING.get(token)
    if not entry:
        raise HTTPException(status_code=404, detail="This review session has expired. Please upload the file again.")
    return entry


@app.post("/api/analytics/stage")
def stage_analytics_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    filename = file.filename or "uploaded_file"
    raw_content = file.file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    parsed_rows = parse_analytics_file(filename, raw_content)
    if not parsed_rows:
        raise HTTPException(
            status_code=400,
            detail="No readable sales rows found. Ensure the file has columns for Date, Item, Sales Amt, Cost Amt and Profit/Loss (Division, Qty, and Vch No are optional).",
        )

    staged_rows = build_staged_rows(parsed_rows)

    if len(ANALYTICS_STAGING) >= ANALYTICS_STAGING_LIMIT:
        oldest_token = next(iter(ANALYTICS_STAGING))
        ANALYTICS_STAGING.pop(oldest_token, None)

    token = uuid4().hex
    ANALYTICS_STAGING[token] = {
        "rows": staged_rows,
        "filename": filename,
        "created_by": current_user.username,
    }

    return {
        "staging_token": token,
        "file_name": filename,
        "rows": [serialize_staged_row(r) for r in staged_rows],
        "summary": build_staging_summary(staged_rows),
    }


@app.post("/api/analytics/stage/{token}/reassign")
def reassign_staged_rows(
    token: str,
    payload: schemas.AnalyticsReassignRequest,
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    entry = _staging_get_or_404(token)
    rows = entry["rows"]
    new_division = (payload.division or "").strip().upper() or "UNCATEGORIZED"
    row_ids = set(payload.row_ids)
    updated = 0
    for row in rows:
        if row["row_id"] in row_ids:
            row["division"] = new_division
            updated += 1

    return {
        "message": f"Reassigned {updated} row(s) to {DIVISION_NAMES.get(new_division, new_division)}.",
        "rows": [serialize_staged_row(r) for r in rows],
        "summary": build_staging_summary(rows),
    }


@app.get("/api/analytics/stage/{token}/download")
def download_staged_file(
    token: str,
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    entry = _staging_get_or_404(token)
    rows = entry["rows"]

    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Date", "Item", "Division", "Brand", "Qty", "Sales Amt", "Cost Amt", "Profit/Loss", "AC Role", "Note"])
    for row in rows:
        writer.writerow([
            row["sale_date"].strftime("%d-%m-%Y") if row.get("sale_date") else "",
            row["item"],
            DIVISION_NAMES.get(row["division"], row["division"]),
            row.get("brand") or "",
            row.get("qty") if row.get("qty") is not None else "",
            round(row.get("sales_amt") or 0.0, 2),
            round(row.get("cost_amt") or 0.0, 2),
            round(row.get("profit_loss") or 0.0, 2),
            row.get("ac_role") or "",
            row.get("note") or "",
        ])

    base_name = entry["filename"].rsplit(".", 1)[0] if "." in entry["filename"] else entry["filename"]
    # utf-8-sig BOM so Excel opens the ₹/non-ASCII text correctly on Windows.
    csv_bytes = "\ufeff" + buffer.getvalue()
    return Response(
        content=csv_bytes.encode("utf-8"),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="cleaned_{base_name}.csv"'},
    )


@app.delete("/api/analytics/stage/{token}")
def discard_staged_file(
    token: str,
    current_user: models.User = Depends(auth.require_roles("Admin")),
):
    ANALYTICS_STAGING.pop(token, None)
    return {"message": "Discarded"}


# ============================================================
# AGEING STOCK ANALYSIS
# Upload a workbook with an "All Data" sheet (every item, ageing-bucketed)
# plus one sheet per physical location. We parse "All Data" as the master
# item list, then match each item by name against every location sheet to
# work out where it's physically sitting, and classify it into
# Category/Brand for the report.
# ============================================================

# Canonical location code -> (display name, sheet-title aliases to match).
# Matching is done on the normalized (lowercased, non-alnum-stripped) sheet
# title, so "MWH 1", "MWH1", "Warehouse" etc. all resolve to MWH.
AGEING_LOCATION_DEFINITIONS = {
    "ALM": {"name": "Alambagh", "aliases": {"alm", "alambagh"}},
    "HZT": {"name": "Hazratganj", "aliases": {"hzt", "hazratganj"}},
    "ASH": {"name": "Ashiyana", "aliases": {"ash", "ashiyana"}},
    "GNG": {"name": "Gomtinagar", "aliases": {"gng", "gomtinagar"}},
    "VKN": {"name": "Vikas Nagar", "aliases": {"vkn", "vikasnagar"}},
    "MWH": {"name": "Warehouse", "aliases": {"mwh", "mwh1", "warehouse", "mainwarehouse"}},
    # PWH and Vault are intentionally NOT listed here - they're handled by
    # AGEING_EXCLUDED_LOCATION_ALIASES below instead of being tracked as a
    # normal location.
}

# Sheet-title aliases (matched the same normalized way as
# AGEING_LOCATION_DEFINITIONS) for locations that must be excluded from
# Ageing Stock Analysis altogether. Per Admin request, PWH and Vault stock
# is not considered by this report at all: it's not enough to just leave
# these out of AGEING_LOCATION_DEFINITIONS above, because the "All Data"
# master sheet lists an item's total closing qty/ageing buckets across ALL
# locations, including PWH and Vault - so an item sitting only in PWH or
# Vault would otherwise still show up in the report with real numbers, just
# with no location tag. parse_ageing_stock_workbook() reads these sheets
# (when present) purely to build the set of item keys to drop from the
# "All Data" rows entirely, so no PWH/Vault item - and no PWH/Vault
# quantity - ever reaches the report, any export, or the on-screen table.
AGEING_EXCLUDED_LOCATION_ALIASES = {
    "PWH": {"pwh"},
    "VAULT": {"vault"},
}
AGEING_ALL_DATA_SHEET_ALIASES = {"alldata", "all", "master", "masterdata"}

# Category Master workbook (separate upload - see
# parse_ageing_category_master_workbook) - one sheet per category, each
# listing every item name (across all brands) that belongs there. Sheet
# titles are matched the same normalized way as the location sheets above,
# so "HA", "Home Appliances", "Mobile", "Computer" etc. all resolve.
AGEING_CATEGORY_MASTER_SHEET_ALIASES = {
    "HA": {"ha", "homeappliances", "homeappliance"},
    "HE": {"he", "homeentertainment"},
    "MH": {"mobile", "mh", "mobiles", "mobilephones"},
    "IT": {"computer", "computers", "it", "informationtechnology"},
    "DC": {"dc", "digitalcamera", "digitalcameras", "camera", "cameras"},
}
AGEING_CATEGORY_MASTER_SHEET_NAMES = {
    "HA": "Home Appliances", "HE": "Home Entertainment",
    "MH": "Mobile", "IT": "Computer",
    "DC": "Digital Camera",
    "OTHER": "Accessories",
}

# Header aliases for the "All Data" (and location) sheets - matched the
# same way as ANALYTICS_HEADER_ALIASES above, via normalize_header_name().
AGEING_HEADER_ALIASES = {
    # "groupname" covers Tally's alternate "Stock Ageing Analysis" export
    # (grouped by "All Groups" instead of "All Items"). Both export styles
    # use the same row shape (name + Closing Qty + Unit + age buckets), so
    # treating "Group Name" as the item column lets that export style
    # parse identically to the normal "Item Details" export, across All
    # Data and every location/MWH sheet.
    "item_details": {"itemdetails", "item", "itemname", "description", "product", "groupname"},
    "closing_qty": {"closingqty", "closingquantity", "qty", "quantity"},
    "unit": {"unit", "uom"},
    # So the Ageing Stock Analysis search box can match a typed-in model
    # number even when it's its own column rather than embedded in Item
    # Details - see parse_ageing_stock_workbook() and the search filter in
    # compute_ageing_stock_report().
    "model_no": {"modelno", "model", "modelnumber", "modelnos", "modno"},
    "age_0_60": {"060days", "060", "0to60days"},
    "age_61_90": {"6190days", "6190", "61to90days"},
    "age_91_150": {"91150days", "91150", "91to150days"},
    "age_151_180": {"151180days", "151180", "151to180days"},
    "age_181_365": {"181365days", "181365", "181to365days"},
    "age_366_plus": {"366days", "366plusdays", "366", "gte366days", "morethan365days", "above365days"},
}


def ageing_canonical_column(normalized_header: str) -> Optional[str]:
    for canonical, aliases in AGEING_HEADER_ALIASES.items():
        if normalized_header in aliases:
            return canonical
    return None


# A handful of brands get written inconsistently across sheets (e.g. "Blue
# Star" vs "Bluestar"). Collapsing these to one canonical spelling here -
# before classification AND before location-sheet matching - means both
# spellings land in the same brand bucket instead of splitting into two.
# Add more (SOURCE -> CANONICAL, both upper-cased, whole-word) here if
# another brand shows the same problem.
AGEING_BRAND_TEXT_ALIASES = {
    r"\bBLUE STAR\b": "BLUESTAR",
}


def normalize_ageing_item_key(value) -> str:
    """Normalizes an Item Details string for cross-sheet matching: upper-
    cased, punctuation/extra-whitespace collapsed. Two sheets writing the
    same item slightly differently (extra spaces, a stray hyphen) still
    match; genuinely different items still don't."""
    text = re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()
    text = re.sub(r"\s+", " ", text)
    for pattern, replacement in AGEING_BRAND_TEXT_ALIASES.items():
        text = re.sub(pattern, replacement, text)
    return text


def is_ageing_nos_unit(unit_value) -> bool:
    """True only for the "Nos." unit (piece-count stock). Per Admin
    request, Ageing Stock Analysis only ever tracks items sold/counted in
    Nos. - any other unit (Mtr, Kg, Ltr, Set, Box, etc.), and any row with
    a blank/missing unit, is left out of the report entirely rather than
    being classified into a category. Comparison strips punctuation and
    case so "Nos.", "NOS", "nos" etc. all still match."""
    text = re.sub(r"[^A-Z]", "", str(unit_value or "").upper())
    return text in {"NOS", "UNITS", "UNIT"}


def find_ageing_header_row(table_rows: List[List]) -> tuple:
    scan_limit = min(len(table_rows), 15)
    best_index, best_score = 0, -1
    for row_index in range(scan_limit):
        row = table_rows[row_index] or []
        normalized = [normalize_header_name(cell) for cell in row]
        canonicals = {ageing_canonical_column(name) for name in normalized if ageing_canonical_column(name)}
        canonicals.discard(None)
        score = len(canonicals)
        if "item_details" in canonicals:
            score += 2
        if "closing_qty" in canonicals:
            score += 1
        if score > best_score:
            best_index, best_score = row_index, score
    return best_index, best_score


def read_ageing_sheet_rows(worksheet) -> List[dict]:
    """Reads one worksheet (All Data or a location sheet) into row dicts
    keyed by the canonical AGEING_HEADER_ALIASES names. Returns [] if the
    sheet doesn't look like an ageing-stock table at all."""
    raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    if not raw_rows:
        return []
    header_index, score = find_ageing_header_row(raw_rows)
    if score < 2:
        return []
    headers = [normalize_header_name(cell) for cell in raw_rows[header_index]]
    canonical_headers = [ageing_canonical_column(h) for h in headers]
    if "item_details" not in canonical_headers:
        return []

    out = []
    for row in raw_rows[header_index + 1:]:
        row_dict = {}
        for idx, canonical in enumerate(canonical_headers):
            if canonical:
                row_dict[canonical] = row[idx] if idx < len(row) else None
        item_details = str(row_dict.get("item_details") or "").strip()
        # Skip blank rows and the sheet's own "Total"/"Grand Total" rows.
        if not item_details or item_details.strip().lower() in {"total", "grand total"}:
            continue
        out.append(row_dict)
    return out


def detect_ageing_location_code(worksheet) -> Optional[str]:
    """Resolve a location from its tab name or Tally Material Centre label."""
    candidates = [str(worksheet.title or "")]
    for row in worksheet.iter_rows(min_row=1, max_row=10, max_col=6, values_only=True):
        for value in row:
            text = str(value or "").strip()
            if "material" in text.lower() and "centre" in text.lower():
                candidates.append(text)

    definitions = {**AGEING_LOCATION_DEFINITIONS, **AGEING_EXCLUDED_LOCATION_ALIASES}
    for candidate in candidates:
        normalized = re.sub(r"[^a-z0-9]", "", candidate.lower())
        for code, definition in definitions.items():
            aliases = definition["aliases"] if isinstance(definition, dict) else definition
            if normalized in aliases or any(normalized.endswith(alias) for alias in aliases):
                return code
    return None


# Category classification keyword rules, checked in this order. Each is
# (category_code, category_name, [substrings to look for in the
# normalized/uppercased item text]). Matched top-down - HE/IT/HA product-
# type keywords are checked before falling back to a brand's usual
# category, since the same brand (e.g. Samsung, Lenovo) sells across
# multiple categories and the product name is the more reliable signal.
AGEING_CATEGORY_KEYWORD_RULES = [
    ("HE", "Home Entertainment", [
        " LED ", "LED ", " TV ", "TELEVISION", "HOME THEATRE", "HOME THEATER",
        "SOUNDBAR", "SOUND BAR", " QLED", "OLED ", " REMOTE",
        # Speakers (Bluetooth, party/home, smart speakers) are Home
        # Entertainment regardless of brand - without this, a speaker from
        # a brand whose *other* products are normally HA (e.g. Aisen,
        # mostly coolers/geysers) would fall through to that brand's
        # default HA category via the Brand Master fallback instead.
        "SPEAKER",
        # Smart-speaker/streaming product lines that don't literally say
        # "speaker" in the name (e.g. "Amazon ECHO 4th Gen", "Google
        # Speaker Chromecast"). Padded to whole-word to avoid any
        # accidental partial-word match.
        " ECHO ", "CHROMECAST",
        # Samsung's Smart Monitor/Display line (e.g. "Samsung Display 32
        # QM32C", "Samsung Display 43 QB43C") is a TV-like all-in-one
        # display, filed under Home Entertainment per Admin request.
        # Scoped to "SAMSUNG DISPLAY" (not bare "DISPLAY") so an actual
        # computer monitor from another brand - which has no dedicated
        # "DISPLAY" keyword and relies on the "MONITOR" keyword below
        # instead - is never pulled into Home Entertainment by mistake.
        "SAMSUNG DISPLAY",
    ]),
    ("IT", "Computer", [
        "LAPTOP", "DESKTOP", "MACBOOK", " MBA ", " MBP ", "NOTEBOOK",
        "PRINTER", "MONITOR", " AIO ",
        # Any other Apple "Mac" product line (Mac mini, Mac Studio, and
        # any oddly-named Mac SKU) is a computer, not a phone.
        " MAC ",
        # NOTE: Tabs/Tablets/iPad are intentionally NOT here - per Admin
        # request, all Tab/Pad devices (any brand) are grouped under
        # Mobile instead of Computer/IT. See the MH keyword list below.
    ]),
    ("HA", "Home Appliances", [
        "REFRIGERATOR", "FRIDGE", "WASHING MACHINE", "MICROWAVE", "AIR CONDITIONER",
        " AC ", "OVEN", "DISHWASHER", "CHIMNEY", "GEYSER", "WATER HEATER", "COOLER",
        # " REF " is the company's own shorthand for Refrigerator (as used
        # throughout ageing_category_master.json, e.g. "Godrej Ref ...").
        " REF ",
        # Short-form/abbreviated appliance codes used in item names across
        # brands (e.g. "Samsung WM ...", "LG SAC ID ...", "IFB MWO ...",
        # "Havells Mixer Grinder ...", "Voltas WAC ..."). Without these, an
        # item using the short form (rather than the full word above) that
        # isn't an exact hit in ageing_category_master.json falls through
        # to the item's brand's *default* category, which can wrongly be
        # HE for a brand that's normally filed there for other products
        # (e.g. TVs) - these keywords make sure the product type itself
        # decides the category, same as the full-word keywords above.
        " MW ", " WM ", " SAC ", " MWO ", " WAC ", "MIXER", "GRINDER",
        "WATER PURIFIER", "VACUUM CLEANER",
        # Additional appliance product types/product-lines not covered
        # above (e.g. "Sunflame LPG Cooktop...", "Sunflame Built In HOB
        # ...", "Sharp Air Purifier...", "Eureka Aquaguard...", "Haier Dry
        # (Iron) Gift", "Amaze (Inverter) AQ675") - without these, a new
        # item not yet listed in ageing_category_master.json falls
        # through to the Brand Master default or Unclassified/Other
        # instead of landing in Home Appliances.
        "COOKTOP", " HOB ", "AIR PURIFIER", "AQUAGUARD", "INVERTER",
        "DRY IRON", "STEAM IRON",
    ]),
    ("MH", "Mobile", [
        "IPHONE", "SMARTPHONE",
        # Per Admin request: every Tab/Pad device, regardless of brand
        # (Samsung Tab, Lenovo Tab, Xiaomi Pad, iPad, etc.), is grouped
        # under Mobile rather than Computer/IT. "IPAD"/"I PAD" are bare
        # substrings (no padding) since "iPad" normalizes to one token
        # with no internal space; " TAB "/" TABLET"/" PAD " stay padded
        # since those are always written as standalone words.
        "IPAD", "I PAD", " TAB ", " TABLET", " PAD ",
    ]),
    ("DC", "Digital Camera", [
        # New category added per Admin request. Any camera - mirrorless,
        # DSLR, point-and-shoot, camcorder - regardless of brand, lands
        # here. Bare "CAMERA" is safe: "SECURITY CAMERA" is checked (and
        # wins) via AGEING_ACCESSORY_KEYWORDS before this rule loop is
        # ever reached, so a security/CCTV camera still correctly lands
        # in Accessories, not Digital Camera.
        "CAMERA", "MIRRORLESS", "DSLR", "CAMCORDER",
    ]),
]

# Fixed display order for the report - core categories in this order, then
# Other last, regardless of how the alphabetical item_count sort would fall.
AGEING_CATEGORY_ORDER = ["HA", "HE", "MH", "IT", "DC", "OTHER"]

# Everything that is an accessory/consumable rather than a sellable core
# unit - chargers, cables, cases, AC installation bits, etc. - is bucketed
# into "Other" instead of its own row under HA/HE/IT/Mobile, no matter
# which brand or product-type keyword it would otherwise match. Checked
# BEFORE the category rules above and the brand hints below, so an
# accessory from a mobile/TV/AC brand still lands in Other. Substrings are
# intentionally broad (e.g. "ACCESSOR" catches both ACCESSORY and
# ACCESSORIES) - extend this list as new accessory item names show up.
AGEING_ACCESSORY_KEYWORDS = [
    "ADAPTOR", "ADAPTER", "BATTERY", "BATTERIES", "CABLE", "BAG PACK",
    "BACKPACK", "BACK PACK", "ACCESSOR", "CHARGER", "POWER BANK", "POWERBANK",
    "TEMPERED GLASS", "SCREEN GUARD", "SCREEN PROTECTOR", "FLIP COVER",
    "BACK COVER", "MOBILE COVER", "PHONE CASE", "POUCH", "STRAP",
    "PENDRIVE", "PEN DRIVE", "USB HUB", "HDMI CABLE",
    # Any cover, regardless of what it's a cover for (phone, tablet, watch,
    # etc.) or how the rest of the name is worded (e.g. "Samsung Cover
    # Flip 4 GP-TOF721AM7RI", "Samsung (Cover) Silicone S25 Edge White") -
    # per Admin request, always an accessory. Bare/padded so it still
    # catches names where "Cover" isn't glued to another word from the
    # FLIP COVER/BACK COVER/MOBILE COVER phrases above.
    " COVER ",
    # Samsung's "The Frame" TV replacement bezel/frame kits (e.g. "Samsung
    # Frame PVC VG SCFA50WTBXL") - a decorative accessory sold separately
    # from the TV itself, not a TV unit. Scoped to the "Frame PVC" phrase
    # (rather than bare "Frame") so an actual Frame-series TV item, which
    # would say "LED"/"QLED" and belongs in Home Entertainment, is never
    # caught by this rule.
    "SAMSUNG FRAME PVC",
    # Samsung Gear wearables (smartwatches, earbuds, etc. under the Gear
    # product line, e.g. "Samsung Gear R 7000") - a wearable accessory,
    # not a phone, per Admin request.
    "SAMSUNG GEAR",
    # Wearables/audio/input accessories - not core sellable Mobile units
    # even when sold under a phone brand (e.g. "Samsung Earphones...",
    # "Apple Watch...", "Apple Airpods...", "Lenovo Mouse...", "Vivo
    # Buds...", "CMF Buds..."). Per Admin request, these always land in
    # Accessories regardless of brand. " WATCH" is padded with a leading
    # space so it only matches the whole word (Smart Watch, Apple Watch)
    # and not an unrelated substring.
    "EARPHONE", "NECKBAND", "NECK BAND", " WATCH", "AIRPOD", "BUDS",
    "KEYBOARD", "MOUSE", "PENCIL", "SMART BAND", "SMARTBAND",
    # Mounting/installation hardware - not a phone, TV, or appliance unit
    # itself even when named after the brand it's mounted to.
    "WALL MOUNT",
    # Branded promotional/gift-with-purchase items (e.g. "Vivo Gift
    # Bottle", "Vivo Gift Tifin Box", "Vivo Pen") - not a sellable phone
    # unit even though named after a phone brand. " PEN " padded so it
    # only matches the standalone word, not "PENCIL"/"PENDRIVE"/"OPEN".
    # NOTE: bare "GIFT" is intentionally NOT included here - it's too
    # broad and was wrongly catching genuine appliances that happen to be
    # sold as a gift-boxed pack (e.g. "Haier Dry (Iron) Gift"), routing a
    # real Home Appliances item into Accessories before it ever got a
    # chance to match the HA "DRY IRON" keyword below. "BOTTLE"/"TIFIN"/
    # "TIFFIN"/" PEN " alone already cover the documented promotional-gift
    # examples above.
    "BOTTLE", "TIFIN", "TIFFIN", " PEN ",
    # Networking/charging accessories sold under a phone-making brand
    # (e.g. "Apple Router WiFi MC 414 HN/A", "Apple Magsafe 2 MD 504
    # ZM/A") - neither is a phone unit itself.
    "ROUTER", "MAGSAFE",
    # AC-specific accessories/consumables and installation material.
    "STABILIZER", "INSTALLATION KIT", "COPPER PIPE", "DRAIN PIPE",
    "OUTDOOR BRACKET", "AC COVER", "AC ACCESSOR", "AC STAND", "VOLTAGE GUARD",
    # Small accessory product lines sold under a phone/appliance brand -
    # cases, headphones, fitness/smart bands, garment steamers, and
    # security cameras are never the core sellable unit (a phone, TV, or
    # major appliance) even when named after that brand, so they always
    # land in Accessories/Other rather than that brand's usual category.
    # "CASE" is intentionally a bare substring (not padded) because some
    # item names have no space before the trailing SKU code (e.g. "iPhone
    # 12 Pro Max Silicone CaseMHLG3ZM/A"), so a space-padded " CASE " check
    # would miss it; "CASE" alone still doesn't collide with any real
    # product word in this catalog (e.g. "CASSETTE" AC units don't contain
    # "CASE" as a substring). " BAND " stays padded since that name is
    # always written as a standalone word. RAM/memory modules (e.g. "RAM
    # DDR3 8GB Laptop") are handled separately below by
    # _is_ram_memory_module(), NOT as a bare " RAM " substring here - some
    # AC model numbers (e.g. "Akabishi SAC ID 1.5T RAM-AE18VG-X1-KT",
    # where the hyphens normalize to spaces) start with "RAM" too, and a
    # bare substring match would wrongly route those into Accessories
    # instead of Home Appliances/AC.
    "CASE", "HEADPHONE", " BAND ", "GARMENT STEAMER", "SECURITY CAMERA",
    # Cleaning/maintenance kits (e.g. "Computer/Laptop Cleaning Kit") are a
    # consumable accessory, not a computer itself - without this, "LAPTOP"
    # in the name would match the IT keyword rule below and, with no real
    # registered brand to match against, the item's first word ("Computer")
    # would get used as a bogus brand grouping instead of landing in
    # Accessories/Other where it belongs.
    "CLEANING KIT",
    # Small laptop/desktop accessories sold under a Computer/IT brand
    # (e.g. "Laptop Keyguard", "Laptop Fan", "HP Flash Drive 3.2 USB
    # 64GB", "HP Cartridge 680 Tri Colour F6V26AA", "Lenovo Headset H110
    # GXD1P46879", "Lenovo Smart Clock ZA4R0023IN", "Lenovo USB C HUB 150
    # GX91M7394") - not a core sellable laptop/desktop unit itself, even
    # under a registered Computer brand like HP/Lenovo. Without these,
    # "Laptop Keyguard"/"Laptop Fan" would have no real brand to match and
    # get grouped under a bogus "Laptop" brand via the IT first-word
    # fallback, and the rest would get counted as full HP/Lenovo Computer
    # stock instead of Accessories.
    "KEYGUARD", "LAPTOP FAN", "FLASH DRIVE", "CARTRIDGE", "HEADSET",
    "SMART CLOCK", "USB C HUB",
    # Branded gift/promotional items that aren't the core sellable unit
    # (e.g. "LG Glass Bowl", "LG Duffele Bag (Wildcraft) Gift", "Haier
    # Platinum Coupon", "Haier Bowl Sets") - same reasoning as the
    # BOTTLE/TIFIN/" PEN " gift items above. Bare "BOWL" (not just "GLASS
    # BOWL") covers any bowl-set gift regardless of material; "COUPON"
    # covers gift vouchers/coupons issued under a brand name.
    "GLASS BOWL", "BOWL", "DUFFEL", "DUFFLE", "COUPON",
]


def _is_ram_memory_module(padded: str) -> bool:
    """True only for an actual RAM/memory module (e.g. "RAM DDR3 8GB
    Laptop", "8GB DDR4 RAM"), which is a component, not the appliance/
    computer itself - routed to Accessories the same as the substring
    list above. Requires "RAM" AND a memory-spec word (DDR/GB) together,
    not just the bare word "RAM" - some AC model numbers also contain
    "RAM" as a standalone token once hyphens normalize to spaces (e.g.
    "Akabishi SAC ID 1.5T RAM-AE18VG-X1-KT" -> "... RAM AE18VG X1 KT"),
    and those are real Home Appliances/AC units, not memory modules."""
    return " RAM " in padded and ("DDR" in padded or "GB" in padded)



# Guard applied ONLY to items that would otherwise land in "MH" (Mobile /
# Handset) - keeps that category strictly to actual mobile phone units,
# no matter which rule put them there (including a bad/legacy entry in
# the bundled ageing_category_master.json, which is checked first and
# would otherwise be unoverridable). Any MH classification is re-routed:
#  - to Home Appliances if it's really an appliance sold under a phone
#    brand (e.g. "Samsung WM ..." = Washing Machine, "Samsung Ref ..." =
#    Refrigerator, not a phone),
#  - to Information Technology if it's really a laptop/desktop/monitor/
#    printer sold under a phone brand (tablets stay in Mobile - see the
#    IT demote list below),
#  - to Other/Accessories for wearables, audio, input, promotional/gift,
#    and consumable accessories sold under a phone brand.
AGEING_MH_DEMOTE_TO_HA_KEYWORDS = [
    " WM ", " REF ", "WASHING MACHINE", "REFRIGERATOR", "MICROWAVE", "CHIMNEY",
    "AIR CONDITIONER", " AC ",
]
AGEING_MH_DEMOTE_TO_IT_KEYWORDS = [
    "LAPTOP", "NOTEBOOK", "MACBOOK", "DESKTOP", "MONITOR", "PRINTER",
    " MAC ",
    # NOTE: " TAB "/" TABLET" intentionally removed - Tab/Pad devices now
    # belong in Mobile (see the MH keyword rule above), so they should no
    # longer be demoted back out to Computer/IT here.
]
AGEING_MH_DEMOTE_TO_OTHER_KEYWORDS = [
    "BUDS", "EARBUD", "NECKBAND", "NECK BAND", " WATCH", "SMART BAND",
    "SMARTBAND", "WALL MOUNT", "MOUSE", "KEYBOARD", "AIRPOD", "EARPHONE",
    "PENCIL", "CHARGER", "CABLE", "ADAPTOR", "ADAPTER", "POWER BANK",
    "POWERBANK", "TEMPERED GLASS", "SCREEN GUARD", "SCREEN PROTECTOR",
    "FLIP COVER", "BACK COVER", "MOBILE COVER", "PHONE CASE", "POUCH",
    "STRAP", "BATTERY", "BATTERIES", "GIFT", "BOTTLE", "TIFIN", "TIFFIN",
    " PEN ", "ROUTER", "MAGSAFE",
]


def _enforce_mobile_category_is_phones_only(result: dict, padded: str, brand_lookup: List[tuple]) -> dict:
    if result.get("category_code") != "MH":
        return result
    if any(keyword in padded for keyword in AGEING_MH_DEMOTE_TO_HA_KEYWORDS):
        fixed = dict(result)
        fixed["category_code"] = "HA"
        fixed["category_name"] = "Home Appliances"
        fixed["classification_source"] = "Mobile Guard"
        return fixed
    if any(keyword in padded for keyword in AGEING_MH_DEMOTE_TO_IT_KEYWORDS):
        fixed = dict(result)
        fixed["category_code"] = "IT"
        fixed["category_name"] = "Computer"
        fixed["classification_source"] = "Mobile Guard"
        return fixed
    if any(keyword in padded for keyword in AGEING_MH_DEMOTE_TO_OTHER_KEYWORDS):
        fixed = dict(result)
        fixed["category_code"] = "OTHER"
        fixed["category_name"] = "Accessories"
        fixed["brand_name"] = _resolve_known_brand_only(padded, brand_lookup)
        fixed["classification_source"] = "Mobile Guard"
        return fixed
    return result


# LED items default to Home Entertainment (TVs) via the " LED "/"LED "
# keyword rule above. However HP, Dell, and Lenovo's LED-branded lineup
# here is computer monitors, not TVs (e.g. "HP LED 21.5 Inch M22F",
# "Lenovo LED 27 Inch 67B6GAC1IN") - so per Admin request, an LED item
# under one of these three brands always belongs in Computer/IT instead,
# regardless of the general LED keyword rule. Every other brand's LED item
# (Samsung, LG, Sony, TCL, etc.) still lands in Home Entertainment as
# before.
AGEING_LED_COMPUTER_BRANDS = {"HP", "DELL", "LENOVO"}


def _enforce_led_computer_brands_are_computer(result: dict, padded: str, brand_lookup: List[tuple]) -> dict:
    if result.get("category_code") != "HE":
        return result
    if "LED" not in padded:
        return result
    first_word = padded.strip().split(" ")[0] if padded.strip() else ""
    if first_word not in AGEING_LED_COMPUTER_BRANDS:
        return result
    fixed = dict(result)
    fixed["category_code"] = "IT"
    fixed["category_name"] = "Computer"
    fixed["brand_name"] = _resolve_ageing_brand_name(padded, padded.strip(), brand_lookup)
    fixed["classification_source"] = "Keyword"
    return fixed


# A "Remote"/"Remote Controller" item defaults to Home Entertainment (TV
# remote) via the " REMOTE" keyword rule above. Daikin's AC remote
# controllers, however, are always modeled with a "BRC..." code (e.g.
# "Daikin Remote Controller BRC91A152") - never a TV remote. Per Admin
# request this specific pattern is re-routed to Home Appliances instead,
# same as the AC unit itself would be. " BRC" is padded so it only matches
# the standalone model-code token, not an accidental substring elsewhere.
def _enforce_ac_remote_is_home_appliance(result: dict, padded: str, brand_lookup: List[tuple]) -> dict:
    if result.get("category_code") != "HE":
        return result
    if "REMOTE" not in padded or " BRC" not in padded:
        return result
    fixed = dict(result)
    fixed["category_code"] = "HA"
    fixed["category_name"] = "Home Appliances"
    fixed["brand_name"] = _resolve_ageing_brand_name(padded, padded.strip(), brand_lookup)
    fixed["classification_source"] = "Keyword"
    return fixed


def _enforce_ac_is_not_a_brand(result: dict, padded: str, brand_lookup: List[tuple]) -> dict:
    """"AC" is a product-type marker (air conditioner spare part/install
    item, e.g. "AC Remote Control", "AC Filter Net", "AC Gas Charging"),
    never a real registered Brand. When no keyword/brand rule already
    caught the item, the brand-guessing fallback in
    _resolve_ageing_brand_name naively takes the item's first word as its
    brand - for one of these AC-prefixed items that produces a bogus "AC"
    brand grouping. Whenever that happens, the item is redirected into
    Accessories/Other (where an AC spare part/consumable belongs anyway)
    and re-resolved through the accessory-only brand lookup, which never
    invents a brand from the first word. Genuine AC units (e.g. "Voltas AC
    1.5 Ton", "LG AC ID 1.5T") are unaffected - their first word is the
    real manufacturer brand, not literally "AC", so this never fires for
    them."""
    if (result.get("brand_name") or "").strip().upper() != "AC":
        return result
    fixed = dict(result)
    fixed["category_code"] = "OTHER"
    fixed["category_name"] = "Accessories"
    fixed["brand_name"] = _resolve_known_brand_only(padded, brand_lookup)
    fixed["classification_source"] = "Accessory"
    return fixed


# Words that are a generic product-type descriptor rather than a real
# registered Brand, but can still get guessed as the brand (via the
# item's first word) when no known Brand record matches - e.g. "Angle
# Grinder" has no "Angle" brand on file, so the fallback in
# _resolve_ageing_brand_name would otherwise invent an "Angle" brand
# grouping. Unlike the AC case above, these items ARE genuinely correctly
# categorized already (e.g. "GRINDER" already matches the HA keyword
# rule) - only the bogus brand name needs fixing, so this guard leaves
# category_code/category_name untouched and only re-resolves the brand
# through the known-brand-only lookup (falls back to "Others").
AGEING_NON_BRAND_FIRST_WORDS = {
    "ANGLE",
    # "Laptop" is a product-type word (subcategory), not a real registered
    # Brand - without this, "Laptop Speaker" (and any other "Laptop ..."
    # item with no real brand on file) would get grouped under a bogus
    # "Laptop" brand card via the first-word fallback instead of "Others".
    "LAPTOP",
}


def _enforce_generic_words_are_not_brands(result: dict, padded: str, brand_lookup: List[tuple]) -> dict:
    if (result.get("brand_name") or "").strip().upper() not in AGEING_NON_BRAND_FIRST_WORDS:
        return result
    fixed = dict(result)
    fixed["brand_name"] = _resolve_known_brand_only(padded, brand_lookup)
    return fixed


# Brand casings/spellings that should always display as one canonical
# name in the Ageing Stock report, keyed by the UPPERCASE normalized form.
# A brand that exists twice in the Brand master under different casing
# (e.g. "Oneplus" and "OnePlus", created at different times) still shows
# up as a single, consistent brand grouping here regardless of which
# underlying Brand record a given item happened to match. Add an entry
# whenever a duplicate/variant brand casing shows up in the report.
AGEING_BRAND_CANONICAL_NAMES = {
    "ONEPLUS": "OnePlus",
    "IQOO": "iQOO",
}


def _canonicalize_ageing_brand_name(brand_name: Optional[str]) -> Optional[str]:
    if not brand_name:
        return brand_name
    return AGEING_BRAND_CANONICAL_NAMES.get(brand_name.strip().upper(), brand_name)

# Brands treated as Mobile Phone by default when no product-type keyword
# above matched (covers plain phone-model rows like "Samsung A17 5G...").
# NOTE: Lenovo is intentionally NOT here - Lenovo's real business here is
# laptops/tablets, and a plain "Lenovo <model-number>" row (no LAPTOP/TAB
# keyword in the name) should default to Computer/IT (see
# AGEING_COMPUTER_BRAND_HINTS below), not Mobile. Lenovo Tab items still
# correctly land in Mobile via the " TAB " keyword rule above, which is
# checked before this fallback.
AGEING_MOBILE_BRAND_HINTS = {
    "IQOO", "MOTOROLA", "NOTHING", "ONEPLUS", "OPPO",
    "REALME", "SAMSUNG", "VIVO", "XIAOMI", "REDMI", "CMF",
}
# Brands treated as Computer/IT by default when no product-type keyword
# above matched. Lenovo included here (rather than in the Mobile hints)
# so a bare "Lenovo <model-number>" row - with no LAPTOP/TAB keyword and
# no explicit ageing_category_master.json entry - defaults to Computer,
# matching Lenovo's actual business here.
AGEING_COMPUTER_BRAND_HINTS = {"ASUS", "DELL", "HP", "EPSON", "ACER", "LENOVO"}
# Brands treated as Digital Camera by default when no product-type keyword
# above matched (covers plain camera-model rows like "Nikon Coolpix P900"
# that don't literally say "Camera"/"Mirrorless"/"DSLR" in the name). Per
# Admin request, every Nikon item defaults here.
AGEING_CAMERA_BRAND_HINTS = {"NIKON"}


def build_ageing_brand_lookup(db: Session) -> List[tuple]:
    """(normalized brand name, brand.name, category_code, category_name)
    for every existing Brand, longest name first so 'Samsung Galaxy' style
    multi-word brands (if any) match before a shorter brand substring
    would. Used as a secondary signal after the keyword rules."""
    rows = (
        db.query(models.Brand, models.Category)
        .outerjoin(models.SubCategory, models.Brand.subcategory_id == models.SubCategory.id)
        .outerjoin(models.Category, models.SubCategory.category_id == models.Category.id)
        .all()
    )
    # A Brand record whose name duplicates a Category's own display name
    # (e.g. a brand literally named "Mobiles / Handset", matching the MH
    # category's full name) is a data-entry mistake, not a real brand -
    # left in, it shows up as a bogus "brand" grouping in the ageing
    # report. Excluded here so such rows fall back to "Others" instead.
    category_names = {
        normalize_ageing_item_key(c.name) for c in db.query(models.Category).all()
    }
    lookup = []
    for brand, category in rows:
        normalized = normalize_ageing_item_key(brand.name)
        if not normalized or normalized in category_names:
            continue
        lookup.append((normalized, brand.name, category.code if category else None, category.name if category else None))
    lookup.sort(key=lambda entry: len(entry[0]), reverse=True)
    return lookup


def _resolve_ageing_brand_name(padded: str, normalized: str, brand_lookup: List[tuple]) -> Optional[str]:
    """Shared brand-name resolution: prefer a match against the existing
    Brand master, else fall back to the item's first word. Only used for
    items landing in a real product category (HA/HE/Mobile/IT) - those
    names reliably start with the brand (e.g. "Samsung LED...", "HP
    Laptop..."), so the first-word guess is safe there."""
    for brand_key, original_name, _cat_code, _cat_name in brand_lookup:
        if brand_key and padded.startswith(f" {brand_key} "):
            return original_name
    return normalized.split(" ")[0].title() if normalized else None


def _resolve_known_brand_only(padded: str, brand_lookup: List[tuple]) -> Optional[str]:
    """Brand resolution for Accessories/Unclassified items: match against
    the existing Brand master ONLY - never guess the first word as a
    brand. Accessory names (e.g. "Flare Natt 1/2", "Floor Stand", "Foam
    Sheet") don't follow a brand-first naming convention, so guessing
    would invent fake brands like "Flare"/"Floor"/"Foam" instead of
    grouping them sensibly. Falls back to "Others" when no real brand
    matches - NOT "General": "General" (as in "O General" air
    conditioners) is itself a real registered Brand, so using that word
    as the no-match fallback made unrelated items look like they belonged
    to the General brand."""
    for brand_key, original_name, _cat_code, _cat_name in brand_lookup:
        if brand_key and padded.startswith(f" {brand_key} "):
            return original_name
    return "Others"


def build_ageing_category_master_lookup(db: Session) -> dict:
    """normalized item key -> {category_code, category_name, brand_name}.
    Loaded once from the bundled static/ageing_category_master.json file
    (generated from the company's HA/HE/Mobile/Computer item lists) - see
    _AGEING_CATEGORY_MASTER_LOOKUP below. classify_ageing_item() checks
    this first; anything not on the list falls through to the keyword/
    brand rules exactly as before. The db argument is unused - kept so
    call sites don't need to change if this is ever swapped back to a
    DB-backed lookup."""
    return _AGEING_CATEGORY_MASTER_LOOKUP


def _load_ageing_category_master() -> dict:
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "ageing_category_master.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            rows = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        rows = []
    lookup = {
        row["item_key"]: {
            "category_code": row["category_code"],
            "category_name": AGEING_CATEGORY_MASTER_SHEET_NAMES.get(row["category_code"], row["category_code"]),
            "brand_name": None,  # resolved dynamically at classify time against the current Brand master
        }
        for row in rows
        if row.get("item_key")
    }
    # Manual overrides added directly here in code (Admin request), on top
    # of whatever the bundled static/ageing_category_master.json says for
    # these specific items - so they land in the right Ageing Stock
    # Analysis category even if that JSON file is missing, stale, or
    # doesn't list them yet. Keyed by the same normalized item key used
    # everywhere else in Ageing Stock Analysis (normalize_ageing_item_key),
    # and applied last so an override always wins over the JSON entry.
    AGEING_CATEGORY_MASTER_MANUAL_OVERRIDES = {
        "SAMSUNG COVER SILICONE S25 EDGE WHITE": "OTHER",   # Samsung (Cover) Silicone S25 Edge White -> Accessories
        "SAMSUNG FRAME PVC VG SCFA43WTBXL": "OTHER",        # Samsung Frame PVC VG SCFA43WTBXL -> Accessories
        "SAMSUNG GEAR R 3500 BLACK": "OTHER",                # Samsung Gear R 3500 Black -> Accessories
        "SAMSUNG NOTE 10 COVER EF NN970PBEGIN": "OTHER",     # Samsung Note 10 Cover EF-NN970PBEGIN -> Accessories
        "SAMSUNG DISPLAY 32 QM32C": "HE",                    # Samsung Display 32 QM32C -> Home Entertainment
        "SAMSUNG COVER FLIP 4 GP TOF721AM7RI": "OTHER",      # Samsung Cover Flip 4 GP-TOF721AM7RI -> Accessories
        "SAMSUNG FRAME PVC VG SCFA50WTBXL": "OTHER",         # Samsung Frame PVC VG SCFA50WTBXL -> Accessories
        "SAMSUNG GEAR R 7000": "OTHER",                      # Samsung Gear R 7000 -> Accessories
        "SAMSUNG DISPLAY 43 QB43C": "HE",                    # Samsung Display 43 QB43C -> Home Entertainment
        "NIKON MIRRORLESS Z5 II 24 70": "DC",                # Nikon Mirrorless Z5 II 24-70 -> Digital Camera
        "ELICA HOOD KITTY SLIM EDS HE LTW 90": "HA",         # Elica Hood KITTY SLIM EDS HE LTW 90 -> Home Appliances
        "KUHL FAN LUXUS C14 1200 WHITE": "HA",                # KUHL Fan Luxus C14 1200 White -> Home Appliances
        "MORPHY RICHARDS OFR 13F 290012REVIEW": "HA",        # Morphy Richards OFR 13F 290012REVIEW -> Home Appliances
        "SUNFLAME COOKER HOOD MAGNUM 60": "HA",              # Sunflame Cooker Hood Magnum 60 -> Home Appliances
        "JBL SOUND BAR SB595": "HE",                          # JBL Sound Bar SB595 -> Home Entertainment
        "LAPTOP SPEAKER": "IT",                               # Laptop Speaker -> Computer (would otherwise match the "SPEAKER" HE keyword)
        "HAVELLS LED PRIDE PLUS 10W": "OTHER",                # Havells LED Pride Plus 10W -> Accessories (was landing in HE)
    }
    for item_key, category_code in AGEING_CATEGORY_MASTER_MANUAL_OVERRIDES.items():
        lookup[item_key] = {
            "category_code": category_code,
            "category_name": AGEING_CATEGORY_MASTER_SHEET_NAMES.get(category_code, category_code),
            "brand_name": None,
        }
    return lookup


# Loaded once at startup from static/ageing_category_master.json - the
# company's own HA/HE/Mobile/Computer item lists - plus the manual
# overrides above. Every ageing-stock item is matched against this first
# (exact, normalized name); anything not on it falls through to the
# keyword/brand-hint rules below, and finally to Other (Accessories) if
# nothing matches at all.
_AGEING_CATEGORY_MASTER_LOOKUP = _load_ageing_category_master()


def classify_ageing_item(item_details: str, brand_lookup: List[tuple], category_master_lookup: Optional[dict] = None) -> dict:
    normalized = normalize_ageing_item_key(item_details)
    padded = f" {normalized} "
    result = _classify_ageing_item_raw(item_details, normalized, padded, brand_lookup, category_master_lookup)
    result = _enforce_mobile_category_is_phones_only(result, padded, brand_lookup)
    result = _enforce_led_computer_brands_are_computer(result, padded, brand_lookup)
    result = _enforce_ac_remote_is_home_appliance(result, padded, brand_lookup)
    result = _enforce_ac_is_not_a_brand(result, padded, brand_lookup)
    result = _enforce_generic_words_are_not_brands(result, padded, brand_lookup)
    result = dict(result)
    result["brand_name"] = _canonicalize_ageing_brand_name(result.get("brand_name"))
    return result


def _classify_ageing_item_raw(
    item_details: str,
    normalized: str,
    padded: str,
    brand_lookup: List[tuple],
    category_master_lookup: Optional[dict],
) -> dict:
    # Exact match against the uploaded Category Master list (HA/HE/Mobile/
    # Computer sheets) - this is the authoritative source when present,
    # since it's the company's own definition of what belongs to each
    # category. Checked before every keyword/brand heuristic below, so a
    # listed item always lands in its master-assigned category even if its
    # name would otherwise match a different keyword rule.
    if category_master_lookup:
        master_match = category_master_lookup.get(normalized)
        if master_match:
            if master_match["brand_name"]:
                resolved_brand = master_match["brand_name"]
            elif master_match["category_code"] == "OTHER":
                # Accessories/gift items in the master list never guess a
                # brand from the item's first word (same reasoning as the
                # Accessory keyword branch below) - a master-listed item
                # like "Laptop Keyguard" would otherwise get grouped under
                # a bogus "Laptop" brand instead of a real registered one
                # or "Others".
                resolved_brand = _resolve_known_brand_only(padded, brand_lookup)
            else:
                resolved_brand = _resolve_ageing_brand_name(padded, normalized, brand_lookup)
            return {
                "category_code": master_match["category_code"],
                "category_name": master_match["category_name"],
                "brand_name": resolved_brand,
                "classification_source": "Category Master",
            }

    # Accessories/consumables always land in "Other", regardless of brand -
    # checked before every other rule so e.g. a Samsung charger doesn't get
    # pulled into the Mobile category with the rest of Samsung's phones.
    if any(keyword in padded for keyword in AGEING_ACCESSORY_KEYWORDS) or _is_ram_memory_module(padded):
        # "iPhone" is a product line, not a brand - same special case as
        # the MH keyword rule below, so an iPhone accessory (e.g. "iPhone
        # 15 Pro Max (Silicon Case) MT1Y3ZM/A") still groups under Apple
        # instead of falling through to "Others" (its item text starts
        # with "Iphone", not "Apple", so the brand-master prefix match in
        # _resolve_known_brand_only would otherwise miss it).
        if "IPHONE" in padded:
            brand_name = "Apple"
        else:
            brand_name = _resolve_known_brand_only(padded, brand_lookup)
        return {
            "category_code": "OTHER", "category_name": "Accessories",
            "brand_name": brand_name,
            "classification_source": "Accessory",
        }

    for code, name, keywords in AGEING_CATEGORY_KEYWORD_RULES:
        if any(keyword in padded for keyword in keywords):
            # "iPhone"/"iPad" are product lines, not a brand - they're
            # sold under Apple, so group them there instead of guessing
            # the brand name is literally "Iphone"/"Ipad" from the item
            # text's first word.
            if code == "MH" and ("IPHONE" in padded or "IPAD" in padded):
                brand_name = "Apple"
            else:
                brand_name = _resolve_ageing_brand_name(padded, normalized, brand_lookup)
            return {
                "category_code": code, "category_name": name,
                "brand_name": brand_name, "classification_source": "Keyword",
            }

    # No product-type keyword matched - fall back to brand-based defaults.
    first_word = normalized.split(" ")[0] if normalized else ""
    if first_word in AGEING_MOBILE_BRAND_HINTS:
        return {
            "category_code": "MH", "category_name": "Mobile",
            "brand_name": first_word.title(), "classification_source": "Keyword",
        }
    if first_word in AGEING_COMPUTER_BRAND_HINTS:
        return {
            "category_code": "IT", "category_name": "Computer",
            "brand_name": first_word.title() if first_word != "HP" else "HP",
            "classification_source": "Keyword",
        }
    if first_word in AGEING_CAMERA_BRAND_HINTS:
        return {
            "category_code": "DC", "category_name": "Digital Camera",
            "brand_name": first_word.title(), "classification_source": "Keyword",
        }

    # Next - match the existing Brand master by longest-prefix, if that
    # brand has a category assigned to it. Only trust the brand's real ERP
    # category when it's one of the four the Ageing Stock report actually
    # tracks (HA/HE/MH/IT) - a brand filed under the main app's own
    # "Others" (OTH) or "Accessories" (ASC) category is intentionally NOT
    # passed through here, since that would create a second, differently-
    # spelled Accessories bucket alongside the report's own OTHER/
    # Accessories catch-all below and split what should be one combined
    # count into two cards.
    for brand_key, original_name, cat_code, cat_name in brand_lookup:
        if brand_key and padded.startswith(f" {brand_key} ") and cat_code in AGEING_CATEGORY_MASTER_SHEET_NAMES and cat_code != "OTHER":
            # Display name always comes from the Ageing Stock report's own
            # naming (AGEING_CATEGORY_MASTER_SHEET_NAMES), not the real ERP
            # Category.name text, so e.g. "IT" always shows as "Computer"
            # here too, even though the real Category record underneath is
            # still named "Information Technology" for the rest of the app.
            return {
                "category_code": cat_code, "category_name": AGEING_CATEGORY_MASTER_SHEET_NAMES[cat_code],
                "brand_name": original_name, "classification_source": "Brand Master",
            }

    # Last resort - nothing recognized this item at all. It still needs a
    # home, so it goes into "Accessories" (flagged Unclassified/REVIEW)
    # rather than a separate "Uncategorized" bucket.
    return {
        "category_code": "OTHER", "category_name": "Accessories",
        "brand_name": _resolve_known_brand_only(padded, brand_lookup),
        "classification_source": "Unclassified",
    }


def parse_ageing_stock_workbook(content: bytes, db: Session) -> tuple:
    """Returns (item_rows, location_sheets_found, unclassified_count) or
    raises HTTPException on a workbook that doesn't look right."""
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {exc}") from exc

    all_data_rows = None
    location_row_maps: dict = {}  # location code -> {normalized item key: qty}
    location_bucket_maps: dict = {}  # location code -> {normalized item key: {age_bucket: qty}}
    location_sheets_found = []
    excluded_item_keys: set = set()  # item keys found on a PWH/Vault sheet

    for worksheet in workbook.worksheets:
        title_normalized = re.sub(r"[^a-z0-9]", "", str(worksheet.title or "").lower())

        if title_normalized in AGEING_ALL_DATA_SHEET_ALIASES:
            all_data_rows = read_ageing_sheet_rows(worksheet)
            continue

        detected_code = detect_ageing_location_code(worksheet)
        excluded_code = detected_code if detected_code in AGEING_EXCLUDED_LOCATION_ALIASES else None
        if excluded_code:
            for row in read_ageing_sheet_rows(worksheet):
                key = normalize_ageing_item_key(row.get("item_details"))
                if key:
                    excluded_item_keys.add(key)
            continue

        matched_code = detected_code if detected_code in AGEING_LOCATION_DEFINITIONS else None
        if not matched_code:
            continue

        rows = read_ageing_sheet_rows(worksheet)
        if not rows:
            continue
        location_sheets_found.append(matched_code)
        qty_map = {}
        # Per-item, per-age-bucket breakdown WITHIN this location, so a
        # duration filter (e.g. ">= 366 Days") can show exactly this
        # branch's quantity for that duration, not its all-time total -
        # see location_qty_for_durations() below and the
        # location_age_buckets_json comment on models.AgeingStockItem.
        bucket_map: dict = {}
        for row in rows:
            key = normalize_ageing_item_key(row.get("item_details"))
            if not key:
                continue
            buckets = {
                bucket: parse_float_value(row.get(bucket), fallback=0.0)
                for bucket in AGEING_AGE_FIELDS
            }
            qty = parse_float_value(row.get("closing_qty"), fallback=None)
            if qty is None:
                # Some location sheets may only carry ageing buckets, no
                # Closing Qty column - fall back to their sum.
                qty = sum(buckets.values())
            qty_map[key] = qty_map.get(key, 0.0) + qty
            if key not in bucket_map:
                bucket_map[key] = dict(buckets)
            else:
                for bucket, value in buckets.items():
                    bucket_map[key][bucket] += value
        location_row_maps[matched_code] = qty_map
        location_bucket_maps[matched_code] = bucket_map

    if all_data_rows is None:
        raise HTTPException(
            status_code=400,
            detail="No 'All Data' sheet found. Rename the master sheet (with every item) to 'All Data' and re-upload.",
        )
    if not all_data_rows:
        raise HTTPException(status_code=400, detail="The 'All Data' sheet has no readable item rows.")

    brand_lookup = build_ageing_brand_lookup(db)
    category_master_lookup = build_ageing_category_master_lookup(db)
    location_columns = {
        "ALM": "qty_alm", "HZT": "qty_hzt", "ASH": "qty_ash", "GNG": "qty_gng",
        "VKN": "qty_vkn", "MWH": "qty_mwh",
    }

    item_rows = []
    unclassified_count = 0
    for row in all_data_rows:
        item_details = str(row.get("item_details") or "").strip()
        # Per Admin request: Ageing Stock Analysis only covers Nos.-unit
        # items. A row measured in any other unit (Mtr, Kg, Ltr, Set, Box,
        # etc.), or with no unit at all, is skipped entirely - it never
        # reaches classification and never appears anywhere in the report
        # or its totals.
        if not is_ageing_nos_unit(row.get("unit")):
            continue
        item_key = normalize_ageing_item_key(item_details)
        # Drop items that belong to PWH or Vault entirely - per Admin
        # request, this report does not consider that stock at all, even
        # though it's part of the "All Data" master sheet's totals.
        if item_key in excluded_item_keys:
            continue

        classification = classify_ageing_item(item_details, brand_lookup, category_master_lookup)
        if classification["classification_source"] == "Unclassified":
            unclassified_count += 1

        built = {
            "item_details": item_details,
            "unit": str(row.get("unit") or "").strip() or None,
            "model_no": str(row.get("model_no") or "").strip() or None,
            "closing_qty": parse_float_value(row.get("closing_qty"), fallback=0.0),
            "age_0_60": parse_float_value(row.get("age_0_60"), fallback=0.0),
            "age_61_90": parse_float_value(row.get("age_61_90"), fallback=0.0),
            "age_91_150": parse_float_value(row.get("age_91_150"), fallback=0.0),
            "age_151_180": parse_float_value(row.get("age_151_180"), fallback=0.0),
            "age_181_365": parse_float_value(row.get("age_181_365"), fallback=0.0),
            "age_366_plus": parse_float_value(row.get("age_366_plus"), fallback=0.0),
            "category_code": classification["category_code"],
            "category_name": classification["category_name"],
            "brand_name": classification["brand_name"],
            "classification_source": classification["classification_source"],
            "qty_alm": 0.0, "qty_hzt": 0.0, "qty_ash": 0.0, "qty_gng": 0.0,
            "qty_vkn": 0.0, "qty_mwh": 0.0,
        }

        present_at = []
        location_buckets_out = {}
        for code, column in location_columns.items():
            qty_map = location_row_maps.get(code)
            if not qty_map:
                continue
            qty_here = qty_map.get(item_key)
            if qty_here:
                built[column] = qty_here
                present_at.append(code)
                bucket_here = (location_bucket_maps.get(code) or {}).get(item_key)
                if bucket_here:
                    location_buckets_out[column] = bucket_here
        built["locations_present"] = ", ".join(present_at) if present_at else None
        # {"qty_alm": {"age_0_60": 1.0, ...}, "qty_hzt": {...}, ...} - lets
        # the report/export show each branch's quantity for exactly the
        # selected duration bucket(s) instead of that branch's all-time
        # total. See location_qty_for_durations() below.
        built["location_age_buckets_json"] = json.dumps(location_buckets_out) if location_buckets_out else None

        item_rows.append(built)

    return item_rows, location_sheets_found, unclassified_count


@app.post("/api/ageing-stock/upload")
def upload_ageing_stock_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    filename = file.filename or "uploaded_file"
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Upload an Excel (.xlsx/.xls) workbook with an 'All Data' sheet plus one sheet per location.")

    raw_content = file.file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    item_rows, location_sheets_found, unclassified_count = parse_ageing_stock_workbook(raw_content, db)

    # Replace the previous dataset wholesale - this report always reflects
    # only the most recently uploaded workbook.
    db.query(models.AgeingStockItem).delete()
    db.query(models.AgeingStockUpload).delete()
    db.commit()

    upload_record = models.AgeingStockUpload(
        source_file=filename,
        uploaded_by=current_user.id,
        uploaded_by_username=current_user.username,
        item_count=len(item_rows),
        location_sheets_found=",".join(location_sheets_found),
        unclassified_count=unclassified_count,
    )
    db.add(upload_record)
    db.commit()
    db.refresh(upload_record)

    for r in item_rows:
        r["upload_id"] = upload_record.id
        r["source_file"] = filename
    db.bulk_insert_mappings(models.AgeingStockItem, item_rows)
    db.commit()

    missing_locations = [code for code in AGEING_LOCATION_DEFINITIONS if code not in location_sheets_found]

    return {
        "message": "File processed successfully",
        "file_name": filename,
        "items_loaded": len(item_rows),
        "location_sheets_found": location_sheets_found,
        "location_sheets_missing": missing_locations,
        "unclassified_count": unclassified_count,
    }


@app.get("/api/ageing-stock/meta")
def ageing_stock_meta(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    upload_record = db.query(models.AgeingStockUpload).order_by(models.AgeingStockUpload.id.desc()).first()
    if not upload_record:
        return {"has_data": False, "last_upload": None, "can_upload": auth.has_admin_access(current_user)}

    return {
        "has_data": True,
        "last_upload": {
            "file_name": upload_record.source_file,
            "uploaded_by": upload_record.uploaded_by_username,
            "uploaded_at": upload_record.created_date,
            "item_count": upload_record.item_count,
            "location_sheets_found": (upload_record.location_sheets_found or "").split(",") if upload_record.location_sheets_found else [],
            "unclassified_count": upload_record.unclassified_count,
        },
        "can_upload": auth.has_admin_access(current_user),
    }


@app.get("/api/ageing-stock/brands")
def ageing_stock_brands(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Every distinct brand name currently in the Ageing Stock Analysis
    data, alphabetically sorted, with how many items each brand has - used
    to power the searchable Brand dropdown on the report page. A brand can
    span several categories (e.g. a brand sold under both HA and IT), so
    this list is category-independent; selecting a brand in the report
    filter shows that brand's items across every category it appears in."""
    rows = (
        db.query(models.AgeingStockItem.brand_name, func.count(models.AgeingStockItem.id))
        .filter(models.AgeingStockItem.brand_name.isnot(None))
        .group_by(models.AgeingStockItem.brand_name)
        .order_by(models.AgeingStockItem.brand_name)
        .all()
    )
    return {"brands": [{"name": name, "item_count": count} for name, count in rows if name]}


AGEING_AGE_FIELDS = ["age_0_60", "age_61_90", "age_91_150", "age_151_180", "age_181_365", "age_366_plus"]
AGEING_AGE_LABELS = {
    "age_0_60": "0-60 Days",
    "age_61_90": "61-90 Days",
    "age_91_150": "91-150 Days",
    "age_151_180": "151-180 Days",
    "age_181_365": "181-365 Days",
    "age_366_plus": "\u2265 366 Days",
}
AGEING_LOCATION_FIELDS = ["qty_alm", "qty_hzt", "qty_ash", "qty_gng", "qty_vkn", "qty_mwh"]

# Product-type classification (LED, Cooler, AC, etc.) used by the Item
# Category filter on the Ageing Stock Analysis page. This list MUST stay in
# sync with the ITEM_CATEGORIES list in ageing_stock.html - it exists here
# too so the Excel export can filter to the same item type the user picked
# on screen instead of only being able to filter by the DB's real
# category/brand/search fields. First matching entry wins, same as the JS.
#
# HOOD and RO must come before AC: kitchen chimneys/hoods often carry an
# "AC" (Auto Clean) spec in their name (e.g. "Sunflame Hood Lara 60 BK AC
# GC", "Faber Hood Zenith FL SC AC BK 60"), and RO water purifiers often
# carry an "AC" (Alkaline Copper) spec (e.g. "Eureka Aquaguard Premier
# RO+UV+MTDS+AC") - neither is an air conditioner, so without HOOD/RO
# checked first, that unrelated "AC" would wrongly land them in AC (All
# Types) instead of their own real category.
AGEING_ITEM_CATEGORIES = [
    ("LED", [r"\bLED\b", r"\bQLED\b", r"\bOLED\b", r"\bTELEVISION\b"]),
    ("SPEAKER", [r"\bSPEAKER\b"]),
    ("REF", [r"\bREF\b", r"\bREFRIGERATOR\b", r"\bFRIDGE\b"]),
    ("WM", [r"\bWM\b", r"\bWASHING MACHINE\b"]),
    ("GEYSER", [r"\bGEYSER\b", r"\bWATER HEATER\b"]),
    ("COOLER", [r"\bCOOLER\b"]),
    ("IRON", [r"\bIRON\b"]),
    ("MIXER", [r"\bMIXER\b"]),
    ("BLENDER", [r"\bBLENDER\b"]),
    ("AIR_PURIFIER", [r"\bAIR\s*PURIFIER\b"]),
    ("HAIR_DRYER", [r"\bHAIR\s*DRYER\b"]),
    ("VACUUM", [r"\bVACUUM\b"]),
    ("MWO", [r"\bMWO\b", r"\bMICROWAVE\b"]),
    ("HOOD", [r"\bHOOD\b", r"\bCHIMNEY\b"]),
    ("RO", [r"\bRO\b", r"\bAQUAGUARD\b", r"\bWATER\s*PURIFIER\b"]),
    ("AC", [r"\bAC\b", r"\bSAC\b", r"\bAIR\s*CONDITION"]),
    ("FAN", [r"\bFAN\b"]),
    ("JUICER", [r"\bJUICER\b"]),
    ("SOUNDBAR", [r"\bSOUND\s*BAR\b"]),
    ("IPAD", [r"\bIPAD\b"]),
    ("TAB", [r"\bTAB\b", r"\bTABLET\b"]),
    ("CAMERA", [r"\bCAMERA\b", r"\bDSLR\b", r"\bCAMCORDER\b"]),
]
_AGEING_ITEM_CATEGORY_PATTERNS = [
    (code, [re.compile(p) for p in patterns]) for code, patterns in AGEING_ITEM_CATEGORIES
]


def classify_ageing_item_category(text_value: Optional[str]) -> Optional[str]:
    t = (text_value or "").upper()
    for code, patterns in _AGEING_ITEM_CATEGORY_PATTERNS:
        if any(p.search(t) for p in patterns):
            return code
    return None


def compute_ageing_stock_report(
    db: Session,
    category: Optional[str],
    brand: Optional[str],
    search: Optional[str],
    item_category: Optional[str] = None,
    locations: Optional[str] = None,
) -> dict:
    """Shared query + aggregation logic behind both the on-screen report
    (/api/ageing-stock/report) and the Excel export (/api/ageing-stock/export),
    so the two can never drift out of sync with each other."""
    query = db.query(models.AgeingStockItem)
    if category and category.upper() != "ALL":
        # Supports either a single code ("HA") or a comma-separated list
        # ("HA,HE,MH") so the on-screen checkbox multi-select and the
        # exports it drives can show/download any combination of
        # categories, not just one at a time.
        codes = [c.strip().upper() for c in category.split(",") if c.strip()]
        conditions = []
        if "UNCATEGORIZED" in codes:
            conditions.append(models.AgeingStockItem.category_code.is_(None))
            codes = [c for c in codes if c != "UNCATEGORIZED"]
        if codes:
            conditions.append(models.AgeingStockItem.category_code.in_(codes))
        if conditions:
            query = query.filter(or_(*conditions))
    if brand:
        query = query.filter(models.AgeingStockItem.brand_name == brand)
    if search:
        # Same search box matches either Item Details or Model No., so
        # typing a model number (e.g. from a price tag or invoice) finds
        # the item even when that model number isn't part of the Item
        # Details text itself.
        query = query.filter(
            or_(
                models.AgeingStockItem.item_details.ilike(f"%{search}%"),
                models.AgeingStockItem.model_no.ilike(f"%{search}%"),
            )
        )

    items = query.order_by(models.AgeingStockItem.category_name, models.AgeingStockItem.brand_name, models.AgeingStockItem.item_details).all()

    if item_category:
        item_category = item_category.upper()
        items = [it for it in items if classify_ageing_item_category(it.item_details) == item_category]

    if locations:
        location_codes = [code.strip().lower() for code in locations.split(",") if code.strip()]
        location_fields = []
        for code in location_codes:
            field = f"qty_{code}"
            if field not in AGEING_LOCATION_FIELDS:
                raise HTTPException(status_code=400, detail=f"Unknown branch '{code}'.")
            location_fields.append(field)
        if location_fields:
            items = [it for it in items if any((getattr(it, f) or 0) > 0 for f in location_fields)]

    age_fields = AGEING_AGE_FIELDS
    location_fields = AGEING_LOCATION_FIELDS

    def empty_totals():
        return {f: 0.0 for f in age_fields + location_fields + ["closing_qty"]}

    def add_into(totals, item):
        for f in age_fields + location_fields:
            totals[f] += getattr(item, f) or 0.0
        totals["closing_qty"] += item.closing_qty or 0.0

    categories_out = {}
    grand_total = empty_totals()
    grand_total_items = 0

    for item in items:
        cat_key = item.category_name or "Uncategorized"
        cat = categories_out.setdefault(cat_key, {
            "category_name": cat_key,
            "category_code": item.category_code,
            "brands": {},
            "totals": empty_totals(),
            "item_count": 0,
        })
        brand_key = item.brand_name or "Unbranded"
        brand_bucket = cat["brands"].setdefault(brand_key, {
            "brand_name": brand_key,
            "items": [],
            "totals": empty_totals(),
        })

        brand_bucket["items"].append({
            "id": item.id,
            "item_details": item.item_details,
            "model_no": item.model_no,
            "unit": item.unit,
            "closing_qty": item.closing_qty,
            "age_0_60": item.age_0_60,
            "age_61_90": item.age_61_90,
            "age_91_150": item.age_91_150,
            "age_151_180": item.age_151_180,
            "age_181_365": item.age_181_365,
            "age_366_plus": item.age_366_plus,
            "qty_alm": item.qty_alm, "qty_hzt": item.qty_hzt, "qty_ash": item.qty_ash,
            "qty_gng": item.qty_gng, "qty_vkn": item.qty_vkn, "qty_mwh": item.qty_mwh,
            "locations_present": item.locations_present,
            "location_age_buckets_json": item.location_age_buckets_json,
            "classification_source": item.classification_source,
        })
        add_into(brand_bucket["totals"], item)
        add_into(cat["totals"], item)
        add_into(grand_total, item)
        cat["item_count"] += 1
        grand_total_items += 1

    def category_sort_key(cat_key):
        code = (categories_out[cat_key].get("category_code") or "").upper()
        try:
            return (AGEING_CATEGORY_ORDER.index(code), cat_key)
        except ValueError:
            # Unrecognized/legacy code (e.g. data from before this
            # classification scheme) - show it after the known categories,
            # alphabetically among themselves, but still before nothing.
            return (len(AGEING_CATEGORY_ORDER), cat_key)

    categories_list = []
    for cat_key in sorted(categories_out.keys(), key=category_sort_key):
        cat = categories_out[cat_key]
        cat["brands"] = [cat["brands"][b] for b in sorted(cat["brands"].keys())]
        categories_list.append(cat)

    upload_record = db.query(models.AgeingStockUpload).order_by(models.AgeingStockUpload.id.desc()).first()

    return {
        "has_data": upload_record is not None,
        "categories": categories_list,
        "grand_total": grand_total,
        "grand_total_items": grand_total_items,
        "location_columns": [
            {"code": code, "name": definition["name"]}
            for code, definition in AGEING_LOCATION_DEFINITIONS.items()
        ],
        "filters": {"category": category or "ALL", "brand": brand, "search": search, "item_category": item_category, "locations": locations},
    }


@app.get("/api/ageing-stock/report")
def ageing_stock_report(
    category: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    item_category: Optional[str] = Query(None),
    locations: Optional[str] = Query(None, description="Comma-separated branch codes (ALM,HZT,ASH,GNG,VKN,MWH) - keeps items with stock at ANY of them. Omit for all branches."),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return compute_ageing_stock_report(db, category, brand, search, item_category, locations)


def location_qty_for_durations(item: dict, location_field: str, active_age_fields: List[str]) -> float:
    """This item's quantity at one branch (qty_alm, qty_hzt, etc.), scoped
    to only the selected duration bucket(s) - fixes branch columns showing
    a branch's ALL-TIME total next to a Closing Qty/duration column that's
    already scoped to the selected duration(s), which could make the
    branch columns sum to more than Closing Qty (e.g. an item aged 91-150
    days at one branch was still being counted when the report was
    filtered to only ">= 366 Days").

    Uses the per-branch age-bucket breakdown captured at upload time
    (location_age_buckets_json). Falls back to the branch's raw all-time
    total when every duration bucket is selected (no filtering needed - the
    two are equivalent) or when the item predates this fix and has no
    bucket breakdown recorded (re-upload the workbook to get exact
    branch-level duration filtering for it)."""
    raw_total = item.get(location_field) or 0.0
    if len(active_age_fields) >= len(AGEING_AGE_FIELDS):
        return raw_total
    buckets_json = item.get("location_age_buckets_json")
    if not buckets_json:
        return raw_total
    try:
        buckets_by_location = json.loads(buckets_json)
    except (TypeError, ValueError):
        return raw_total
    location_buckets = buckets_by_location.get(location_field)
    if not location_buckets:
        return raw_total
    return sum(location_buckets.get(f, 0.0) for f in active_age_fields)


def parse_ageing_durations_param(durations: Optional[str]) -> List[str]:
    """Turns the ?durations=age_0_60,age_181_365 query string into a
    validated, order-preserved list of age bucket keys. Blank/omitted or
    entirely unrecognized input falls back to all six buckets, so an old
    cached link or a stray typo never produces a report with zero age
    columns."""
    if not durations:
        return list(AGEING_AGE_FIELDS)
    requested = {d.strip() for d in durations.split(",") if d.strip()}
    active = [f for f in AGEING_AGE_FIELDS if f in requested]
    return active or list(AGEING_AGE_FIELDS)


AGEING_FONT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "fonts")
AGEING_FONT_REGULAR_PATH = os.path.join(AGEING_FONT_DIR, "DejaVuSans.ttf")
AGEING_FONT_BOLD_PATH = os.path.join(AGEING_FONT_DIR, "DejaVuSans-Bold.ttf")


def build_ageing_export_dataset(report: dict, active_age_fields: List[str], active_location_fields: Optional[List[str]] = None) -> Optional[dict]:
    """Turns the report payload from compute_ageing_stock_report() into a
    flat, format-agnostic structure shared by the .xlsx, .pdf, and .jpg
    exporters below, so all three always show identical numbers and never
    drift apart from each other.

    Deliberately does NOT carry a "Category" column: Category is kept as a
    coloured section heading (matching what's shown on screen) in every
    exported format instead of a flat table column, per the "Category
    shows on screen, not as a download column" requirement. Brand stays a
    section heading too, exactly mirroring the on-screen layout.
    Only items with stock in at least one of the checked duration buckets
    are included, matching the on-screen filtered view.

    active_location_fields narrows which branch columns (qty_alm, qty_hzt,
    etc.) are shown, mirroring the on-screen Branch picker - if the user
    has unchecked some branches there, the download only shows the
    branches still checked instead of always showing all six. Omit (or
    pass None) to show every branch, matching "all branches selected"."""
    loc_labels = {"qty_alm": "ALM", "qty_hzt": "HZT", "qty_ash": "ASH",
                  "qty_gng": "GNG", "qty_vkn": "VKN", "qty_mwh": "MWH"}

    location_fields = active_location_fields if active_location_fields else list(AGEING_LOCATION_FIELDS)

    def item_qty_in_active_durations(item: dict) -> float:
        return sum(item.get(f) or 0 for f in active_age_fields)

    def item_matches_location_total(item: dict) -> bool:
        location_total = sum(
            location_qty_for_durations(item, field, active_age_fields)
            for field in AGEING_LOCATION_FIELDS
        )
        return abs(item_qty_in_active_durations(item) - location_total) < 0.005

    headers = ["Item Details", "Unit", "Closing Qty"]
    headers += [AGEING_AGE_LABELS[f] for f in active_age_fields]
    headers += [loc_labels[f] for f in location_fields]
    headers += ["Present At"]

    zero_totals = lambda: {f: 0.0 for f in ["closing_qty"] + active_age_fields + location_fields}

    categories_out = []
    grand_totals = zero_totals()
    grand_total_items = 0
    any_rows = False

    for cat in report["categories"]:
        brands_out = []
        cat_totals = zero_totals()
        for brand_bucket in cat["brands"]:
            filtered_items = [
                it for it in brand_bucket["items"]
                if item_qty_in_active_durations(it) > 0 and item_matches_location_total(it)
            ]
            if not filtered_items:
                continue
            rows = []
            brand_totals = zero_totals()
            for item in filtered_items:
                # Closing Qty must always equal the sum of the currently
                # selected duration bucket(s), never the item's full raw
                # closing_qty - mirrors the same fix in ageing_stock.html's
                # addIntoTotals(), so the download always matches the
                # on-screen numbers with no discrepancy to explain.
                item_active_qty = round(item_qty_in_active_durations(item), 2)
                # Each branch column is scoped to the same selected
                # duration bucket(s) as Closing Qty above, not that
                # branch's all-time total - see location_qty_for_durations().
                loc_qtys = {f: location_qty_for_durations(item, f, active_age_fields) for f in location_fields}
                present_at_filtered = ", ".join(
                    loc_labels[f] for f in location_fields if loc_qtys[f] > 0
                ) or "-"
                row = [item["item_details"], item.get("unit") or ""]
                row.append(item_active_qty)
                row += [round(item.get(f) or 0, 2) for f in active_age_fields]
                row += [round(loc_qtys[f], 2) for f in location_fields]
                row.append(present_at_filtered)
                rows.append(row)
                brand_totals["closing_qty"] += item_active_qty
                for f in active_age_fields:
                    brand_totals[f] += item.get(f) or 0
                for f in location_fields:
                    brand_totals[f] += loc_qtys[f]
            any_rows = True
            brands_out.append({
                "name": brand_bucket["brand_name"],
                "rows": rows,
                "totals": brand_totals,
            })
            for f in ["closing_qty"] + active_age_fields + location_fields:
                cat_totals[f] += brand_totals[f]
            grand_total_items += len(filtered_items)

        if not brands_out:
            continue
        categories_out.append({
            "name": cat["category_name"],
            "code": cat["category_code"],
            "color": catColor_py(cat["category_code"]),
            "brands": brands_out,
            "totals": cat_totals,
        })
        for f in ["closing_qty"] + active_age_fields + location_fields:
            grand_totals[f] += cat_totals[f]

    if not any_rows:
        return None

    filt = report["filters"]
    subtitle_bits = [f"Category: {filt.get('category') or 'ALL'}"]
    if filt.get("brand"):
        subtitle_bits.append(f"Brand: {filt['brand']}")
    if filt.get("item_category"):
        subtitle_bits.append(f"Item Type: {filt['item_category']}")
    if filt.get("search"):
        subtitle_bits.append(f"Search: \"{filt['search']}\"")
    if filt.get("locations"):
        branch_codes = [c.strip().upper() for c in filt["locations"].split(",") if c.strip()]
        subtitle_bits.append("Branch: " + ", ".join(branch_codes))
    subtitle_bits.append("Durations: " + ", ".join(AGEING_AGE_LABELS[f] for f in active_age_fields))
    subtitle_bits.append(f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")

    return {
        "title": "Initiative ERP - Ageing Stock Analysis",
        "subtitle": "  |  ".join(subtitle_bits),
        "headers": headers,
        "active_age_fields": active_age_fields,
        # Selected branch columns (mirrors on-screen Branch picker) - the
        # xlsx/pdf/jpg builders below must use THIS list everywhere they
        # touch branch columns, never the AGEING_LOCATION_FIELDS constant
        # directly, or totals rows will have a different column count than
        # the header row once branches are filtered.
        "location_fields": location_fields,
        "categories": categories_out,
        "grand_totals": grand_totals,
        "grand_total_items": grand_total_items,
    }


# One accent colour per core category (HA/HE/Mobile/IT), plus Other -
# mirrors CATEGORY_COLORS in ageing_stock.html so every export format uses
# the exact same section-heading colours as the on-screen view.
AGEING_CATEGORY_COLORS_HEX = {"HA": "0F766E", "HE": "6D28D9", "MH": "15803D", "IT": "1D4ED8", "DC": "B45309", "OTHER": "475467"}


def catColor_py(code: Optional[str]) -> str:
    return AGEING_CATEGORY_COLORS_HEX.get((code or "").upper(), "111827")


def _ageing_fmt_num(v) -> str:
    """Shared number formatting for the PDF and JPG exports (Excel uses its
    own native number_format instead): 0 -> '-', whole numbers show with no
    decimal point, everything else rounds to 2 decimals. Matches fmt() in
    ageing_stock.html so every format displays numbers identically."""
    if v is None:
        return "-"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    if v == 0:
        return "-"
    return str(int(v)) if v.is_integer() else f"{round(v, 2)}"


def _ageing_build_xlsx(dataset: dict) -> bytes:
    import openpyxl as _openpyxl
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment, Border as _Border, Side as _Side
    from openpyxl.utils import get_column_letter as _get_column_letter

    headers = dataset["headers"]
    n_cols = len(headers)

    FONT_TITLE = _Font(name="Calibri", size=16, bold=True, color="111827")
    FONT_SUBTITLE = _Font(name="Calibri", size=9.5, italic=True, color="475467")
    FONT_HEADER = _Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    FONT_DATA = _Font(name="Calibri", size=10.5, color="1A1A1A")
    FONT_TOTAL = _Font(name="Calibri", size=10.5, bold=True, color="111827")
    FONT_CAT = _Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    FILL_HEADER = _PatternFill("solid", fgColor="111827")
    FILL_TOTAL = _PatternFill("solid", fgColor="EEF0F4")
    FILL_ROW_ALT = _PatternFill("solid", fgColor="FAFBFC")
    FILL_GRAND = _PatternFill("solid", fgColor="000000")
    FONT_GRAND = _Font(name="Calibri", size=11.5, bold=True, color="FFFFFF")
    THIN = _Side(style="thin", color="D0D5DD")
    BORDER_ALL = _Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUMFMT = '#,##0.##;-#,##0.##;"-"'

    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ageing Stock Report"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=dataset["title"])
    title_cell.font = FONT_TITLE
    title_cell.alignment = _Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    subtitle_cell = ws.cell(row=2, column=1, value=dataset["subtitle"])
    subtitle_cell.font = FONT_SUBTITLE
    subtitle_cell.alignment = _Alignment(horizontal="center", vertical="center")

    r = 4

    def write_row(values, font, fill=None, align_first_left=True):
        nonlocal r
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = font
            if fill:
                cell.fill = fill
            cell.border = BORDER_ALL
            if ci >= 3 and ci < n_cols:
                cell.number_format = NUMFMT
            left = (ci == 1 and align_first_left) or ci == n_cols
            cell.alignment = _Alignment(horizontal="left" if left else "right", vertical="center", wrap_text=(ci == 1))
        r += 1

    def totals_row_values(label, totals):
        vals = ["", label]
        vals[0] = ""
        row = [label, ""]
        row = [label] + [""] + [round(totals["closing_qty"], 2)] + \
              [round(totals[f], 2) for f in dataset["active_age_fields"]] + \
              [round(totals[f], 2) for f in dataset["location_fields"]] + [""]
        return row

    for cat in dataset["categories"]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        cat_cell = ws.cell(row=r, column=1, value=f"  {cat['name']}")
        cat_cell.font = FONT_CAT
        cat_cell.fill = _PatternFill("solid", fgColor=cat["color"])
        cat_cell.alignment = _Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1

        for brand in cat["brands"]:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            b_cell = ws.cell(row=r, column=1, value=f"   {brand['name']}")
            b_cell.font = _Font(name="Calibri", size=10.5, bold=True, italic=True, color="344054")
            b_cell.fill = _PatternFill("solid", fgColor="F2F4F7")
            b_cell.alignment = _Alignment(horizontal="left", vertical="center")
            r += 1

            for ci, h in enumerate(headers, start=1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.border = BORDER_ALL
                cell.alignment = _Alignment(horizontal="center", vertical="center", wrap_text=True)
            r += 1

            for i, row_vals in enumerate(brand["rows"]):
                fill = FILL_ROW_ALT if i % 2 == 1 else None
                write_row(row_vals, FONT_DATA, fill)

            write_row(totals_row_values(f"{brand['name']} Total", brand["totals"]), FONT_TOTAL, FILL_TOTAL)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    grand_label_cell = ws.cell(row=r, column=1, value=f"GRAND TOTAL  ({dataset['grand_total_items']} items)")
    grand_label_cell.font = FONT_GRAND
    grand_label_cell.fill = FILL_GRAND
    grand_label_cell.alignment = _Alignment(horizontal="left", vertical="center")
    for ci in range(2, n_cols + 1):
        ws.cell(row=r, column=ci).fill = FILL_GRAND
    r += 1
    gt = dataset["grand_totals"]
    grand_vals = [""] + [round(gt["closing_qty"], 2)] + \
                 [round(gt[f], 2) for f in dataset["active_age_fields"]] + \
                 [round(gt[f], 2) for f in dataset["location_fields"]] + [""]
    for ci, val in enumerate(grand_vals, start=1):
        cell = ws.cell(row=r, column=ci, value=val)
        cell.font = FONT_GRAND
        cell.fill = FILL_GRAND
        cell.border = BORDER_ALL
        if ci >= 2 and ci < n_cols:
            cell.number_format = NUMFMT
        cell.alignment = _Alignment(horizontal="left" if ci in (1, n_cols) else "right", vertical="center")

    widths = [42, 8, 11] + [11] * len(dataset["active_age_fields"]) + [8] * len(dataset["location_fields"]) + [24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[_get_column_letter(i)].width = w
    ws.freeze_panes = "B5"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _ageing_build_pdf(dataset: dict) -> bytes:
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Prefer the Unicode-capable DejaVu Sans font already bundled for the
    # JPG exporter, so special characters render correctly. But registering
    # a missing/uninstalled font file would raise TTFError and break the
    # whole PDF download, so this is strictly best-effort: if the .ttf
    # isn't actually present on disk (e.g. a dev checkout without
    # static/fonts populated), silently fall back to the built-in
    # Helvetica fonts instead of failing the export.
    font_regular, font_bold = "Helvetica", "Helvetica-Bold"
    if "AgeingSans" not in pdfmetrics.getRegisteredFontNames():
        try:
            pdfmetrics.registerFont(TTFont("AgeingSans", AGEING_FONT_REGULAR_PATH))
            pdfmetrics.registerFont(TTFont("AgeingSans-Bold", AGEING_FONT_BOLD_PATH))
        except Exception:
            pass
    if "AgeingSans" in pdfmetrics.getRegisteredFontNames():
        font_regular, font_bold = "AgeingSans", "AgeingSans-Bold"

    headers = dataset["headers"]
    n_cols = len(headers)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A3),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=12 * mm, bottomMargin=12 * mm,
        title=dataset["title"],
    )

    # Font sizes bumped up across the board (18->21 title, 7.5->10 table
    # body) - the previous sizes were tuned for print but read too small
    # when the PDF/JPG is viewed on screen, which is how most people
    # actually check the downloaded report.
    title_style = ParagraphStyle("AgeingTitle", fontName=font_bold, fontSize=21, leading=25, alignment=TA_CENTER, textColor=colors.HexColor("#111827"), spaceAfter=8)
    subtitle_style = ParagraphStyle("AgeingSubtitle", fontName=font_regular, fontSize=11, leading=15, alignment=TA_CENTER, textColor=colors.HexColor("#475467"), spaceBefore=2, spaceAfter=14)
    cat_style = ParagraphStyle("AgeingCat", fontName=font_bold, fontSize=15, alignment=TA_CENTER, textColor=colors.white, spaceBefore=0, spaceAfter=0, leading=24)
    brand_style = ParagraphStyle("AgeingBrand", fontName=font_bold, fontSize=12, textColor=colors.HexColor("#344054"), spaceBefore=8, spaceAfter=4, leftIndent=2)
    # Item Details cells: bold and wrapped as a Paragraph (not a plain
    # string) so a long product name word-wraps inside its own column
    # instead of overflowing into the Unit/Closing Qty columns next to it.
    item_style = ParagraphStyle("AgeingItem", fontName=font_bold, fontSize=10, leading=12.5, alignment=TA_LEFT, textColor=colors.HexColor("#101828"))

    # Even with the fallback Helvetica font, don't depend on it supporting
    # "\u2265" ("≥") - swap it for plain ">=" so the subtitle always
    # renders correctly regardless of which font ended up being used.
    pdf_subtitle = dataset["subtitle"].replace("\u2265", ">=")

    page_width = landscape(A3)[0] - 28 * mm
    first_col_w = page_width * 0.24
    last_col_w = page_width * 0.14
    remaining = page_width - first_col_w - last_col_w
    other_cols = n_cols - 2
    col_w = remaining / other_cols
    col_widths = [first_col_w] + [col_w] * other_cols + [last_col_w]

    story = []
    story.append(Paragraph(dataset["title"], title_style))
    story.append(Paragraph(pdf_subtitle, subtitle_style))
    story.append(Spacer(1, 6))

    for cat in dataset["categories"]:
        cat_table = Table([[Paragraph(cat["name"], cat_style)]], colWidths=[page_width])
        cat_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{cat['color']}")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(cat_table)

        for brand in cat["brands"]:
            story.append(Paragraph(brand["name"], brand_style))

            table_data = [headers]
            for row in brand["rows"]:
                cells = []
                for col_idx, v in enumerate(row):
                    if col_idx == 0:
                        # Item Details: full text, bold, word-wrapped.
                        cells.append(Paragraph(str(v) if v not in (None, "") else "-", item_style))
                    elif isinstance(v, (int, float)):
                        cells.append(_ageing_fmt_num(v))
                    else:
                        cells.append(v if v not in (None, "") else "-")
                table_data.append(cells)
            bt = brand["totals"]
            total_row = [f"{brand['name']} Total", ""] + \
                [_ageing_fmt_num(bt["closing_qty"])] + \
                [_ageing_fmt_num(bt[f]) for f in dataset["active_age_fields"]] + \
                [_ageing_fmt_num(bt[f]) for f in dataset["location_fields"]] + [""]
            table_data.append(total_row)

            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), font_bold),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 1), (-1, -1), font_regular),
                ("FONTSIZE", (0, 1), (-1, -1), 9.5),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
                ("ALIGN", (-1, 1), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D0D5DD")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEF0F4")),
                ("FONTNAME", (0, -1), (-1, -1), font_bold),
                ("FONTSIZE", (0, -1), (-1, -1), 9.5),
                ("SPAN", (0, -1), (1, -1)),
            ]
            for i in range(1, len(table_data) - 1):
                if i % 2 == 0:
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FAFBFC")))
            t.setStyle(TableStyle(style_cmds))
            story.append(t)
            story.append(Spacer(1, 6))

    gt = dataset["grand_totals"]
    grand_row = [f"GRAND TOTAL  ({dataset['grand_total_items']} items)", ""] + \
                [_ageing_fmt_num(gt["closing_qty"])] + \
                [_ageing_fmt_num(gt[f]) for f in dataset["active_age_fields"]] + \
                [_ageing_fmt_num(gt[f]) for f in dataset["location_fields"]] + [""]
    gt_table = Table([grand_row], colWidths=col_widths)
    gt_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.black),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_bold),
        ("FONTSIZE", (0, 0), (-1, -1), 12),
        ("ALIGN", (1, 0), (-1, 0), "RIGHT"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
        ("SPAN", (0, 0), (1, 0)),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#333333")),
    ]))
    story.append(Spacer(1, 4))
    story.append(gt_table)

    doc.build(story)
    return buffer.getvalue()


def _ageing_render_jpg_page(headers, title, subtitle, categories, active_age_fields,
                             grand_totals=None, grand_total_items=None, location_fields=None) -> bytes:
    """Renders one image containing the given categories (plus the grand
    total footer, only when grand_totals is passed in on the final page).
    Shared by both the single-image and multi-page (zipped) paths below so
    every page looks identical in style."""
    from PIL import Image, ImageDraw, ImageFont

    n_cols = len(headers)
    location_fields = location_fields if location_fields is not None else list(AGEING_LOCATION_FIELDS)

    def load_font(bold, size):
        # 1) Preferred: the bundled DejaVu Sans (best Unicode coverage,
        #    e.g. the "≥" in "≥ 366 Days"), when static/fonts actually has
        #    it. 2) Guaranteed fallback: Bitstream Vera, which ships
        #    inside the reportlab package - reportlab is already a hard
        #    requirement for the PDF export, so this file is always on
        #    disk even when static/fonts isn't. Both are real scalable
        #    TTFs, unlike PIL's own load_default() below, which silently
        #    ignores the requested size and always draws a ~10px font -
        #    that mismatch (sizes set here, but rendered tiny) was why the
        #    downloaded JPG looked fine in code but tiny in practice.
        candidates = [AGEING_FONT_BOLD_PATH if bold else AGEING_FONT_REGULAR_PATH]
        try:
            import reportlab
            rl_fonts_dir = os.path.join(os.path.dirname(reportlab.__file__), "fonts")
            candidates.append(os.path.join(rl_fonts_dir, "VeraBd.ttf" if bold else "Vera.ttf"))
        except Exception:
            pass
        for path in candidates:
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
        try:
            return ImageFont.load_default(size=size)  # Pillow 10.1+
        except TypeError:
            return ImageFont.load_default()

    # Font sizes bumped up across the board so the downloaded image reads
    # clearly at normal viewing zoom, not just when printed large. Every
    # pixel offset below (row heights, vertical-centering offsets, wrapped
    # line spacing) is scaled to match so nothing overlaps or misaligns.
    F_TITLE = load_font(True, 36)
    F_SUBTITLE = load_font(False, 18)
    F_CAT = load_font(True, 23)
    F_BRAND = load_font(True, 18)
    F_HEAD = load_font(True, 17)
    F_DATA = load_font(False, 17)
    F_TOTAL = load_font(True, 17)

    PAD = 34
    ROW_H = 38
    HEAD_H = 44
    CAT_H = 50
    BRAND_H = 40
    SUBTITLE_LINE_H = 26

    tmp_img = Image.new("RGB", (10, 10))
    tmp_draw = ImageDraw.Draw(tmp_img)

    # Middle columns (Unit, Closing Qty, each duration bucket, each
    # location) are sized to fit their own header label rather than a
    # fixed guess - at the larger header font, a fixed width let long
    # labels like "Closing Qty" or "≥ 366 Days" overflow into the next
    # column instead of just being tight.
    first_col_w = 460
    last_col_w = 290
    middle_headers = headers[1:-1]
    middle_col_widths = [max(96, tmp_draw.textlength(h, font=F_HEAD) + 28) for h in middle_headers]
    col_widths = [first_col_w] + middle_col_widths + [last_col_w]
    table_w = sum(col_widths)
    img_w = table_w + PAD * 2

    def wrap_text(draw, text, font, max_w):
        words = str(text).split(" ")
        lines, cur = [], ""
        for w in words:
            trial = (cur + " " + w).strip()
            if draw.textlength(trial, font=font) <= max_w - 10:
                cur = trial
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        return lines or [""]

    # The subtitle line (Category / Brand / Item Type / Search / Durations
    # / Generated, joined with " | ") can run far wider than the image once
    # several filters are active. Drawing it as a single un-wrapped centered
    # line let it overflow past the image edge and visually collapse into
    # the border below. Wrapping it onto as many centered lines as it needs
    # - each its own full-width line, never squeezed - fixes that, and the
    # title block below grows to fit however many lines that turns out to be.
    subtitle_lines = wrap_text(tmp_draw, subtitle, F_SUBTITLE, img_w - PAD * 2)
    subtitle_block_h = len(subtitle_lines) * SUBTITLE_LINE_H
    TITLE_BLOCK_H = 80 + subtitle_block_h + 17

    row_defs = [("title", None, PAD + 80), ("subtitle", subtitle_lines, subtitle_block_h), ("gap", None, 17)]
    for cat in categories:
        row_defs.append(("cat", cat, CAT_H))
        for brand in cat["brands"]:
            row_defs.append(("brand", brand, BRAND_H))
            row_defs.append(("head", None, HEAD_H))
            for row in brand["rows"]:
                wrapped = wrap_text(tmp_draw, row[0], F_DATA, first_col_w)
                h = max(ROW_H, 26 * len(wrapped) + 12)
                row_defs.append(("data", (row, wrapped), h))
            row_defs.append(("btotal", brand, ROW_H + 6))
            row_defs.append(("bgap", None, 8))
    if grand_totals is not None:
        row_defs.append(("gtotal", (grand_totals, grand_total_items), ROW_H + 16))
    row_defs.append(("footer_gap", None, PAD))

    img_h = sum(h for _, _, h in row_defs) + PAD
    img = Image.new("RGB", (int(img_w), int(img_h)), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    ink = "#0B1220"
    muted = "#475467"
    line_col = "#D0D5DD"
    header_fill = "#111827"

    y = PAD
    x0 = PAD

    edges = [x0]
    for w in col_widths:
        edges.append(edges[-1] + w)

    def center_text(text, font, cy, fill):
        w = draw.textlength(str(text), font=font)
        draw.text((img_w / 2 - w / 2, cy), str(text), font=font, fill=fill)

    def draw_row_bg(y_top, y_bot, color):
        draw.rectangle([x0, y_top, edges[-1], y_bot], fill=color)

    def draw_cell_text(ci, text, y_top, y_bot, font, fill, align="right"):
        cell_x0, cell_x1 = edges[ci], edges[ci + 1]
        w = draw.textlength(str(text), font=font)
        ty = y_top + (y_bot - y_top) / 2 - 10
        if align == "left":
            tx = cell_x0 + 8
        elif align == "center":
            tx = cell_x0 + (cell_x1 - cell_x0) / 2 - w / 2
        else:
            tx = cell_x1 - w - 8
        draw.text((tx, ty), str(text), font=font, fill=fill)

    def numfmt(v):
        return "-" if v == 0 else (int(v) if float(v).is_integer() else round(v, 2))

    for kind, payload, h in row_defs:
        if kind == "title":
            center_text(title, F_TITLE, y + 16, "#111827")
        elif kind == "subtitle":
            for li, line in enumerate(payload):
                center_text(line, F_SUBTITLE, y + 3 + li * SUBTITLE_LINE_H, muted)
        elif kind in ("gap", "bgap", "footer_gap"):
            pass
        elif kind == "cat":
            draw.rectangle([x0, y, edges[-1], y + h], fill=f"#{payload['color']}")
            cw = draw.textlength(payload["name"], font=F_CAT)
            draw.text((img_w / 2 - cw / 2, y + h / 2 - 13), payload["name"], font=F_CAT, fill="#FFFFFF")
        elif kind == "brand":
            draw.rectangle([x0, y, edges[-1], y + h], fill="#F2F4F7")
            draw.text((x0 + 8, y + h / 2 - 11), payload["name"], font=F_BRAND, fill="#344054")
        elif kind == "head":
            draw.rectangle([x0, y, edges[-1], y + h], fill=header_fill)
            for ci, htext in enumerate(headers):
                align = "left" if ci == 0 or ci == n_cols - 1 else "center"
                draw_cell_text(ci, htext, y, y + h, F_HEAD, "#FFFFFF", align)
        elif kind == "data":
            row, wrapped = payload
            for li, line in enumerate(wrapped):
                draw.text((x0 + 8, y + 7 + li * 26), line, font=F_DATA, fill=ink)
            for ci in range(1, n_cols):
                val = row[ci]
                if isinstance(val, (int, float)):
                    val = numfmt(val)
                align = "left" if ci == n_cols - 1 else "right"
                draw_cell_text(ci, val, y, y + h, F_DATA, ink, align)
        elif kind == "btotal":
            draw_row_bg(y, y + h, "#EEF0F4")
            draw_cell_text(0, f"{payload['name']} Total", y, y + h, F_TOTAL, "#111827", "left")
            totals = payload["totals"]
            vals = [totals["closing_qty"]] + [totals[f] for f in active_age_fields] + [totals[f] for f in location_fields]
            for i, v in enumerate(vals, start=2):
                draw_cell_text(i - 1, numfmt(v), y, y + h, F_TOTAL, "#111827", "right")
        elif kind == "gtotal":
            gt, gt_items = payload
            draw_row_bg(y, y + h, "#000000")
            draw_cell_text(0, f"GRAND TOTAL  ({gt_items} items)", y, y + h, F_TOTAL, "#FFFFFF", "left")
            vals = [gt["closing_qty"]] + [gt[f] for f in active_age_fields] + [gt[f] for f in location_fields]
            for i, v in enumerate(vals, start=2):
                draw_cell_text(i - 1, numfmt(v), y, y + h, F_TOTAL, "#FFFFFF", "right")
        y += h

    draw.rectangle([x0, PAD + TITLE_BLOCK_H, edges[-1], img_h - PAD], outline=line_col, width=1)

    out = BytesIO()
    img.convert("RGB").save(out, format="JPEG", quality=92, optimize=True)
    return out.getvalue()


def _ageing_estimate_category_height(headers, cat) -> int:
    """Quick row-count-based estimate (no font metrics needed) of how tall
    a single category's block will render at, used only to decide where to
    split pages - doesn't need to be pixel-perfect, just larger than the
    real height would ever exceed it by much, since row heights only grow
    with wrapped text, never shrink below the base row height."""
    ROW_H, HEAD_H, CAT_H, BRAND_H = 38, 44, 50, 40
    h = CAT_H
    for brand in cat["brands"]:
        h += BRAND_H + HEAD_H + len(brand["rows"]) * ROW_H + (ROW_H + 6) + 8
    return h


# A single JPEG can't exceed this many pixels tall (Pillow/libjpeg's hard
# limit is 65500) - stay comfortably under it so per-row wrapped-text
# growth never tips a page over the edge.
AGEING_JPG_MAX_PAGE_HEIGHT = 58000


def _ageing_build_jpg(dataset: dict):
    """Builds the JPG export. Small/medium filtered reports come back as a
    single .jpg. Large, mostly-unfiltered reports (many thousand rows)
    would exceed JPEG's maximum image height as one image, so those are
    automatically split one-image-per-category and bundled into a single
    .zip - still one click, still every category included, just delivered
    as a small set of images instead of one impossibly tall picture.
    Returns (content_bytes, media_type, file_extension)."""
    headers = dataset["headers"]
    title = dataset["title"]
    subtitle = dataset["subtitle"]
    categories = dataset["categories"]
    active_age_fields = dataset["active_age_fields"]
    location_fields = dataset["location_fields"]

    base_header_h = 34 + 80 + 31 + 17 + 34
    total_est = base_header_h + sum(_ageing_estimate_category_height(headers, c) for c in categories) + 60

    if total_est <= AGEING_JPG_MAX_PAGE_HEIGHT:
        content = _ageing_render_jpg_page(
            headers, title, subtitle, categories, active_age_fields,
            grand_totals=dataset["grand_totals"], grand_total_items=dataset["grand_total_items"],
            location_fields=location_fields,
        )
        return content, "image/jpeg", "jpg"

    # Too tall for one image: bucket categories into pages, splitting a
    # single very-large category across pages if even that alone would be
    # too tall (defensive - not hit by today's data volumes).
    pages = []
    current, current_h = [], base_header_h
    for cat in categories:
        cat_h = _ageing_estimate_category_height(headers, cat)
        if cat_h > AGEING_JPG_MAX_PAGE_HEIGHT:
            if current:
                pages.append(current)
                current, current_h = [], base_header_h
            n_brands = len(cat["brands"])
            chunk = max(1, n_brands // (cat_h // AGEING_JPG_MAX_PAGE_HEIGHT + 1))
            for i in range(0, n_brands, chunk):
                sub_brands = cat["brands"][i:i + chunk]
                sub_totals = {f: 0.0 for f in ["closing_qty"] + active_age_fields + location_fields}
                for b in sub_brands:
                    for f in sub_totals:
                        sub_totals[f] += b["totals"][f]
                pages.append([{**cat, "brands": sub_brands, "totals": sub_totals,
                               "name": f"{cat['name']} (part {i // chunk + 1})"}])
            continue
        if current and current_h + cat_h > AGEING_JPG_MAX_PAGE_HEIGHT:
            pages.append(current)
            current, current_h = [], base_header_h
        current.append(cat)
        current_h += cat_h
    if current:
        pages.append(current)

    import zipfile
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, page_cats in enumerate(pages, start=1):
            is_last = (i == len(pages))
            page_subtitle = f"{subtitle}  |  Page {i} of {len(pages)}"
            page_bytes = _ageing_render_jpg_page(
                headers, title, page_subtitle, page_cats, active_age_fields,
                grand_totals=dataset["grand_totals"] if is_last else None,
                grand_total_items=dataset["grand_total_items"] if is_last else None,
                location_fields=location_fields,
            )
            cat_label = re.sub(r"[^A-Za-z0-9]+", "_", page_cats[0]["name"])[:30]
            zf.writestr(f"Page_{i:02d}_{cat_label}.jpg", page_bytes)

    return zip_buffer.getvalue(), "application/zip", "zip"


@app.get("/api/ageing-stock/export")
def ageing_stock_export(
    category: Optional[str] = Query(None),
    brand: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    item_category: Optional[str] = Query(None, description="Item-type filter code (e.g. LED, COOLER, AC) matching the on-screen Item Category picker. Omit for all types."),
    locations: Optional[str] = Query(None, description="Comma-separated branch codes (ALM,HZT,ASH,GNG,VKN,MWH) matching the on-screen Branch picker - keeps items with stock at ANY of them. Omit for all branches."),
    durations: Optional[str] = Query(None, description="Comma-separated age bucket keys to include, e.g. age_0_60,age_181_365. Omit for all."),
    format: str = Query("xlsx", description="Download format: xlsx, pdf, or jpg."),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    """Downloads the currently-filtered Ageing Stock Analysis report - same
    category/brand/search/item-category/branch scope as the on-screen view, plus
    whichever ageing-duration columns the user has checked - as .xlsx,
    .pdf, or .jpg, with a Grand Total row/section at the bottom. Every
    filter here mirrors what's visible on screen so the download always
    matches exactly what the user is currently looking at. Category is
    shown as a coloured section heading in every format (matching the
    on-screen layout) rather than as a flat table column."""
    fmt = (format or "xlsx").strip().lower()
    if fmt not in ("xlsx", "pdf", "jpg"):
        raise HTTPException(status_code=400, detail="format must be one of: xlsx, pdf, jpg.")

    report = compute_ageing_stock_report(db, category, brand, search, item_category, locations)
    if not report["has_data"] or not report["categories"]:
        raise HTTPException(status_code=404, detail="No ageing stock data available for the selected filters.")

    active_age_fields = parse_ageing_durations_param(durations)

    # Narrow the export's branch columns to whichever branches the user
    # left checked in the on-screen Branch picker - mirrors activeLocCols()
    # in ageing_stock.html so the download never shows more branch columns
    # than what's currently visible on screen.
    active_location_fields = None
    if locations:
        codes = [c.strip().lower() for c in locations.split(",") if c.strip()]
        active_location_fields = [f"qty_{c}" for c in codes if f"qty_{c}" in AGEING_LOCATION_FIELDS]

    dataset = build_ageing_export_dataset(report, active_age_fields, active_location_fields)
    if dataset is None:
        raise HTTPException(status_code=404, detail="No items have stock in the selected duration(s).")

    fname_bits = [report["filters"].get("category") or "ALL"]
    if report["filters"].get("item_category"):
        fname_bits.append(report["filters"]["item_category"])
    if report["filters"].get("search"):
        fname_bits.append("search")
    base_filename = f"Ageing_Stock_Report_{'_'.join(fname_bits)}".replace(" ", "_")

    if fmt == "xlsx":
        content = _ageing_build_xlsx(dataset)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{base_filename}.xlsx"
    elif fmt == "pdf":
        content = _ageing_build_pdf(dataset)
        media_type = "application/pdf"
        filename = f"{base_filename}.pdf"
    else:
        content, media_type, ext = _ageing_build_jpg(dataset)
        filename = f"{base_filename}.{ext}"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/api/ageing-stock/clear")
def clear_ageing_stock_data(
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    deleted_count = db.query(models.AgeingStockItem).delete()
    db.query(models.AgeingStockUpload).delete()
    db.commit()
    return {"message": "Ageing Stock Analysis data cleared", "deleted": deleted_count}


INCENTIVE_MONTH_ALIASES = {"month", "months", "period", "sale month", "sales month"}
INCENTIVE_OUTLET_ALIASES = {"outlet", "outlet name", "branch", "branch name", "store", "store name", "manager", "manager name"}
INCENTIVE_SALES_ALIASES = {"total sale", "total sales", "sales", "sale", "sales amount", "sale amount", "monthly sales", "net sales"}


def _incentive_header(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _incentive_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return float(value)
    cleaned = re.sub(r"[^0-9.()-]", "", str(value).strip()).replace("(", "-").replace(")", "")
    try:
        return float(cleaned) if cleaned not in {"", "-", "."} else None
    except ValueError:
        return None


def _incentive_month(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%b %Y")
    if isinstance(value, (int, float)) and 20000 < value < 80000:
        # Excel's 1900 date system (including its historical leap-year offset).
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).strftime("%b %Y")
    return str(value or "").strip()


def _incentive_month_from_filename(filename: Optional[str]) -> str:
    stem = re.sub(r"[_-]+", " ", (filename or "").rsplit(".", 1)[0])
    match = re.search(
        r"\b(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec)\b.*?\b(20\d{2})\b",
        stem,
        re.IGNORECASE,
    )
    if not match:
        return stem.strip() or "Uploaded Report"
    parsed = datetime.strptime(match.group(1)[:3].title(), "%b")
    return f"{parsed.strftime('%b')} {match.group(2)}"


def _parse_incentive_outlet_category_rows(rows: list, month: str) -> Optional[list]:
    """Parse reports shaped as Outlets | Category | Sale.

    An outlet name starts a block and each category sale becomes one source
    row. The following TOTAL line is skipped so outlet totals can be rebuilt
    without double-counting while category detail remains available.
    """
    header_idx = next((
        idx for idx, row in enumerate(rows[:15])
        if any(_incentive_header(value) == "outlets" for value in row)
        and any(_incentive_header(value) == "category" for value in row)
        and any(_incentive_header(value) in {"sale", "sales", "total sale", "total sales"} for value in row)
    ), None)
    if header_idx is None:
        return None
    parsed, current_outlet = [], None
    ignored_outlets = {"outlets", "sum all branch", "grand total", "total"}
    for row in rows[header_idx + 1:]:
        outlet_value = str(row[0] or "").strip() if row else ""
        category_value = str(row[1] or "").strip() if len(row) > 1 else ""
        sale = _incentive_number(row[2]) if len(row) > 2 else None
        if outlet_value and _incentive_header(outlet_value) not in ignored_outlets:
            current_outlet = outlet_value
        if not current_outlet:
            continue
        if _incentive_header(category_value) == "total":
            current_outlet = None
        elif category_value and sale is not None:
            parsed.append({"month": month, "outlet": current_outlet, "category": category_value, "total_sales": sale})
    return parsed


def parse_incentive_workbook(content: bytes, filename: Optional[str] = None) -> list:
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {exc}")

    result = []
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10000), values_only=True)]
        if not rows:
            continue
        width = min(max((len(r) for r in rows), default=0), 500)
        rows = [(r + [None] * width)[:width] for r in rows]
        scan_limit = min(15, len(rows))

        outlet_category_rows = _parse_incentive_outlet_category_rows(
            rows, _incentive_month_from_filename(filename)
        )
        if outlet_category_rows is not None:
            result.extend(outlet_category_rows)
            continue

        # Long layout: Month | Outlet | Total Sales.
        handled = False
        for header_idx in range(scan_limit):
            normalized = [_incentive_header(v) for v in rows[header_idx]]
            outlet_col = next((i for i, h in enumerate(normalized) if h in INCENTIVE_OUTLET_ALIASES), None)
            sales_col = next((i for i, h in enumerate(normalized) if h in INCENTIVE_SALES_ALIASES), None)
            month_col = next((i for i, h in enumerate(normalized) if h in INCENTIVE_MONTH_ALIASES), None)
            if outlet_col is None or sales_col is None:
                continue
            for row in rows[header_idx + 1:]:
                sales = _incentive_number(row[sales_col])
                outlet = str(row[outlet_col] or "").strip()
                if sales is None or not outlet:
                    continue
                month = _incentive_month(row[month_col]) if month_col is not None else ws.title
                result.append({"month": month or ws.title, "outlet": outlet, "total_sales": sales})
            handled = True
            break
        if handled:
            continue

        # Grouped reference layout: manager/outlet names on one row and
        # repeated Total Sale / Salary / Incentive columns on the next row.
        for subheader_idx in range(1, scan_limit):
            sales_cols = [i for i, v in enumerate(rows[subheader_idx]) if _incentive_header(v) in INCENTIVE_SALES_ALIASES]
            if not sales_cols:
                continue
            month_col = next((i for i, v in enumerate(rows[subheader_idx - 1]) if _incentive_header(v) in INCENTIVE_MONTH_ALIASES), 0)
            outlets = {}
            for col in sales_cols:
                name = ""
                for lookup_col in range(col, -1, -1):
                    candidate = str(rows[subheader_idx - 1][lookup_col] or "").strip()
                    if candidate and _incentive_header(candidate) not in INCENTIVE_MONTH_ALIASES:
                        name = candidate
                        break
                outlets[col] = name or f"Outlet {col + 1}"
            for row in rows[subheader_idx + 1:]:
                month = _incentive_month(row[month_col])
                if not month:
                    continue
                for col, outlet in outlets.items():
                    sales = _incentive_number(row[col])
                    if sales is not None:
                        result.append({"month": month, "outlet": outlet, "total_sales": sales})
            handled = True
            break
        if handled:
            continue

        # Wide layout: Month | Outlet A | Outlet B | ...
        for header_idx in range(scan_limit):
            normalized = [_incentive_header(v) for v in rows[header_idx]]
            month_col = next((i for i, h in enumerate(normalized) if h in INCENTIVE_MONTH_ALIASES), None)
            if month_col is None:
                continue
            outlet_cols = [i for i, value in enumerate(rows[header_idx]) if i != month_col and str(value or "").strip()]
            for row in rows[header_idx + 1:]:
                month = _incentive_month(row[month_col])
                if not month:
                    continue
                for col in outlet_cols:
                    sales = _incentive_number(row[col])
                    if sales is not None:
                        result.append({"month": month, "outlet": str(rows[header_idx][col]).strip(), "total_sales": sales})
            break

    if not result:
        raise HTTPException(status_code=400, detail="No readable sales rows found. Use Month, Outlet and Total Sales columns, or a Month row with outlet-wise sales columns.")
    return result


def parse_incentive_pdf(content: bytes, filename: Optional[str] = None) -> list:
    try:
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(content))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read the PDF file: {exc}")
    if not text.strip():
        raise HTTPException(status_code=400, detail="The PDF has no selectable text. Export it from Excel instead of uploading a scanned image PDF.")

    month = _incentive_month_from_filename(filename)
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
    category_pattern = r"(?:MOB|COM|DC|HA|HE|ACC(?:\s+COM|\s+UNB)?)"
    ignored = {"outlets", "category", "sale", "base table total monthly sale"}
    parsed, current_outlet = [], None
    for index, line in enumerate(lines):
        # Some PDF exporters keep all three table cells on one text line.
        category_row = re.match(rf"^(.+?)\s+({category_pattern})\s+([₹,0-9.() -]+)$", line, re.IGNORECASE)
        if category_row:
            sale = _incentive_number(category_row.group(3))
            if sale is not None:
                parsed.append({"month": month, "outlet": category_row.group(1).strip(), "category": category_row.group(2).upper(), "total_sales": sale})
            continue
        combined = re.match(rf"^(.+?)\s+TOTAL\s+([₹,0-9.() -]+)$", line, re.IGNORECASE)
        if combined:
            continue
        total = re.match(r"^TOTAL\s+([₹,0-9.() -]+)$", line, re.IGNORECASE)
        if total and current_outlet:
            current_outlet = None
            continue
        category_only = re.match(rf"^({category_pattern})\s+([₹,0-9.() -]+)$", line, re.IGNORECASE)
        if category_only and current_outlet:
            sale = _incentive_number(category_only.group(2))
            if sale is not None:
                parsed.append({"month": month, "outlet": current_outlet, "category": category_only.group(1).upper(), "total_sales": sale})
            continue
        normalized = _incentive_header(line)
        next_line = lines[index + 1] if index + 1 < len(lines) else ""
        if normalized not in ignored and not re.match(rf"^(?:{category_pattern}|TOTAL)\b", line, re.IGNORECASE) and re.match(rf"^{category_pattern}\b", next_line, re.IGNORECASE):
            current_outlet = line
    if not parsed:
        raise HTTPException(status_code=400, detail="No readable outlet/category sales rows found in the PDF. Use an exported text PDF with Outlets, Category and Sale columns.")
    return parsed


def parse_incentive_upload(content: bytes, filename: Optional[str]) -> list:
    extension = (filename or "").lower().rsplit(".", 1)[-1]
    if extension == "pdf":
        return parse_incentive_pdf(content, filename)
    if extension in {"xlsx", "xlsm"}:
        return parse_incentive_workbook(content, filename)
    raise HTTPException(status_code=400, detail="Upload an Excel (.xlsx/.xlsm) or PDF (.pdf) file.")


def calculate_incentive_rows(rows: list, profit_rate: float, incentive_rate: float) -> list:
    calculated = []
    for row in rows:
        sales = round(float(row["total_sales"]), 2)
        avg_profit = round(sales * profit_rate / 100.0, 2)
        incentive = round(avg_profit * incentive_rate / 100.0, 2)
        calculated.append({**row, "outlet": _incentive_outlet_short_name(row["outlet"]), "total_sales": sales, "avg_profit": avg_profit, "total_incentive": incentive})
    return calculated


def _incentive_outlet_short_name(value: Optional[str]) -> str:
    original = str(value or "").strip()
    normalized = _incentive_header(original).replace(" ", "")
    return {
        "alambagh": "ALM", "alm": "ALM",
        "ashiyana": "ASH", "ash": "ASH",
        "gomtinagar": "GNG", "gng": "GNG",
        "hazratganj": "HZT", "hzt": "HZT",
        "vikasnagar": "VKN", "vkn": "VKN",
    }.get(normalized, original)


def _incentive_report_category(value: Optional[str]) -> str:
    normalized = _incentive_header(value).upper()
    return normalized if normalized in {"HA", "HE", "MOB", "COM", "DC"} else "ACC"


def _incentive_outlet_groups(outlet: str) -> list:
    normalized = _incentive_header(outlet).replace(" ", "")
    if normalized in {"alambagh", "alm"}:
        return [("HA + HE", ("HA", "HE")), ("MOB + DC + ACC", ("MOB", "DC", "ACC")), ("COM", ("COM",))]
    if normalized in {"hazratganj", "hzt"}:
        return [("HA + HE", ("HA", "HE")), ("MOB + COM + DC + ACC", ("MOB", "COM", "DC", "ACC"))]
    if normalized in {"gomtinagar", "gng", "ashiyana", "ash", "vikasnagar", "vkn"}:
        return [("ALL", ("HA", "HE", "MOB", "COM", "DC", "ACC"))]
    return [("ALL", ("HA", "HE", "MOB", "COM", "DC", "ACC"))]


@app.post("/api/incentive/calculate")
def calculate_incentive_report(
    file: UploadFile = File(...),
    profit_rate: float = Form(7.0),
    incentive_rate: float = Form(2.5),
    current_user: models.User = Depends(auth.require_roles("Admin", "Owner", "Accounts", "MISExecutive", "HR")),
):
    if not (0 <= profit_rate <= 100 and 0 <= incentive_rate <= 100):
        raise HTTPException(status_code=400, detail="Both rates must be between 0 and 100.")
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    rows = calculate_incentive_rows(parse_incentive_upload(content, file.filename), profit_rate, incentive_rate)
    by_outlet = defaultdict(lambda: {"total_sales": 0.0, "avg_profit": 0.0, "total_incentive": 0.0})
    by_outlet_category = defaultdict(lambda: {category: {"total_sales": 0.0, "avg_profit": 0.0, "total_incentive": 0.0} for category in ("HA", "HE", "MOB", "COM", "DC", "ACC")})
    for row in rows:
        for field in ("total_sales", "avg_profit", "total_incentive"):
            by_outlet[row["outlet"]][field] += row[field]
        category = _incentive_report_category(row.get("category"))
        by_outlet_category[row["outlet"]][category]["total_sales"] += row["total_sales"]
        by_outlet_category[row["outlet"]][category]["avg_profit"] += row["avg_profit"]
        by_outlet_category[row["outlet"]][category]["total_incentive"] += row["total_incentive"]
    # Preserve the original outlet report's calculation order: calculate
    # profit from the outlet total, then incentive from that profit. This
    # avoids category-level rounding changing the existing report by cents.
    for totals in by_outlet.values():
        totals["avg_profit"] = round(totals["total_sales"] * profit_rate / 100.0, 2)
        totals["total_incentive"] = round(totals["avg_profit"] * incentive_rate / 100.0, 2)
    grand_sales = round(sum(totals["total_sales"] for totals in by_outlet.values()), 2)
    grand_profit = round(grand_sales * profit_rate / 100.0, 2)
    grand_incentive = round(grand_profit * incentive_rate / 100.0, 2)
    grouped_summary = []
    for outlet, categories in sorted(by_outlet_category.items()):
        for label, category_names in _incentive_outlet_groups(outlet):
            grouped_summary.append({
                "outlet": outlet,
                "group": label,
                "total_sales": round(sum(categories[name]["total_sales"] for name in category_names), 2),
                "avg_profit": round(sum(categories[name]["avg_profit"] for name in category_names), 2),
                "total_incentive": round(sum(categories[name]["total_incentive"] for name in category_names), 2),
            })
    return {
        "rows": rows,
        "summary": [{"outlet": outlet, **{k: round(v, 2) for k, v in totals.items()}} for outlet, totals in sorted(by_outlet.items())],
        "category_summary": [{
            "outlet": outlet,
            "categories": {category: {field: round(value, 2) for field, value in values.items()} for category, values in categories.items()},
            "total": {
                "avg_profit": by_outlet[outlet]["avg_profit"],
                "total_incentive": by_outlet[outlet]["total_incentive"],
            },
        } for outlet, categories in sorted(by_outlet_category.items())],
        "grouped_summary": grouped_summary,
        "totals": {"total_sales": grand_sales, "avg_profit": grand_profit, "total_incentive": grand_incentive},
        "profit_rate": profit_rate,
        "incentive_rate": incentive_rate,
    }


@app.post("/api/incentive/export")
def export_incentive_report(
    file: UploadFile = File(...),
    profit_rate: float = Form(7.0),
    incentive_rate: float = Form(2.5),
    current_user: models.User = Depends(auth.require_roles("Admin", "Owner", "Accounts", "MISExecutive", "HR")),
):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if not (0 <= profit_rate <= 100 and 0 <= incentive_rate <= 100):
        raise HTTPException(status_code=400, detail="Both rates must be between 0 and 100.")
    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    parsed_rows = parse_incentive_upload(content, file.filename)
    source_totals = defaultdict(float)
    for row in parsed_rows:
        source_totals[(row["month"], _incentive_outlet_short_name(row["outlet"]))] += float(row["total_sales"])
    # Keep the existing downloadable outlet report unchanged even though the
    # parser now retains category detail for the additional on-screen matrix.
    source_rows = [
        {"month": month, "outlet": outlet, "total_sales": round(total, 2)}
        for (month, outlet), total in sorted(source_totals.items())
    ]
    book = Workbook()
    sheet = book.active
    sheet.title = "Incentive Report"
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "OUTLET-WISE MONTHLY SALES & INCENTIVE REPORT"
    sheet["A2"], sheet["B2"] = "AVG Profit Rate", profit_rate / 100
    sheet["D2"], sheet["E2"] = "Incentive Rate", incentive_rate / 100
    sheet["B2"].number_format = sheet["E2"].number_format = "0.00%"
    headers = ["Month", "Outlet", "Total Sales", "AVG Profit", "Total Incentive"]
    for col, header in enumerate(headers, 1):
        sheet.cell(4, col, header)
    for row_no, row in enumerate(source_rows, 5):
        sheet.cell(row_no, 1, row["month"])
        sheet.cell(row_no, 2, row["outlet"])
        sheet.cell(row_no, 3, round(float(row["total_sales"]), 2))
        sheet.cell(row_no, 4, f"=C{row_no}*$B$2")
        sheet.cell(row_no, 5, f"=D{row_no}*$E$2")
    end_row = 4 + len(source_rows)
    total_row = end_row + 1
    sheet.cell(total_row, 2, "GRAND TOTAL")
    for col in range(3, 6):
        letter = get_column_letter(col)
        sheet.cell(total_row, col, f"=SUM({letter}5:{letter}{end_row})")
    navy, blue, pale, white = "17365D", "4472C4", "D9EAF7", "FFFFFF"
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].font = Font(color=white, bold=True, size=15)
    sheet["A1"].alignment = Alignment(horizontal="center")
    for cell in sheet[4]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="left" if cell.column <= 2 else "right", vertical="center")
    for cell in sheet[total_row]:
        cell.fill = PatternFill("solid", fgColor=pale); cell.font = Font(bold=True)
    thin = Side(style="thin", color="B8C2CC")
    for row in sheet.iter_rows(min_row=4, max_row=total_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(bottom=thin)
    for col in range(3, 6):
        for row_no in range(5, total_row + 1):
            sheet.cell(row_no, col).number_format = '#,##0.00'
            sheet.cell(row_no, col).alignment = Alignment(horizontal="right", vertical="center")
    for row_no in range(5, total_row + 1):
        for col in range(1, 3):
            sheet.cell(row_no, col).alignment = Alignment(horizontal="left", vertical="center")
    for col, width in enumerate((15, 28, 18, 18, 20), 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:E{end_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 26
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1

    calculated_detail = calculate_incentive_rows(parsed_rows, profit_rate, incentive_rate)
    export_categories = ("HA", "HE", "MOB", "COM", "DC", "ACC")
    category_totals = defaultdict(lambda: {category: {"avg_profit": 0.0, "total_incentive": 0.0, "total_sales": 0.0} for category in export_categories})
    for row in calculated_detail:
        category = _incentive_report_category(row.get("category"))
        for field in ("total_sales", "avg_profit", "total_incentive"):
            category_totals[row["outlet"]][category][field] += row[field]

    category_sheet = book.create_sheet("Category-wise Report")
    category_sheet.merge_cells("A1:I1")
    category_sheet["A1"] = "CATEGORY-WISE AVG PROFIT & INCENTIVE"
    category_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    category_sheet["A1"].font = Font(color=white, bold=True, size=14)
    category_sheet["A1"].alignment = Alignment(horizontal="center")
    category_sheet.append([])
    category_sheet.append(["Outlet", "Metric", *export_categories, "TOTAL"])
    for cell in category_sheet[3]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="left" if cell.column <= 2 else "right", vertical="center")
    row_no = 4
    for outlet, categories in sorted(category_totals.items()):
        for metric_label, field in (("AVG Profit", "avg_profit"), ("Total Incentive", "total_incentive")):
            category_sheet.cell(row_no, 1, outlet)
            category_sheet.cell(row_no, 2, metric_label)
            for col_no, category in enumerate(export_categories, 3):
                category_sheet.cell(row_no, col_no, round(categories[category][field], 2))
            category_sheet.cell(row_no, 9, f"=SUM(C{row_no}:H{row_no})")
            for cell in category_sheet[row_no][:9]:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(horizontal="left" if cell.column <= 2 else "right", vertical="center")
                if cell.column >= 3: cell.number_format = '#,##0.00'
            row_no += 1
    category_sheet.freeze_panes = "C4"
    category_sheet.sheet_view.showGridLines = False
    category_sheet.column_dimensions["A"].width = 14
    category_sheet.column_dimensions["B"].width = 20
    for letter in "CDEFGHI": category_sheet.column_dimensions[letter].width = 17

    group_sheet = book.create_sheet("Group-wise Report")
    group_sheet.merge_cells("A1:E1")
    group_sheet["A1"] = "OUTLET GROUP-WISE SALES, AVG PROFIT & INCENTIVE"
    group_sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    group_sheet["A1"].font = Font(color=white, bold=True, size=14)
    group_sheet["A1"].alignment = Alignment(horizontal="center")
    group_sheet.append([])
    group_sheet.append(["Outlet", "Category Group", "Total Sales", "AVG Profit", "Total Incentive"])
    for cell in group_sheet[3]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="left" if cell.column <= 2 else "right", vertical="center")
    row_no = 4
    for outlet, categories in sorted(category_totals.items()):
        for label, category_names in _incentive_outlet_groups(outlet):
            group_sheet.cell(row_no, 1, outlet)
            group_sheet.cell(row_no, 2, label)
            for col_no, field in ((3, "total_sales"), (4, "avg_profit"), (5, "total_incentive")):
                group_sheet.cell(row_no, col_no, round(sum(categories[name][field] for name in category_names), 2))
                group_sheet.cell(row_no, col_no).number_format = '#,##0.00'
            for cell in group_sheet[row_no][:5]:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(horizontal="left" if cell.column <= 2 else "right", vertical="center")
            row_no += 1
    group_sheet.freeze_panes = "A4"
    group_sheet.sheet_view.showGridLines = False
    group_sheet.column_dimensions["A"].width = 14
    group_sheet.column_dimensions["B"].width = 30
    for letter in "CDE": group_sheet.column_dimensions[letter].width = 20

    output = BytesIO()
    book.save(output)
    filename = f"Outlet_Incentive_Report_{india_today().isoformat()}.xlsx"
    return Response(output.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/analytics/stage/{token}/commit")
def commit_staged_file(
    token: str,
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    entry = _staging_get_or_404(token)
    rows = entry["rows"]
    filename = entry["filename"]

    # Replace the previous dataset wholesale - this dashboard always reflects
    # only the most recently committed file.
    db.query(models.AnalyticsSalesRow).delete()
    db.query(models.AnalyticsUpload).delete()
    db.commit()

    sheet_names = {r.get("source_sheet") for r in rows if r.get("source_sheet")}
    dates = [r["sale_date"] for r in rows if r.get("sale_date")]

    upload_record = models.AnalyticsUpload(
        source_file=filename,
        uploaded_by=current_user.id,
        uploaded_by_username=current_user.username,
        row_count=len(rows),
        sheet_count=len(sheet_names) or 1,
        date_from=min(dates) if dates else None,
        date_to=max(dates) if dates else None,
    )
    db.add(upload_record)
    db.commit()
    db.refresh(upload_record)

    insert_mappings = [
        {
            "upload_id": upload_record.id,
            "sale_date": r["sale_date"],
            "item": r["item"],
            "division": r["division"],
            "brand": r.get("brand"),
            "qty": r.get("qty"),
            "sales_amt": r.get("sales_amt") or 0.0,
            "cost_amt": r.get("cost_amt") or 0.0,
            "profit_loss": r.get("profit_loss") or 0.0,
            "source_sheet": r.get("source_sheet"),
            "source_file": r.get("source_file"),
        }
        for r in rows
    ]
    db.bulk_insert_mappings(models.AnalyticsSalesRow, insert_mappings)
    db.commit()

    ANALYTICS_STAGING.pop(token, None)

    return {
        "message": "File processed successfully",
        "file_name": filename,
        "rows_loaded": len(rows),
        "sheets_read": len(sheet_names) or 1,
        "date_from": upload_record.date_from,
        "date_to": upload_record.date_to,
    }


@app.post("/api/analytics/upload")
def upload_analytics_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    filename = file.filename or "uploaded_file"
    raw_content = file.file.read()
    if not raw_content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    parsed_rows = parse_analytics_file(filename, raw_content)
    if not parsed_rows:
        raise HTTPException(
            status_code=400,
            detail="No readable sales rows found. Ensure the file has columns for Date, Item, Sales Amt, Cost Amt and Profit/Loss (Division and Qty are optional).",
        )

    # Replace the previous dataset wholesale - this dashboard always reflects
    # only the most recently uploaded file.
    db.query(models.AnalyticsSalesRow).delete()
    db.query(models.AnalyticsUpload).delete()
    db.commit()

    sheet_names = {r["source_sheet"] for r in parsed_rows if r["source_sheet"]}
    dates = [r["sale_date"] for r in parsed_rows]

    upload_record = models.AnalyticsUpload(
        source_file=filename,
        uploaded_by=current_user.id,
        uploaded_by_username=current_user.username,
        row_count=len(parsed_rows),
        sheet_count=len(sheet_names) or 1,
        date_from=min(dates) if dates else None,
        date_to=max(dates) if dates else None,
    )
    db.add(upload_record)
    db.commit()
    db.refresh(upload_record)

    for r in parsed_rows:
        r["upload_id"] = upload_record.id
    db.bulk_insert_mappings(models.AnalyticsSalesRow, parsed_rows)
    db.commit()

    return {
        "message": "File processed successfully",
        "file_name": filename,
        "rows_loaded": len(parsed_rows),
        "sheets_read": len(sheet_names) or 1,
        "date_from": upload_record.date_from,
        "date_to": upload_record.date_to,
    }


@app.get("/api/analytics/meta")
def analytics_meta(
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    upload_record = db.query(models.AnalyticsUpload).order_by(models.AnalyticsUpload.id.desc()).first()
    divisions = [
        row[0] for row in
        db.query(models.AnalyticsSalesRow.division).distinct().order_by(models.AnalyticsSalesRow.division).all()
        if row[0]
    ]
    if not upload_record:
        return {"has_data": False, "divisions": [], "last_upload": None, "can_upload": auth.has_admin_access(current_user)}

    return {
        "has_data": True,
        "divisions": divisions,
        "last_upload": {
            "file_name": upload_record.source_file,
            "uploaded_by": upload_record.uploaded_by_username,
            "uploaded_at": upload_record.created_date,
            "row_count": upload_record.row_count,
            "sheet_count": upload_record.sheet_count,
            "date_from": upload_record.date_from,
            "date_to": upload_record.date_to,
        },
        "can_upload": auth.has_admin_access(current_user),
    }


@app.get("/api/analytics/dashboard")
def analytics_dashboard(
    division: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    current_user: models.User = Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(models.AnalyticsSalesRow)
    if division and division.upper() != "ALL":
        query = query.filter(models.AnalyticsSalesRow.division == division.upper())
    if start_date:
        query = query.filter(models.AnalyticsSalesRow.sale_date >= start_date)
    if end_date:
        query = query.filter(models.AnalyticsSalesRow.sale_date <= end_date)

    rows = query.all()
    result = build_analytics_dashboard(rows)
    result["filters"] = {
        "division": division or "ALL",
        "start_date": start_date,
        "end_date": end_date,
    }
    return result


@app.delete("/api/analytics/clear")
def clear_analytics_data(
    current_user: models.User = Depends(auth.require_roles("Admin")),
    db: Session = Depends(get_db),
):
    deleted_count = db.query(models.AnalyticsSalesRow).delete()
    db.query(models.AnalyticsUpload).delete()
    db.commit()
    return {"message": "AI Analysis data cleared", "deleted": deleted_count}


if __name__ == "__main__":
    import uvicorn

    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    reload_enabled = os.getenv("UVICORN_RELOAD", "1").strip().lower() in {"1", "true", "yes", "on"}
    uvicorn.run("main:app", host=host, port=port, reload=reload_enabled)
